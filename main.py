#!/usr/bin/env python3

import aci_lib
import getpass
import ipaddress
import openpyxl
import pathlib
import re, os, sys
import urllib3
from openpyxl import load_workbook,workbook
from openpyxl.styles import Alignment, colors, Border, Font, NamedStyle, PatternFill, Protection, Side 
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path

# Global Variables
excel_workbook = None
home = Path.home()

Access_regex = re.compile('(add_apg|vlan_pool)')
Admin_regex = re.compile('(backup|radius|tacacs|realm|web_security)')
Contracts_regex = re.compile('(add_contract|add_subject|add_filter)')
DHCP_regex = re.compile('(add_vrf|ctx_common)')
Fabric_regex = re.compile('(bgp_(as|rr)|dns|dns_mgmt|domain|ntp|smartcallhome|snmp_(client|comm|info|trap|user)|syslog_(dg|rmt))')
Inventory_regex = re.compile('(apic_inb|inb_subnet|switch|vpc_pair)')
L3Out_regex = re.compile('(add_l3out|node_intf|node_prof)')
networks_regex = re.compile('(add_net)')
Sites_regex = re.compile('(site_id|grp_id)')
Tenant_regex = re.compile('(add_tenant)')
VRF_regex = re.compile('(add_vrf|ctx_common)')
VMM_regex = re.compile('(add_vrf|ctx_common)')

def process_Access(wb):
    # Evaluate Access Worksheet
    ws = wb['Access']
    aci_lib_ref = 'aci_lib.Access_Policies'
    func_regex = Access_regex
    read_worksheet(wb, ws, aci_lib_ref, func_regex)
    
    # Evaluate Inventory Worksheet
    ws = wb['Inventory']
    func_regex = Inventory_regex
    read_worksheet(wb, ws, aci_lib_ref, func_regex)

def process_Admin(wb):
    # Evaluate Admin Worksheet
    ws = wb['Admin']
    aci_lib_ref = 'aci_lib.Admin_Policies'
    func_regex = Admin_regex
    read_worksheet(wb, ws, aci_lib_ref, func_regex)
    
def process_Contracts(wb):
    # Evaluate Fabric Worksheet
    ws = wb['Contracts']
    aci_lib_ref = 'aci_lib.Contracts_Policies'
    func_regex = Contracts_regex
    read_worksheet(wb, ws, aci_lib_ref, func_regex)

def process_Fabric(wb):
    # Evaluate Fabric Worksheet
    ws = wb['Fabric']
    aci_lib_ref = 'aci_lib.Fabric_Policies'
    func_regex = Fabric_regex
    read_worksheet(wb, ws, aci_lib_ref, func_regex)

def process_L3Out(wb):
    # Evaluate Fabric Worksheet
    ws = wb['L3Out']
    aci_lib_ref = 'aci_lib.L3Out_Policies'
    func_regex = L3Out_regex
    read_worksheet(wb, ws, aci_lib_ref, func_regex)

def process_Sites(wb):
    # Evaluate Sites Worksheet
    ws = wb['Sites']
    aci_lib_ref = 'aci_lib.Site_Policies'
    func_regex = Sites_regex
    read_worksheet(wb, ws, aci_lib_ref, func_regex)

def process_Tenants(wb):
    # Evaluate Tenants Worksheet
    ws = wb['Tenants']
    aci_lib_ref = 'aci_lib.Tenant_Policies'
    func_regex = Tenant_regex
    read_worksheet(wb, ws, aci_lib_ref, func_regex)

    # Evaluate VRF Worksheet
    ws = wb['VRF']
    func_regex = VRF_regex
    read_worksheet(wb, ws, aci_lib_ref, func_regex)

    # Evaluate Network Segments Worksheet
    ws = wb['Networks']
    func_regex = networks_regex
    read_worksheet(wb, ws, aci_lib_ref, func_regex)

    # Evaluate DHCP Relay
    # ws = wb['DHCP Relay']
    # func_regex = DHCP_regex
    # read_worksheet(wb, ws, aci_lib_ref, func_regex)
    
def process_VMM(wb):
    # Evaluate Sites Worksheet
    ws = wb['VMM']
    aci_lib_ref = 'aci_lib.VMM_Policies'
    func_regex = VMM_regex
    read_worksheet(wb, ws, aci_lib_ref, func_regex)

def read_worksheet(wb, ws, aci_lib_ref, func_regex):
    rows = ws.max_row
    func_list = aci_lib.findKeys(ws, func_regex)
    class_init = '%s(ws)' % (aci_lib_ref)
    aci_lib.stdout_log(ws, None)
    for func in func_list:
        count = aci_lib.countKeys(ws, func)
        var_dict = aci_lib.findVars(ws, func, rows, count)
        for pos in var_dict:
            row_num = var_dict[pos]['row']
            del var_dict[pos]['row']
            for x in list(var_dict[pos].keys()):
                if var_dict[pos][x] == '':
                    del var_dict[pos][x]
            aci_lib.stdout_log(ws, row_num)
            eval("%s.%s(wb, ws, row_num, **var_dict[pos])" % (class_init, func))

def wb_update(wr_ws, status, i):
    # build green and red style sheets for excel
    bd1 = Side(style="thick", color="8EA9DB")
    bd2 = Side(style="medium", color="8EA9DB")
    wsh1 = NamedStyle(name="wsh1")
    wsh1.alignment = Alignment(horizontal="center", vertical="center", wrap_text="True")
    wsh1.border = Border(left=bd1, top=bd1, right=bd1, bottom=bd1)
    wsh1.font = Font(bold=True, size=15, color="FFFFFF")
    wsh2 = NamedStyle(name="wsh2")
    wsh2.alignment = Alignment(horizontal="center", vertical="center", wrap_text="True")
    wsh2.border = Border(left=bd2, top=bd2, right=bd2, bottom=bd2)
    wsh2.fill = PatternFill("solid", fgColor="305496")
    wsh2.font = Font(bold=True, size=15, color="FFFFFF")
    green_st = NamedStyle(name="ws_odd")
    green_st.alignment = Alignment(horizontal="center", vertical="center")
    green_st.border = Border(left=bd2, top=bd2, right=bd2, bottom=bd2)
    green_st.fill = PatternFill("solid", fgColor="D9E1F2")
    green_st.font = Font(bold=False, size=12, color="44546A")
    red_st = NamedStyle(name="ws_even")
    red_st.alignment = Alignment(horizontal="center", vertical="center")
    red_st.border = Border(left=bd2, top=bd2, right=bd2, bottom=bd2)
    red_st.font = Font(bold=False, size=12, color="44546A")
    yellow_st = NamedStyle(name="ws_even")
    yellow_st.alignment = Alignment(horizontal="center", vertical="center")
    yellow_st.border = Border(left=bd2, top=bd2, right=bd2, bottom=bd2)
    yellow_st.font = Font(bold=False, size=12, color="44546A")
    # green_st = xlwt.easyxf('pattern: pattern solid;')
    # green_st.pattern.pattern_fore_colour = 3
    # red_st = xlwt.easyxf('pattern: pattern solid;')
    # red_st.pattern.pattern_fore_colour = 2
    # yellow_st = xlwt.easyxf('pattern: pattern solid;')
    # yellow_st.pattern.pattern_fore_colour = 5
    # if stanzas to catch the status code from the request
    # and then input the appropriate information in the workbook
    # this then writes the changes to the doc
    if status == 200:
        wr_ws.write(i, 1, 'Success (200)', green_st)
    if status == 400:
        print("Error 400 - Bad Request - ABORT!")
        print("Probably have a bad URL or payload")
        wr_ws.write(i, 1, 'Bad Request (400)', red_st)
        pass
    if status == 401:
        print("Error 401 - Unauthorized - ABORT!")
        print("Probably have incorrect credentials")
        wr_ws.write(i, 1, 'Unauthorized (401)', red_st)
        pass
    if status == 403:
        print("Error 403 - Forbidden - ABORT!")
        print("Server refuses to handle your request")
        wr_ws.write(i, 1, 'Forbidden (403)', red_st)
        pass
    if status == 404:
        print("Error 404 - Not Found - ABORT!")
        print("Seems like you're trying to POST to a page that doesn't"
              " exist.")
        wr_ws.write(i, 1, 'Not Found (400)', red_st)
        pass
    if status == 666:
        print("Error - Something failed!")
        print("The POST failed, see stdout for the exception.")
        wr_ws.write(i, 1, 'Unkown Failure', yellow_st)
        pass
    if status == 667:
        print("Error - Invalid Input!")
        print("Invalid integer or other input.")
        wr_ws.write(i, 1, 'Unkown Failure', yellow_st)
        pass

def main():
    # Disable urllib3 warnings
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

    # Ask user for required Information: ACI_DEPLOY_FILE
    if sys.argv[1:]:
        if os.path.isfile(sys.argv[1]):
            excel_workbook = sys.argv[1]
        else:
            print('\nWorkbook not Found.  Please enter a valid /path/filename for the source you will be using.')
            while True:
                print('Please enter a valid /path/filename for the source you will be using.')
                excel_workbook = input('/Path/Filename: ')
                if os.path.isfile(excel_workbook):
                    print(f'\n-----------------------------------------------------------------------------\n')
                    print(f'   {excel_workbook} exists.  Will Now Check for API Variables...')
                    print(f'\n-----------------------------------------------------------------------------\n')
                    break
                else:
                    print('\nWorkbook not Found.  Please enter a valid /path/filename for the source you will be using.')
    else:
        while True:
            print('Please enter a valid /path/filename for the source you will be using.')
            excel_workbook = input('/Path/Filename: ')
            if os.path.isfile(excel_workbook):
                print(f'\n-----------------------------------------------------------------------------\n')
                print(f'   {excel_workbook} exists.  Will Now Check for API Variables...')
                print(f'\n-----------------------------------------------------------------------------\n')
                break
            else:
                print('\nWorkbook not Found.  Please enter a valid /path/filename for the source you will be using.')

    # Load Workbook
    wb = aci_lib.read_in(excel_workbook)
    
    # Run Proceedures for Worksheets in the Workbook
    process_Sites(wb)

    # Either Run All Remaining Proceedures or Just Specific based on sys.argv[2:]
    if sys.argv[2:]:
        if re.search('access', str(sys.argv[2:])):
            process_Access(wb)
        elif re.search('admin', str(sys.argv[2:])):
            process_Admin(wb)
        elif re.search('contracts', str(sys.argv[2:])):
            process_Contracts(wb)
        elif re.search('fabric', str(sys.argv[2:])):
            process_Fabric(wb)
        elif re.search('l3out', str(sys.argv[2:])):
            process_L3Out(wb)
        elif re.search('tenant', str(sys.argv[2:])):
            process_Tenants(wb)
        elif re.search('vmm', str(sys.argv[2:])):
            process_VMM(wb)
        else:
            process_Fabric(wb)
            process_Access(wb)
            process_Admin(wb)
            process_Contracts(wb)
            process_L3Out(wb)
            process_Tenants(wb)
    else:
        process_Fabric(wb)
        process_Access(wb)
        process_Admin(wb)
        process_Tenants(wb)

    # Save Workbook Changes
    # wb.save(excel_workbook)
    # wb.close()
    
if __name__ == '__main__':
    main()