#!/usr/bin/env python3

from openpyxl import load_workbook, workbook, Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, colors, Border, Font, NamedStyle, PatternFill, Protection, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from ordered_set import OrderedSet
from subprocess import check_output
import ast
import getpass
import ipaddress
import jinja2
import json
import os, re, sys
import pkg_resources
import requests
import time
import validating

# Global options for debugging
print_payload = True
print_response_always = True
print_response_on_fail = True

# Log levels 0 = None, 1 = Class only, 2 = Line
log_level = 2

# Global path to main Template directory
aci_template_path = pkg_resources.resource_filename('aci_lib', 'ACI/templates/')

# Exception Classes
class InsufficientArgs(Exception):
    pass

class ErrException(Exception):
    pass

class InvalidArg(Exception):
    pass

class LoginFailed(Exception):
    pass

# Terraform ACI Provider - Access Policies
# Class must be instantiated with Variables
class Access_Policies(object):
    def __init__(self, ws):
        self.templateLoader = jinja2.FileSystemLoader(
            searchpath=(aci_template_path + 'Access_Policies/'))
        self.templateEnv = jinja2.Environment(loader=self.templateLoader)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def aep_profile(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Infra_VLAN': ''}
        optional_args = {'Description': '',
                         'Physical_Domains': '',
                         'L3_Domains': '',
                         'VMM_Domains': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.values(row_num, ws, 'Infra_VLAN', templateVars['Infra_VLAN'], ['no', 'yes'])
            if not templateVars['Physical_Domains'] == None:
                templateVars['domains'] = 'yes'
                if re.search(r',', templateVars['Physical_Domains']):
                    x = templateVars['Physical_Domains'].split(',')
                    for domain in x:
                        validating.name_rule(row_num, ws, 'Physical_Domains', domain)
                else:
                    validating.name_rule(row_num, ws, 'Physical_Domains', templateVars['Physical_Domains'])
            if not templateVars['L3_Domains'] == None:
                templateVars['domains'] = 'yes'
                if re.search(r',', templateVars['L3_Domains']):
                    x = templateVars['L3_Domains'].split(',')
                    for domain in x:
                        validating.name_rule(row_num, ws, 'L3_Domains', domain)
                else:
                    validating.name_rule(row_num, ws, 'L3_Domains', templateVars['L3_Domains'])
            if not templateVars['VMM_Domains'] == None:
                templateVars['domains'] = 'yes'
                if re.search(r',', templateVars['VMM_Domains']):
                    x = templateVars['VMM_Domains'].split(',')
                    for domain in x:
                        validating.name_rule(row_num, ws, 'VMM_Domains', domain)
                else:
                    validating.name_rule(row_num, ws, 'VMM_Domains', templateVars['VMM_Domains'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        templateVars['phys_count'] = 0
        templateVars['l3_count'] = 0
        templateVars['vmm_count'] = 0
        if not templateVars['Physical_Domains'] == None:
            if re.search(r',', templateVars['Physical_Domains']):
                x = templateVars['Physical_Domains'].split(',')
                templateVars['Physical_Domains'] = []
                for domain in x:
                    templateVars['Physical_Domains'].append(domain)
                    templateVars['phys_count'] =+ 1
            else:
                templateVars['Physical_Domains'] = [templateVars['Physical_Domains']]
                templateVars['phys_count'] =+ 1
        if not templateVars['L3_Domains'] == None:
            if re.search(r',', templateVars['L3_Domains']):
                x = templateVars['L3_Domains'].split(',')
                templateVars['L3_Domains'] = []
                for domain in x:
                    templateVars['L3_Domains'].append(domain)
                    templateVars['l3_count'] =+ 1
            else:
                templateVars['L3_Domains'] = [templateVars['L3_Domains']]
                templateVars['l3_count'] =+ 1
        if not templateVars['VMM_Domains'] == None:
            if re.search(r',', templateVars['VMM_Domains']):
                x = templateVars['VMM_Domains'].split(',')
                templateVars['VMM_Domains'] = []
                for domain in x:
                    templateVars['VMM_Domains'].append(domain)
                    templateVars['vmm_count'] =+ 1
            else:
                templateVars['VMM_Domains'] = [templateVars['VMM_Domains']]
                templateVars['vmm_count'] =+ 1

        # Define the Template Source
        template_file = "global_aep.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'policies_global_aep_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def apic_inb(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Node_ID': '',
                         'Pod_ID': '',
                         'Inband_IPv4': '',
                         'Inband_GWv4': ''}
        optional_args = {'Inband_IPv6': '',
                         'Inband_GWv6': '',}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        templateVars['Device_Type'] = 'apic'

        # Initialize the Class
        aci_lib_ref = 'Access_Policies'
        class_init = '%s(ws)' % (aci_lib_ref)

        # Assign the APIC Inband Management IP's
        eval("%s.%s(wb, ws, row_num, **templateVars)" % (class_init, 'mgmt_inband'))

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def cdp(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Admin_State': ''}
        optional_args = {'Description': '',
                         'Alias': ''}
        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.values(row_num, ws, 'Admin_State', templateVars['Admin_State'], ['disabled', 'enabled'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)


        # Define the Template Source
        template_file = "policy_intf_cdp.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'policies_interface_cdp_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def fibre_channel(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Port_Mode': '',
                         'Trunk_Mode': '',
                         'Speed': '',
                         'Auto_Max_Speed': '',
                         'Fill_Pattern': '',
                         'Buffer_Credit': ''}
        optional_args = {'Description': '',
                         'Alias': ''}
        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.number_check(row_num, ws, 'Buffer_Credit', templateVars['Buffer_Credit'], 16, 64)
            validating.values(row_num, ws, 'Port_Mode', templateVars['Port_Mode'], ['f', 'np'])
            validating.values(row_num, ws, 'Trunk_Mode', templateVars['Trunk_Mode'], ['auto', 'trunk-off', 'trunk-on'])
            validating.values(row_num, ws, 'Speed', templateVars['Speed'], ['auto', '4G', '8G', '16G', '32G'])
            validating.values(row_num, ws, 'Auto_Max_Speed', templateVars['Auto_Max_Speed'], ['4G', '8G', '16G', '32G'])
            validating.values(row_num, ws, 'Fill_Pattern', templateVars['Fill_Pattern'], ['ARBFF', 'IDLE'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)


        # Define the Template Source
        template_file = "policy_intf_fc.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'policies_interface_fc_interface_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def intf_profile(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Switch_Role': '',
                         'Name': '',
                         'Dest_Folder': ''}
        optional_args = {'Description': ''}
        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.name_rule(row_num, ws, 'Dest_Folder', templateVars['Dest_Folder'])
            validating.values(row_num, ws, 'Switch_Role', templateVars['Switch_Role'], ['leaf', 'spine'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        if templateVars['Switch_Role'] == 'leaf':
            # Define the Template Source
            template_file = "leaf_interface_profile.template"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = '%s_interface_profile.tf' % (templateVars['Name'])
        elif templateVars['Switch_Role'] == 'spine':
            # Define the Template Source
            template_file = "spine_interface_profile.template"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = '%s_interface_profile.tf' % (templateVars['Name'])

        if not templateVars['Dest_Folder'] == None:
            dest_dir = '%s' % (templateVars['Dest_Folder'])
        else:
            dest_dir = 'Access'

        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def intf_selector(self, wb, ws, row_num, wr_file, Site_Group, Site_Name, Switch_Role, **kwargs):
        if not kwargs.get('Policy_Group') == None:
            # Dicts for required and optional args
            required_args = {'Pod_ID': '',
                            'Node_ID': '',
                            'Interface_Profile': '',
                            'Interface_Selector': '',
                            'Port': '',
                            'Policy_Group': '',
                            'Port_Type': ''}
            optional_args = {'LACP_Policy': '',
                            'Bundle_ID': '',
                            'Description': '',
                            'Switchport_Mode': '',
                            'Access_or_Native': '',
                            'Trunk_Allowed_VLANs': ''}

            if not kwargs.get('Port_Type') == None:
                if re.search('(port-channel|vpc)', kwargs.get('Port_Type')):

                    # Open the Access Worksheet and Find the Policy Group
                    ws_pg = wb['Access']
                    rowcount = ws_pg.max_row
                    row_count = ''
                    func = 'pg_access'
                    count = countKeys(ws_pg, func)
                    var_dict = findVars(ws_pg, func, rowcount, count)
                    for pos in var_dict:
                        if var_dict[pos].get('Name') == kwargs.get('Policy_Group'):
                            row_count = var_dict[pos]['row']
                            del var_dict[pos]['row']
                            kwargs = {**kwargs, **var_dict[pos]}
                            break

                    # Open the Access Policies Worksheet to get the Interface_Policy
                    ws_ac = wb['Access Policies']
                    rowcount = ws_ac.max_row
                    row_count = ''

                    # Get the Interface Policies from the Access Policies Tab
                    func = 'intf_polgrp'
                    count = countKeys(ws_ac, func)
                    row_count = ''
                    var_dict = findVars(ws_ac, func, rowcount, count)
                    for pos in var_dict:
                        if var_dict[pos].get('Policy_Name') == kwargs.get('Interface_Policy'):
                            row_count = var_dict[pos]['row']
                            del var_dict[pos]['row']
                            kwargs = {**kwargs, **var_dict[pos]}
                            break

                    # Validate inputs, return dict of template vars

                    kwargs['Site_Group'] = Site_Group
                    kwargs['Site_Name'] = Site_Name
                    kwargs['Switch_Role'] = Switch_Role

                    if kwargs.get('Port_Type') == 'vpc':
                        kwargs['Lag_Type'] = 'node'
                        ws_vpc = wb['Inventory']
                        for row in ws_vpc.rows:
                            if row[0].value == 'vpc_pair' and int(row[1].value) == int(Site_Group) and str(row[5].value) == str(kwargs.get('Node_ID')):
                                kwargs['VPC_Name'] = row[2].value
                                kwargs['Name'] = '%s_vpc%s' % (row[2].value, kwargs.get('Bundle_ID'))

                            elif row[0].value == 'vpc_pair' and str(row[1].value) == str(Site_Group) and str(row[6].value) == str(kwargs.get('Node_ID')):
                                kwargs['VPC_Name'] = row[2].value
                                kwargs['Name'] = '%s_vpc%s' % (row[2].value, kwargs.get('Bundle_ID'))
                    elif kwargs.get('Port_Type') == 'port-channel':
                        kwargs['Lag_Type'] = 'link'
                        kwargs['Name'] = '%s_pc%s' % (kwargs.get('Interface_Profile'), kwargs.get('Bundle_ID'))

                    # Create the Bundle Policy Group
                    aci_lib_ref = 'Access_Policies'
                    class_init = '%s(ws)' % (aci_lib_ref)
                    func = 'pg_bundle'
                    eval("%s.%s(wb, ws, row_num, **kwargs)" % (class_init, func))

            # Validate inputs, return dict of template vars
            templateVars = process_kwargs(required_args, optional_args, **kwargs)
            # leafx = Name
            templateVars['Site_Group'] = Site_Group
            templateVars['Site_Name'] = Site_Name
            templateVars['Switch_Role'] = Switch_Role

            xa = templateVars['Port'].split('/')
            xcount = len(xa)
            templateVars['Module'] = xa[0]
            templateVars['Port'] = xa[1]
            if Switch_Role == 'leaf':
                if not templateVars['Port_Type'] == 'breakout':
                    templateVars['Resource_Type'] = 'aci_leaf_access_port_policy_group'
                    if templateVars['Port_Type'] == 'access':
                        templateVars['PG_Type'] = 'accportgrp'
                elif templateVars['Port_Type'] == 'breakout':
                    templateVars['Resource_Type'] = 'aci_rest'
                    templateVars['PG_Type'] = 'brkoutportgrp'

                # If Policy Group Exists then Add Policy Group to templateVars for Port Selector
                if not templateVars['Policy_Group'] == None:
                    if templateVars['Port_Type'] == 'breakout':
                        templateVars['DN_Policy_Group'] = 'uni/infra/funcprof/brkoutportgrp-%s' % (templateVars['Policy_Group'])
                    elif templateVars['Port_Type'] == 'individual':
                        templateVars['DN_Policy_Group'] = 'uni/infra/funcprof/accportgrp-%s' % (templateVars['Policy_Group'])
                    elif templateVars['Port_Type'] == 'port-channel':
                        templateVars['DN_Policy_Group'] = 'uni/infra/funcprof/accbundle-%s' % (templateVars['Policy_Group'])
                    elif templateVars['Port_Type'] == 'vpc':
                        templateVars['DN_Policy_Group'] = 'uni/infra/funcprof/accbundle-%s' % (templateVars['Policy_Group'])

                # Define the Template Source
                template_file = "leaf_portselect.template"
                template = self.templateEnv.get_template(template_file)

                # Process the template and write to file
                payload = template.render(templateVars)
                wr_file.write(payload + '\n\n')

                # Define the Template Source
                if xcount == 3:
                    templateVars['Sub_Port'] = xa[2]
                    template_file = "leaf_portblock_sub.template"
                else:
                    template_file = "leaf_portblock.template"
                template = self.templateEnv.get_template(template_file)

                # Process the template and write to file
                payload = template.render(templateVars)
                wr_file.write(payload + '\n\n')

            elif Switch_Role == 'spine':
                # Define the Template Source
                template_file = "spine_portselect.template"
                template = self.templateEnv.get_template(template_file)

                # Process the template and write to file
                payload = template.render(templateVars)
                wr_file.write(payload + '\n\n')

                # Define the Template Source
                if not templateVars['Policy_Group'] == None:
                    if Switch_Role == 'spine':
                        template_file = "spine_pg_to_select.template"
                    template = self.templateEnv.get_template(template_file)

                    # Process the template and write to file
                    payload = template.render(templateVars)
                    wr_file.write(payload + '\n\n')

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def l2_interface(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'QinQ': '',
                         'Reflective_Relay': '',
                         'VLAN_Scope': ''}
        optional_args = {'Description': '',
                         'Alias': ''}
        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.values(row_num, ws, 'QinQ', templateVars['QinQ'], ['disabled', 'enabled'])
            validating.values(row_num, ws, 'Reflective_Relay', templateVars['Reflective_Relay'], ['disabled', 'enabled'])
            validating.values(row_num, ws, 'VLAN_Scope', templateVars['VLAN_Scope'], ['global', 'portlocal'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)


        # Define the Template Source
        template_file = "policy_intf_l2_interface.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'policies_interface_l2_interface_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def l3_domain(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'VLAN_Pool': ''}
        optional_args = { }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.name_rule(row_num, ws, 'VLAN_Pool', templateVars['VLAN_Pool'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)


        # Define the Template Source
        template_file = "domain_l3.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'policies_domain_l3_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def leaf_pg(self, wb, ws, row_num, **kwargs):
        # Open the Access Policies Worksheet
        ws_ac = wb['Access Policies']
        rowcount = ws_ac.max_row

        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'CDP_Policy': '',
                         'LLDP_Policy': '',
                         'topoctrlFwdScaleProfilePol': '',
                         'monInfraPol': '',
                         'Leaf_Profile_Policy': '',
                         'Policy_Name': '',
                         'l2NodeAuthPol': '',
                         'bfdIpv4InstPol': '',
                         'bfdIpv6InstPol': '',
                         'bfdMhIpv4InstPol': '',
                         'bfdMhIpv6InstPol': '',
                         'coppLeafProfile': '',
                         'iaclLeafProfile': '',
                         'equipmentFlashConfigPol': '',
                         'topoctrlFastLinkFailoverInstPol': '',
                         'fcFabricPol': '',
                         'fcInstPol': '',
                         'poeInstPol': '',
                         'stpInstPol': ''}
        optional_args = {'Description': '',
                         'netflowNodePol': ''}

        # Get the Application Profile Policies from the Network Policies Tab
        func = 'leaf_pg'
        count = countKeys(ws_ac, func)
        row_count = ''
        var_dict = findVars(ws_ac, func, rowcount, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('Leaf_Profile_Policy'):
                row_count = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}
                break

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.name_rule(row_num, ws, 'CDP_Policy', templateVars['CDP_Policy'])
            validating.name_rule(row_num, ws, 'LLDP_Policy', templateVars['LLDP_Policy'])
            validating.name_rule(row_num, ws, 'topoctrlFwdScaleProfilePol', templateVars['topoctrlFwdScaleProfilePol'])
            validating.name_rule(row_num, ws, 'monInfraPol', templateVars['monInfraPol'])
            validating.name_rule(row_count, ws_ac, 'l2NodeAuthPol', templateVars['l2NodeAuthPol'])
            validating.name_rule(row_count, ws_ac, 'bfdIpv4InstPol', templateVars['bfdIpv4InstPol'])
            validating.name_rule(row_count, ws_ac, 'bfdIpv6InstPol', templateVars['bfdIpv6InstPol'])
            validating.name_rule(row_count, ws_ac, 'bfdMhIpv4InstPol', templateVars['bfdMhIpv4InstPol'])
            validating.name_rule(row_count, ws_ac, 'bfdMhIpv6InstPol', templateVars['bfdMhIpv6InstPol'])
            validating.name_rule(row_count, ws_ac, 'coppLeafProfile', templateVars['coppLeafProfile'])
            validating.name_rule(row_count, ws_ac, 'iaclLeafProfile', templateVars['iaclLeafProfile'])
            validating.name_rule(row_count, ws_ac, 'equipmentFlashConfigPol', templateVars['equipmentFlashConfigPol'])
            validating.name_rule(row_count, ws_ac, 'topoctrlFastLinkFailoverInstPol', templateVars['topoctrlFastLinkFailoverInstPol'])
            validating.name_rule(row_count, ws_ac, 'fcFabricPol', templateVars['fcFabricPol'])
            validating.name_rule(row_count, ws_ac, 'fcInstPol', templateVars['fcInstPol'])
            validating.name_rule(row_count, ws_ac, 'poeInstPol', templateVars['poeInstPol'])
            validating.name_rule(row_count, ws_ac, 'stpInstPol', templateVars['stpInstPol'])
            if not templateVars['netflowNodePol'] == None:
                validating.name_rule(row_count, ws_ac, 'netflowNodePol', templateVars['netflowNodePol'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "leaf_policy_group.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'leaf_policy_group_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def link_level(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Auto_Neg': '',
                         'Speed': '',
                         'Port_Delay': '',
                         'Debounce_Interval': '',
                         'FEC_Mode': ''}
        optional_args = {'Description': '',
                         'Alias': ''}
        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.number_check(row_num, ws, 'Port_Delay', templateVars['Port_Delay'], 0, 10000)
            validating.number_check(row_num, ws, 'Debounce_Interval', templateVars['Debounce_Interval'], 0, 5000)
            validating.values(row_num, ws, 'Auto_Neg', templateVars['Auto_Neg'], ['off', 'on'])
            validating.values(row_num, ws, 'Speed', templateVars['Speed'], ['inherit', '100M', '1G', '10G', '25G', '40G', '50G', '100G', '200G', '400G'])
            validating.values(row_num, ws, 'FEC_Mode', templateVars['FEC_Mode'], ['inherit', 'auto-fec', 'cl74-fc-fec', 'cl91-rs-fec', 'cons16-rs-fec', 'disable-fec', 'ieee-rs-fec', 'kp-fec'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)


        # Define the Template Source
        template_file = "policy_intf_link_level.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'policies_interface_link_level_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def lldp(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Receive_State': '',
                         'Transmit_State': ''}
        optional_args = {'Description': '',
                         'Alias': ''}
        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.values(row_num, ws, 'Receive_State', templateVars['Receive_State'], ['disabled', 'enabled'])
            validating.values(row_num, ws, 'Transmit_State', templateVars['Transmit_State'], ['disabled', 'enabled'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)


        # Define the Template Source
        template_file = "policy_intf_lldp.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'policies_interface_lldp_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def mcp(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Admin_State': ''}
        optional_args = {'Description': '',
                         'Alias': ''}
        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.values(row_num, ws, 'Admin_State', templateVars['Admin_State'], ['disabled', 'enabled'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)


        # Define the Template Source
        template_file = "policy_intf_mcp.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'policies_interface_mcp_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def mgmt_inband(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Node_ID': '',
                         'Pod_ID': '',
                         'Inband_IPv4': '',
                         'Inband_GWv4': '',
                         'Device_Type': ''}
        optional_args = {'Inband_IPv6': '',
                         'Inband_GWv6': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.hostname(row_num, ws, 'Name', templateVars['Name'])
            validating.number_check(row_num, ws, 'Pod_ID', templateVars['Pod_ID'], 1, 12)
            validating.mgmt_network(row_num, ws, 'Inband_IPv4', templateVars['Inband_IPv4'], 'Inband_GWv4', templateVars['Inband_GWv4'])
            if templateVars['Device_Type'] == 'apic':
                validating.number_check(row_num, ws, 'Node_ID', templateVars['Node_ID'], 1, 7)
            else:
                validating.number_check(row_num, ws, 'Node_ID', templateVars['Node_ID'], 101, 4001)
            if not templateVars['Inband_IPv6'] == None:
                validating.mgmt_network(row_num, ws, 'Inband_IPv6', templateVars['Inband_IPv6'], 'Inband_GWv6', templateVars['Inband_GWv6'])
            else:
                templateVars['Inband_IPv6'] = '::'
                templateVars['Inband_GWv6'] = '::'
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "mgmt_inb.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = '%s_Mgmt_Inband.tf' % (templateVars['Name'])
        dest_dir = 'Tenant_mgmt'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def mgmt_oob(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Node_ID': '',
                         'Pod_ID': '',
                         'OOB_IPv4': '',
                         'OOB_GWv4': '',
                         'Device_Type': ''}
        optional_args = {'OOB_IPv6': '',
                         'OOB_GWv6': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.hostname(row_num, ws, 'Name', templateVars['Name'])
            validating.number_check(row_num, ws, 'Pod_ID', templateVars['Pod_ID'], 1, 12)
            validating.mgmt_network(row_num, ws, 'OOB_IPv4', templateVars['OOB_IPv4'], 'OOB_GWv4', templateVars['OOB_GWv4'])
            if templateVars['Device_Type'] == 'apic':
                validating.number_check(row_num, ws, 'Node_ID', templateVars['Node_ID'], 1, 7)
            else:
                validating.number_check(row_num, ws, 'Node_ID', templateVars['Node_ID'], 101, 4001)
            if not templateVars['OOB_IPv6'] == None:
                validating.mgmt_network(row_num, ws, 'OOB_IPv6', templateVars['OOB_IPv6'], 'OOB_GWv6', templateVars['OOB_GWv6'])
            else:
                templateVars['OOB_IPv6'] = '::'
                templateVars['OOB_GWv6'] = '::'
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "mgmt_oob.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = '%s_Mgmt_Out_of_Band.tf' % (templateVars['Name'])
        dest_dir = 'Tenant_mgmt'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def pg_access(self, wb, ws, row_num, **kwargs):
        # Open the Access Policies Worksheet
        ws_ac = wb['Access Policies']
        rowcount = ws_ac.max_row

        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'AEP_Policy': '',
                         'CDP_Policy': '',
                         'Link_Level': '',
                         'LLDP_Policy': '',
                         'MCP_Policy': '',
                         'STP_Policy': '',
                         'Interface_Policy': '',
                         'Policy_Name': ''}
        optional_args = {'Description': '',
                         'Alias': '',
                         'Pol_802_1x': '',
                         'poeIfPol': '',
                         'monFabricPol': '',
                         'dwdmIfPol': '',
                         'coppIfPol': '',
                         'qosDppPol_egress': '',
                         'qosDppPol_ingress': '',
                         'Fibre_Channel': '',
                         'L2_Interface': '',
                         'fabricLinkFlapPol': '',
                         'qosLlfcIfPol': '',
                         'macsecIfPol': '',
                         'netflowMonitorPol': '',
                         'Port_Security': '',
                         'qosPfcIfPol': '',
                         'qosSdIfPol': '',
                         'stormctrlIfPol': ''}

        # Get the Application Profile Policies from the Network Policies Tab
        func = 'intf_polgrp'
        count = countKeys(ws_ac, func)
        row_count = ''
        var_dict = findVars(ws_ac, func, rowcount, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('Interface_Policy'):
                row_count = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}
                break

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.name_rule(row_num, ws, 'AEP_Policy', templateVars['AEP_Policy'])
            validating.name_rule(row_num, ws, 'CDP_Policy', templateVars['CDP_Policy'])
            validating.name_rule(row_num, ws, 'Link_Level', templateVars['Link_Level'])
            validating.name_rule(row_num, ws, 'LLDP_Policy', templateVars['LLDP_Policy'])
            validating.name_rule(row_num, ws, 'MCP_Policy', templateVars['MCP_Policy'])
            validating.name_rule(row_num, ws, 'STP_Policy', templateVars['STP_Policy'])
            if not templateVars['Fibre_Channel'] == None:
                validating.name_rule(row_count, ws_ac, 'Fibre_Channel', templateVars['Fibre_Channel'])
            validating.name_rule(row_count, ws_ac, 'L2_Interface', templateVars['L2_Interface'])
            validating.name_rule(row_count, ws_ac, 'Port_Security', templateVars['Port_Security'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
            if not templateVars['Pol_802_1x'] == None:
                validating.name_rule(row_num, ws, 'Pol_802_1x', templateVars['Pol_802_1x'])
                templateVars['Pol_802_1x'] = 'uni/infra/portauthpol-%s' % (templateVars['Pol_802_1x'])
            if not templateVars['poeIfPol'] == None:
                validating.name_rule(row_num, ws, 'poeIfPol', templateVars['poeIfPol'])
                templateVars['poeIfPol'] = 'uni/infra/poeIfP-%s' % (templateVars['poeIfPol'])
            if not templateVars['monFabricPol'] == None:
                validating.name_rule(row_num, ws, 'monFabricPol', templateVars['monFabricPol'])
                templateVars['monFabricPol'] = 'uni/fabric/monfab-%s' % (templateVars['monFabricPol'])
            if not templateVars['dwdmIfPol'] == None:
                validating.name_rule(row_num, ws, 'dwdmIfPol', templateVars['dwdmIfPol'])
                templateVars['dwdmIfPol'] = 'uni/infra/dwdmifpol-%s' % (templateVars['dwdmIfPol'])
            if not templateVars['coppIfPol'] == None:
                validating.name_rule(row_count, ws_ac, 'coppIfPol', templateVars['coppIfPol'])
                templateVars['coppIfPol'] = 'uni/infra/coppifpol-%s' % (templateVars['coppIfPol'])
            if not templateVars['qosDppPol_egress'] == None:
                validating.name_rule(row_count, ws_ac, 'qosDppPol_egress', templateVars['qosDppPol_egress'])
                templateVars['qosDppPol_egress'] = 'uni/infra/qosdpppol-%s' % (templateVars['qosDppPol_egress'])
            if not templateVars['qosDppPol_ingress'] == None:
                validating.name_rule(row_count, ws_ac, 'qosDppPol_ingress', templateVars['qosDppPol_ingress'])
                templateVars['qosDppPol_ingress'] = 'uni/infra/qosdpppol-%s' % (templateVars['qosDppPol_ingress'])
            if not templateVars['fabricLinkFlapPol'] == None:
                validating.name_rule(row_count, ws_ac, 'fabricLinkFlapPol', templateVars['fabricLinkFlapPol'])
                templateVars['fabricLinkFlapPol'] = 'uni/infra/linkflappol-%s' % (templateVars['fabricLinkFlapPol'])
            if not templateVars['qosLlfcIfPol'] == None:
                validating.name_rule(row_count, ws_ac, 'qosLlfcIfPol', templateVars['qosLlfcIfPol'])
                templateVars['qosLlfcIfPol'] = 'uni/infra/llfc-%s' % (templateVars['qosLlfcIfPol'])
            if not templateVars['macsecIfPol'] == None:
                validating.name_rule(row_count, ws_ac, 'macsecIfPol', templateVars['macsecIfPol'])
                templateVars['macsecIfPol'] = 'uni/infra/macsecifpol-%s' % (templateVars['macsecIfPol'])
            if not templateVars['netflowMonitorPol'] == None:
                validating.name_rule(row_count, ws_ac, 'netflowMonitorPol', templateVars['netflowMonitorPol'])
                templateVars['netflowMonitorPol'] = 'uni/infra/poeIfP-%s' % (templateVars['netflowMonitorPol'])
            if not templateVars['qosPfcIfPol'] == None:
                validating.name_rule(row_count, ws_ac, 'qosPfcIfPol', templateVars['qosPfcIfPol'])
                templateVars['qosPfcIfPol'] = 'uni/infra/pfc-%s' % (templateVars['qosPfcIfPol'])
            if not templateVars['qosSdIfPol'] == None:
                validating.name_rule(row_count, ws_ac, 'qosSdIfPol', templateVars['qosSdIfPol'])
                templateVars['qosSdIfPol'] = 'uni/infra/qossdpol-%s' % (templateVars['qosSdIfPol'])
            if not templateVars['stormctrlIfPol'] == None:
                validating.name_rule(row_count, ws_ac, 'stormctrlIfPol', templateVars['stormctrlIfPol'])
                templateVars['stormctrlIfPol'] = 'uni/infra/stormctrlifp-%s' % (templateVars['stormctrlIfPol'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "leaf_intf_pg_access.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'leaf_interface_pg_access_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def pg_bundle(self, wb, ws, row_num, **kwargs):
        # Open the Access Policies Worksheet
        ws_ac = wb['Access Policies']
        rowcount = ws_ac.max_row

        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Lag_Type': '',
                         'AEP_Policy': '',
                         'CDP_Policy': '',
                         'Link_Level': '',
                         'LACP_Policy': '',
                         'LLDP_Policy': '',
                         'MCP_Policy': '',
                         'STP_Policy': '',
                         'Interface_Policy': '',
                         'Policy_Name': ''}
        optional_args = {'Description': '',
                         'Alias': '',
                         'monFabricPol': '',
                         'coppIfPol': '',
                         'qosDppPol_egress': '',
                         'qosDppPol_ingress': '',
                         'Fibre_Channel': '',
                         'L2_Interface': '',
                         'fabricLinkFlapPol': '',
                         'qosLlfcIfPol': '',
                         'macsecIfPol': '',
                         'netflowMonitorPol': '',
                         'Port_Security': '',
                         'qosPfcIfPol': '',
                         'qosSdIfPol': '',
                         'stormctrlIfPol': ''}

        # Get the Application Profile Policies from the Network Policies Tab
        func = 'intf_polgrp'
        count = countKeys(ws_ac, func)
        row_count = ''
        var_dict = findVars(ws_ac, func, rowcount, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('Interface_Policy'):
                row_count = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}
                break

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.name_rule(row_num, ws, 'AEP_Policy', templateVars['AEP_Policy'])
            validating.name_rule(row_num, ws, 'CDP_Policy', templateVars['CDP_Policy'])
            validating.name_rule(row_num, ws, 'Link_Level', templateVars['Link_Level'])
            validating.name_rule(row_num, ws, 'LACP_Policy', templateVars['LACP_Policy'])
            validating.name_rule(row_num, ws, 'LLDP_Policy', templateVars['LLDP_Policy'])
            validating.name_rule(row_num, ws, 'MCP_Policy', templateVars['MCP_Policy'])
            validating.name_rule(row_num, ws, 'STP_Policy', templateVars['STP_Policy'])
            if not templateVars['Fibre_Channel'] == None:
                validating.name_rule(row_count, ws_ac, 'Fibre_Channel', templateVars['Fibre_Channel'])
            validating.name_rule(row_count, ws_ac, 'L2_Interface', templateVars['L2_Interface'])
            validating.name_rule(row_count, ws_ac, 'Port_Security', templateVars['Port_Security'])
            validating.values(row_num, ws, 'Lag_Type', templateVars['Lag_Type'], ['link', 'node'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
            if not templateVars['monFabricPol'] == None:
                validating.name_rule(row_num, ws, 'monFabricPol', templateVars['monFabricPol'])
                templateVars['monFabricPol'] = 'uni/fabric/monfab-%s' % (templateVars['monFabricPol'])
            if not templateVars['coppIfPol'] == None:
                validating.name_rule(row_count, ws_ac, 'coppIfPol', templateVars['coppIfPol'])
                templateVars['coppIfPol'] = 'uni/infra/coppifpol-%s' % (templateVars['coppIfPol'])
            if not templateVars['qosDppPol_egress'] == None:
                validating.name_rule(row_count, ws_ac, 'qosDppPol_egress', templateVars['qosDppPol_egress'])
                templateVars['qosDppPol_egress'] = 'uni/infra/qosdpppol-%s' % (templateVars['qosDppPol_egress'])
            if not templateVars['qosDppPol_ingress'] == None:
                validating.name_rule(row_count, ws_ac, 'qosDppPol_ingress', templateVars['qosDppPol_ingress'])
                templateVars['qosDppPol_ingress'] = 'uni/infra/qosdpppol-%s' % (templateVars['qosDppPol_ingress'])
            if not templateVars['fabricLinkFlapPol'] == None:
                validating.name_rule(row_count, ws_ac, 'fabricLinkFlapPol', templateVars['fabricLinkFlapPol'])
                templateVars['fabricLinkFlapPol'] = 'uni/infra/linkflappol-%s' % (templateVars['fabricLinkFlapPol'])
            if not templateVars['qosLlfcIfPol'] == None:
                validating.name_rule(row_count, ws_ac, 'qosLlfcIfPol', templateVars['qosLlfcIfPol'])
                templateVars['qosLlfcIfPol'] = 'uni/infra/llfc-%s' % (templateVars['qosLlfcIfPol'])
            if not templateVars['macsecIfPol'] == None:
                validating.name_rule(row_count, ws_ac, 'macsecIfPol', templateVars['macsecIfPol'])
                templateVars['macsecIfPol'] = 'uni/infra/macsecifpol-%s' % (templateVars['macsecIfPol'])
            if not templateVars['netflowMonitorPol'] == None:
                validating.name_rule(row_count, ws_ac, 'netflowMonitorPol', templateVars['netflowMonitorPol'])
                templateVars['netflowMonitorPol'] = 'uni/infra/poeIfP-%s' % (templateVars['netflowMonitorPol'])
            if not templateVars['qosPfcIfPol'] == None:
                validating.name_rule(row_count, ws_ac, 'qosPfcIfPol', templateVars['qosPfcIfPol'])
                templateVars['qosPfcIfPol'] = 'uni/infra/pfc-%s' % (templateVars['qosPfcIfPol'])
            if not templateVars['qosSdIfPol'] == None:
                validating.name_rule(row_count, ws_ac, 'qosSdIfPol', templateVars['qosSdIfPol'])
                templateVars['qosSdIfPol'] = 'uni/infra/qossdpol-%s' % (templateVars['qosSdIfPol'])
            if not templateVars['stormctrlIfPol'] == None:
                validating.name_rule(row_count, ws_ac, 'stormctrlIfPol', templateVars['stormctrlIfPol'])
                templateVars['stormctrlIfPol'] = 'uni/infra/stormctrlifp-%s' % (templateVars['stormctrlIfPol'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "leaf_intf_pg_bundle.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'leaf_interface_pg_bundle_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def pg_breakout(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Breakout_Map': ''}
        optional_args = {'Description': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.values(row_num, ws, 'Breakout_Map', templateVars['Breakout_Map'], ['100g-2x', '100g-4x', '10g-4x', '25g-4x', '50g-8x'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "leaf_intf_pg_breakout.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'leaf_interface_pg_breakout_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def pg_spine(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'AEP_Policy': '',
                         'CDP_Policy': '',
                         'Link_Level': ''}
        optional_args = {'Description': '',
                         'Alias': '',
                         'fabricLinkFlapPol': '',
                         'macsecIfPol': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.name_rule(row_num, ws, 'AEP_Policy', templateVars['AEP_Policy'])
            validating.name_rule(row_num, ws, 'CDP_Policy', templateVars['CDP_Policy'])
            validating.name_rule(row_num, ws, 'Link_Level', templateVars['Link_Level'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
            if not templateVars['fabricLinkFlapPol'] == None:
                validating.name_rule(row_num, ws, 'fabricLinkFlapPol', templateVars['fabricLinkFlapPol'])
                templateVars['fabricLinkFlapPol'] = 'uni/infra/linkflappol-%s' % (templateVars['fabricLinkFlapPol'])
            if not templateVars['macsecIfPol'] == None:
                validating.name_rule(row_num, ws, 'macsecIfPol', templateVars['macsecIfPol'])
                templateVars['macsecIfPol'] = 'uni/infra/macsecifpol-%s' % (templateVars['macsecIfPol'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "spine_intf_pg_access.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'spine_interface_pg_access_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def phys_dom(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'VLAN_Pool': ''}
        optional_args = { }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.name_rule(row_num, ws, 'VLAN_Pool', templateVars['VLAN_Pool'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)


        # Define the Template Source
        template_file = "domain_phys.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'policies_domain_phys_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def port_channel(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Mode': '',
                         'Min_Links': '',
                         'Max_Links': '',
                         'Fast_Select': '',
                         'Graceful': '',
                         'Load_Defer': '',
                         'Suspend_Individual': '',
                         'Symmetric_Hash': ''}
        optional_args = {'Description': '',
                         'Alias': ''}
        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.number_check(row_num, ws, 'Min_Links', templateVars['Min_Links'], 1, 16)
            validating.number_check(row_num, ws, 'Max_Links', templateVars['Max_Links'], 1, 16)
            validating.values(row_num, ws, 'Mode', templateVars['Mode'], ['active', 'explicit-failover', 'mac-pin', 'mac-pin-nicload', 'off', 'passive'])
            validating.values(row_num, ws, 'Fast_Select', templateVars['Fast_Select'], ['no', 'yes'])
            validating.values(row_num, ws, 'Graceful', templateVars['Graceful'], ['no', 'yes'])
            validating.values(row_num, ws, 'Load_Defer', templateVars['Load_Defer'], ['no', 'yes'])
            validating.values(row_num, ws, 'Suspend_Individual', templateVars['Suspend_Individual'], ['no', 'yes'])
            validating.values(row_num, ws, 'Symmetric_Hash', templateVars['Symmetric_Hash'], ['no', 'yes'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)


        # Create ctrl templateVars
        ctrl = ''
        if templateVars['Fast_Select'] == 'yes':
            ctrl = ctrl + '"fast-sel-hot-stdby"'
        if templateVars['Graceful'] == 'yes':
            ctrl = ctrl + ', ' + '"graceful-conv"'
        if templateVars['Load_Defer'] == 'yes':
            ctrl = ctrl + ', ' + '"load-defer"'
        if templateVars['Suspend_Individual'] == 'yes':
            ctrl = ctrl + ', ' + '"susp-individual"'
        if templateVars['Symmetric_Hash'] == 'yes':
            ctrl = ctrl + ', ' + '"symmetric-hash"'

        templateVars['ctrl'] = '[%s]' % (ctrl)

        # Define the Template Source
        template_file = "policy_intf_lacp.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'policies_interface_port_channel_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def port_cnvt(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Node_ID': '',
                         'Port': ''}
        optional_args = { }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.hostname(row_num, ws, 'Name', templateVars['Name'])
            validating.number_check(row_num, ws, 'Node_ID', templateVars['Node_ID'], 101, 4001)
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Create Port Name Var
        zz = templateVars['Port'].split('/')
        templateVars['Port_Name'] = '%s_%s' % (zz[0], zz[1])

        # Define the Template Source
        template_file = "downlink.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'downlink_convert_%s.tf' % (templateVars['Port_Name'])
        dest_dir = 'Access/%s' % (templateVars['Name'])
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def port_security(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Timeout': '',
                         'Maximum_Endpoints': ''}
        optional_args = {'Description': '',
                         'Alias': ''}
        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.number_check(row_num, ws, 'Timeout', templateVars['Timeout'], 60, 3600)
            validating.number_check(row_num, ws, 'Maximum_Endpoints', templateVars['Maximum_Endpoints'], 0, 12000)
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)


        # Define the Template Source
        template_file = "policy_intf_port_sec.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'policies_interface_port_security_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def spine_pg(self, wb, ws, row_num, **kwargs):
        # Open the Access Policies Worksheet
        ws_ac = wb['Access Policies']
        rowcount = ws_ac.max_row

        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'AEP_Policy': '',
                         'CDP_Policy': '',
                         'Link_Level': '',
                         'LLDP_Policy': '',
                         'MCP_Policy': '',
                         'STP_Policy': '',
                         'Interface_Policy': '',
                         'Policy_Name': ''}
        optional_args = {'Description': '',
                         'Alias': '',
                         '802.1x': '',
                         'poeIfPol': '',
                         'monFabricPol': '',
                         'dwdmIfPol': '',
                         'coppIfPol': '',
                         'qosDppPol_egress': '',
                         'qosDppPol_ingress': '',
                         'Fibre_Channel': '',
                         'L2_Interface': '',
                         'fabricLinkFlapPol': '',
                         'qosLlfcIfPol': '',
                         'macsecIfPol': '',
                         'netflowMonitorPol': '',
                         'Port_Security': '',
                         'qosPfcIfPol': '',
                         'qosSdIfPol': '',
                         'stormctrlIfPol': ''}

        # Get the Application Profile Policies from the Network Policies Tab
        func = 'intf_polgrp'
        count = countKeys(ws_ac, func)
        row_count = ''
        var_dict = findVars(ws_ac, func, rowcount, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('Interface_Policy'):
                row_count = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}
                break

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.name_rule(row_num, ws, 'AEP_Policy', templateVars['AEP_Policy'])
            validating.name_rule(row_num, ws, 'CDP_Policy', templateVars['CDP_Policy'])
            validating.name_rule(row_num, ws, 'Link_Level', templateVars['Link_Level'])
            validating.name_rule(row_num, ws, 'LLDP_Policy', templateVars['LLDP_Policy'])
            validating.name_rule(row_num, ws, 'MCP_Policy', templateVars['MCP_Policy'])
            validating.name_rule(row_num, ws, 'STP_Policy', templateVars['STP_Policy'])
            validating.name_rule(row_count, ws_ac, 'Fibre_Channel', templateVars['Fibre_Channel'])
            validating.name_rule(row_count, ws_ac, 'L2_Interface', templateVars['L2_Interface'])
            validating.name_rule(row_count, ws_ac, 'Port_Security', templateVars['Port_Security'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
            if not templateVars['poeIfPol'] == None:
                validating.name_rule(row_num, ws, 'poeIfPol', templateVars['poeIfPol'])
                templateVars['poeIfPol'] = 'uni/infra/poeIfP-%s' % (templateVars['poeIfPol'])
            if not templateVars['monFabricPol'] == None:
                validating.name_rule(row_num, ws, 'monFabricPol', templateVars['monFabricPol'])
                templateVars['monFabricPol'] = 'uni/fabric/monfab-%s' % (templateVars['monFabricPol'])
            if not templateVars['dwdmIfPol'] == None:
                validating.name_rule(row_num, ws, 'dwdmIfPol', templateVars['dwdmIfPol'])
                templateVars['dwdmIfPol'] = 'uni/infra/dwdmifpol-%s' % (templateVars['dwdmIfPol'])
            if not templateVars['coppIfPol'] == None:
                validating.name_rule(row_count, ws_ac, 'coppIfPol', templateVars['coppIfPol'])
                templateVars['coppIfPol'] = 'uni/infra/coppifpol-%s' % (templateVars['coppIfPol'])
            if not templateVars['qosDppPol_egress'] == None:
                validating.name_rule(row_count, ws_ac, 'qosDppPol_egress', templateVars['qosDppPol_egress'])
                templateVars['qosDppPol_egress'] = 'uni/infra/qosdpppol-%s' % (templateVars['qosDppPol_egress'])
            if not templateVars['qosDppPol_ingress'] == None:
                validating.name_rule(row_count, ws_ac, 'qosDppPol_ingress', templateVars['qosDppPol_ingress'])
                templateVars['qosDppPol_ingress'] = 'uni/infra/qosdpppol-%s' % (templateVars['qosDppPol_ingress'])
            if not templateVars['fabricLinkFlapPol'] == None:
                validating.name_rule(row_count, ws_ac, 'fabricLinkFlapPol', templateVars['fabricLinkFlapPol'])
                templateVars['fabricLinkFlapPol'] = 'uni/infra/linkflappol-%s' % (templateVars['fabricLinkFlapPol'])
            if not templateVars['qosLlfcIfPol'] == None:
                validating.name_rule(row_count, ws_ac, 'qosLlfcIfPol', templateVars['qosLlfcIfPol'])
                templateVars['qosLlfcIfPol'] = 'uni/infra/llfc-%s' % (templateVars['qosLlfcIfPol'])
            if not templateVars['macsecIfPol'] == None:
                validating.name_rule(row_count, ws_ac, 'macsecIfPol', templateVars['macsecIfPol'])
                templateVars['macsecIfPol'] = 'uni/infra/macsecifpol-%s' % (templateVars['macsecIfPol'])
            if not templateVars['netflowMonitorPol'] == None:
                validating.name_rule(row_count, ws_ac, 'netflowMonitorPol', templateVars['netflowMonitorPol'])
                templateVars['netflowMonitorPol'] = 'uni/infra/poeIfP-%s' % (templateVars['netflowMonitorPol'])
            if not templateVars['qosPfcIfPol'] == None:
                validating.name_rule(row_count, ws_ac, 'qosPfcIfPol', templateVars['qosPfcIfPol'])
                templateVars['qosPfcIfPol'] = 'uni/infra/pfc-%s' % (templateVars['qosPfcIfPol'])
            if not templateVars['qosSdIfPol'] == None:
                validating.name_rule(row_count, ws_ac, 'qosSdIfPol', templateVars['qosSdIfPol'])
                templateVars['qosSdIfPol'] = 'uni/infra/qossdpol-%s' % (templateVars['qosSdIfPol'])
            if not templateVars['stormctrlIfPol'] == None:
                validating.name_rule(row_count, ws_ac, 'stormctrlIfPol', templateVars['stormctrlIfPol'])
                templateVars['stormctrlIfPol'] = 'uni/infra/stormctrlifp-%s' % (templateVars['stormctrlIfPol'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "leaf_intf_pg_access.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'leaf_interface_pg_access_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def stp(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Filter': '',
                         'Guard': ''}
        optional_args = {'Description': '',
                         'Alias': ''}
        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.values(row_num, ws, 'Filter', templateVars['Filter'], ['disabled', 'enabled'])
            validating.values(row_num, ws, 'Guard', templateVars['Guard'], ['disabled', 'enabled'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)


        # Define the Template Source
        template_file = "policy_intf_stp.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'policies_interface_stp_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def switch(self, wb, ws, row_num, **kwargs):
        # Initialize the Class
        aci_lib_ref = 'Access_Policies'
        class_init = '%s(ws)' % (aci_lib_ref)

        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Serial': '',
                         'Node_ID': '',
                         'Name': '',
                         'Profiles': '',
                         'Node_Type': '',
                         'Pod_ID': '',
                         'Switch_Role': '',
                         'Switch_Type': '',
                         'Is_Virtual': '',
                         'Tier-2': '',
                         'Fabric_ID': '',
                         'Inband_IPv4': '',
                         'Inband_GWv4': ''}
        optional_args = {'Policy_Group': '',
                         'Remote_ID': '',
                         'OOB_IPv4': '',
                         'OOB_GWv4': '',
                         'Inband_IPv6': '',
                         'Inband_GWv6': '',
                         'OOB_IPv6': '',
                         'OOB_GWv6': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        # Use Switch_Type to Determine the Number of ports on the switch
        modules,port_count = query_switch_model(row_num, templateVars['Switch_Type'])

        templateVars['Device_Type'] = 'switch'
        if not (templateVars['Inband_IPv4'] == None or templateVars['Inband_IPv6'] == None):
            # Assign the Switch Inband Management IP's
            eval("%s.%s(wb, ws, row_num, **templateVars)" % (class_init, 'mgmt_inband'))

        if not (templateVars['OOB_IPv4'] == None or templateVars['OOB_IPv6'] == None):
            # Assign the Switch Out-of-Band Management IP's
            eval("%s.%s(wb, ws, row_num, **templateVars)" % (class_init, 'mgmt_oob'))

        try:
            # Validate Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.hostname(row_num, ws, 'Name', templateVars['Name'])
            validating.modules(row_num, templateVars['Name'], templateVars['Switch_Role'], modules)
            validating.number_check(row_num, ws, 'Node_ID', templateVars['Node_ID'], 101, 4001)
            validating.number_check(row_num, ws, 'Pod_ID', templateVars['Pod_ID'], 1, 12)
            validating.number_check(row_num, ws, 'Fabric_ID', templateVars['Fabric_ID'], 1, 12)
            validating.port_count(row_num, templateVars['Name'], templateVars['Switch_Role'], port_count)
            validating.values(row_num, ws, 'Profiles', templateVars['Profiles'], ['no', 'yes'])
            validating.values(row_num, ws, 'Node_Type', templateVars['Node_Type'], ['remote-leaf-wan', 'unspecified'])
            validating.values(row_num, ws, 'Switch_Role', templateVars['Switch_Role'], ['leaf', 'spine'])
            validating.values(row_num, ws, 'Is_Virtual', templateVars['Is_Virtual'], ['no', 'yes'])
            validating.values(row_num, ws, 'Tier-2', templateVars['Tier-2'], ['no', 'yes'])
            if not templateVars['Profiles'] == 'yes':
                validating.name_rule(row_num, 'Policy_Group', templateVars['Policy_Group'])
            if not templateVars['Remote_ID'] == None:
                validating.number_check(row_num, ws, 'Remote_ID', templateVars['Remote_ID'], 1, 255)
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        if re.search('Grp_[A-F]', templateVars['Site_Group']):
            print(f"\n-----------------------------------------------------------------------------\n")
            print(f"   Error on Worksheet {ws.title}, Row {row_num} Site_Group, value {templateVars['Site_Group']}.")
            print(f"   A Leaf can only be assigned to one Site.  Exiting....")
            print(f"\n-----------------------------------------------------------------------------\n")
            exit()
        elif re.search(r'\d+', templateVars['Site_Group']):
            Site_ID = 'Site_ID_%s' % (templateVars['Site_Group'])
            site_dict = ast.literal_eval(os.environ[Site_ID])

            # Create templateVars for Site_Name and APIC_URL
            templateVars['Site_Name'] = site_dict.get('Site_Name')
            templateVars['APIC_URL'] = site_dict.get('APIC_URL')
            templateVars['APIC_Version'] = site_dict.get('APIC_Version')
            templateVars['APIC_Auth_Type'] = site_dict.get('APIC_Auth_Type')
            templateVars['Provider_Version'] = site_dict.get('Provider_Version')
            templateVars['Run_Location'] = site_dict.get('Run_Location')
            templateVars['State_Location'] = site_dict.get('State_Location')
        else:
            print(f"\n-----------------------------------------------------------------------------\n")
            print(f"   Error on Worksheet {ws.title}, Row {row_num} Site_Group, value {templateVars['Site_Group']}.")
            print(f"   Unable to Determine if this is a Single or Group of Site(s).  Exiting....")
            print(f"\n-----------------------------------------------------------------------------\n")
            exit()

        # Copy the data file for the Inband EPG Into the Switch Directory
        src_dir = './ACI/templates'
        dest_dir = '%s' % (templateVars['Name'])

        # Write the main.tf to the Appropriate Directories
        self.templateLoader = jinja2.FileSystemLoader(searchpath=('ACI/templates/'))
        self.templateEnv = jinja2.Environment(loader=self.templateLoader)
        template_file = "main.template"
        template = self.templateEnv.get_template(template_file)
        create_tf_file('w', dest_dir, template_file, template, **templateVars)

        # Write the variables.tf to the Appropriate Directories
        template_file = "variables.template"
        template = self.templateEnv.get_template(template_file)
        create_tf_file('w', dest_dir, template_file, template, **templateVars)

        self.templateLoader = jinja2.FileSystemLoader(searchpath=(aci_template_path + 'Access_Policies/'))
        self.templateEnv = jinja2.Environment(loader=self.templateLoader)
        excel_wkbook = '%s_intf_selectors.xlsx' % (templateVars['Site_Name'])

        wb_sw = load_workbook(excel_wkbook)

        # Check if there is a Worksheet for the Switch Already
        if not templateVars['Name'] in wb_sw.sheetnames:
            ws_sw = wb_sw.create_sheet(title = templateVars['Name'])
            ws_sw = wb_sw[templateVars['Name']]
            ws_sw.column_dimensions['A'].width = 15
            ws_sw.column_dimensions['B'].width = 10
            ws_sw.column_dimensions['c'].width = 10
            ws_sw.column_dimensions['d'].width = 20
            ws_sw.column_dimensions['E'].width = 20
            ws_sw.column_dimensions['F'].width = 10
            ws_sw.column_dimensions['G'].width = 20
            ws_sw.column_dimensions['H'].width = 20
            ws_sw.column_dimensions['I'].width = 20
            ws_sw.column_dimensions['J'].width = 15
            ws_sw.column_dimensions['K'].width = 30
            ws_sw.column_dimensions['L'].width = 20
            ws_sw.column_dimensions['M'].width = 20
            ws_sw.column_dimensions['N'].width = 30
            dv1 = DataValidation(type="list", formula1='"intf_selector"', allow_blank=True)
            dv2 = DataValidation(type="list", formula1='"access,breakout,port-channel,vpc"', allow_blank=True)
            ws_sw.add_data_validation(dv1)
            ws_sw.add_data_validation(dv2)
            ws_header = '%s Interface Selectors' % (templateVars['Name'])
            data = [ws_header]
            ws_sw.append(data)
            ws_sw.merge_cells('A1:N1')
            for cell in ws_sw['1:1']:
                cell.style = 'Heading 1'
            data = ['','Notes: Breakout Policy Group Names are 2x100g_pg, 4x10g_pg, 4x25g_pg, 4x100g_pg, 8x50g_pg.']
            ws_sw.append(data)
            ws_sw.merge_cells('B2:N2')
            for cell in ws_sw['2:2']:
                cell.style = 'Heading 2'
            data = ['Type','Pod_ID','Node_ID','Interface_Profile','Interface_Selector','Port','Policy_Group','Port_Type','LACP_Policy','Bundle_ID','Description','Switchport_Mode','Access_or_Native','Trunk_Allowed_VLANs']
            ws_sw.append(data)
            for cell in ws_sw['3:3']:
                cell.style = 'Heading 3'

            ws_sw_row_count = 4
            templateVars['dv1'] = dv1
            templateVars['dv2'] = dv2
            templateVars['port_count'] = port_count
            sw_type = str(templateVars['Switch_Type'])
            sw_name = str(templateVars['Name'])
            if re.search('^(93[0-9][0-9])', sw_type):
                for module in range(1, 2):
                    templateVars['module'] = module
                    ws_sw_row_count = create_selector(ws_sw, ws_sw_row_count, **templateVars)
            if re.search('^(9396|95[0-1][4-8])', sw_type):
                row_count = 1
                for row in ws.rows:
                    if re.search('9396', sw_type):
                        start, end = 2, 2
                    else:
                        start, end = 1, int(modules)
                    if str(row[0].value) == sw_type and str(row[2].value) == sw_name:
                        for module in range(start, end + 2):
                            templateVars['module'] = module
                            module_type = row[module + 2].value
                            if module_type == None:
                                module_type = 'none'
                            elif re.search('(X97|M(4|6|12)P)', module_type):
                                templateVars['port_count'] = query_module_type(row_count, module_type)
                                ws_sw_row_count = create_selector(ws_sw, ws_sw_row_count, **templateVars)
                        row_count += 1
                        break
            wb_sw.save(excel_wkbook)
        else:
            ws_sw = wb_sw[templateVars['Name']]

        # Define the Template Source
        template_file = "inventory.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = '%s.tf' % (templateVars['Name'])
        dest_dir = '%s' % (templateVars['Name'])
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        if templateVars['Profiles'] == 'yes':
            templateVars['Description'] = None
            templateVars['Dest_Folder'] = templateVars['Name']
            eval("%s.%s(wb, ws, row_num, **templateVars)" % (class_init, 'intf_profile'))

            templateVars['Selector_Name'] = templateVars['Name']
            templateVars['Association_Type'] = 'range'
            templateVars['Nodeblk_Name'] = 'blk%s-%s' % (templateVars['Node_ID'], templateVars['Node_ID'])
            templateVars['Node_ID_From'] = templateVars['Node_ID']
            templateVars['Node_ID_To'] = templateVars['Node_ID']
            templateVars['Interface_Profile'] = templateVars['Name']
            eval("%s.%s(wb, ws, row_num, **templateVars)" % (class_init, 'sw_profile'))

            sw_intf_profile = './ACI/%s/%s/%s_interface_profile.tf' % (templateVars['Site_Name'], templateVars['Name'], templateVars['Name'])
            wr_file = open(sw_intf_profile, 'a+')
            aci_lib_ref_sw = 'Access_Policies'
            rows_sw = ws_sw.max_row
            func_regex = re.compile('^intf_selector$')
            func_list_sw = findKeys(ws_sw, func_regex)
            class_init_sw = '%s(ws_sw)' % (aci_lib_ref_sw)
            stdout_log(ws_sw, None)
            for func_sw in func_list_sw:
                count_sw = countKeys(ws_sw, func_sw)
                var_dict_sw = findVars(ws_sw, func_sw, rows_sw, count_sw)
                for pos_sw in var_dict_sw:
                    row_num_sw = var_dict_sw[pos_sw]['row']
                    del var_dict_sw[pos_sw]['row']
                    for x_sw in list(var_dict_sw[pos_sw].keys()):
                        if var_dict_sw[pos_sw][x_sw] == '':
                            del var_dict_sw[pos_sw][x_sw]
                    stdout_log(ws_sw, row_num_sw)
                    site_group = templateVars['Site_Group']
                    sw_role = templateVars['Switch_Role']
                    site_name = templateVars['Site_Name']
                    eval("%s.%s(wb, ws_sw, row_num_sw, wr_file, '%s', '%s', '%s', **var_dict_sw[pos_sw])" % (class_init_sw, func_sw, site_group, site_name, sw_role))
            wr_file.close()
            ws_wr = wb_sw.get_sheet_names()
            for sheetName in ws_wr:
                if sheetName in ['Sites']:
                    sheetToDelete = wb_sw.get_sheet_by_name(sheetName)
                    wb_sw.remove_sheet(sheetToDelete)
                    wb_sw.save(excel_wkbook)
            wb_sw.close()

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def sw_profile(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Switch_Role': '',
                         'Name': '',
                         'Selector_Name': '',
                         'Association_Type': '',
                         'Nodeblk_Name': '',
                         'Node_ID_From': '',
                         'Node_ID_To': '',
                         'Policy_Group': '',
                         'Interface_Profile': '',
                         'Dest_Folder': ''}
        optional_args = {'Description': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.name_rule(row_num, ws, 'Selector_Name', templateVars['Selector_Name'])
            validating.name_rule(row_num, ws, 'Nodeblk_Name', templateVars['Nodeblk_Name'])
            validating.name_rule(row_num, ws, 'Policy_Group', templateVars['Policy_Group'])
            validating.name_rule(row_num, ws, 'Interface_Profile', templateVars['Interface_Profile'])
            validating.name_rule(row_num, ws, 'Dest_Folder', templateVars['Dest_Folder'])
            validating.number_check(row_num, ws, 'Node_ID_From', templateVars['Node_ID_From'], 101, 4001)
            validating.number_check(row_num, ws, 'Node_ID_To', templateVars['Node_ID_To'], 101, 4001)
            validating.values(row_num, ws, 'Switch_Role', templateVars['Switch_Role'], ['leaf', 'spine'])
            validating.values(row_num, ws, 'Association_Type', templateVars['Association_Type'], ['ALL', 'range', 'ALL_IN_POD'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        if templateVars['Switch_Role'] == 'leaf':
            template_file = "leaf_profile.template"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = '%s_leaf_profile.tf' % (templateVars['Name'])
        else:
            template_file = "spine_profile.template"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = '%s_spine_profile.tf' % (templateVars['Name'])

        if not templateVars['Dest_Folder'] == None:
            dest_dir = '%s' % (templateVars['Dest_Folder'])
        else:
            dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def vlan_pool(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Allocation_Mode': '',
                         'VLAN_Grp1': '',
                         'VGRP1_Allocation': ''}
        optional_args = {'Description': '',
                         'Alias': '',
                         'VLAN_Grp1': '',
                         'VGRP1_Allocation': '',
                         'VLAN_Grp2': '',
                         'VGRP2_Allocation': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.values(row_num, ws, 'Allocation_Mode', templateVars['Allocation_Mode'], ['dynamic', 'static'])
            validating.values(row_num, ws, 'VGRP1_Allocation', templateVars['VGRP1_Allocation'], ['dynamic', 'static'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
            if not templateVars['VGRP2_Allocation'] == None:
                validating.values(row_num, ws, 'VGRP2_Allocation', templateVars['VGRP2_Allocation'], ['dynamic', 'static'])
            validating.vlans(row_num, ws, 'VLAN_Grp1', templateVars['VLAN_Grp1'])
            if not templateVars['VLAN_Grp2'] == None:
                validating.vlans(row_num, ws, 'VLAN_Grp2', templateVars['VLAN_Grp2'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        if templateVars['Name'] == None:
            Error_Return = 'Error on Worksheet %s Row %s.  Could not Determine the Name of the VLAN Pool.' % (ws.title, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "vlan_pool.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'vlan_pool_%s.tf' % (templateVars['Name'])
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        template_file = "data_vlan_pool.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'data_vlan_pool_%s.tf' % (templateVars['Name'])
        dest_dir = 'VLANs'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # Add VLAN(s) to VLAN Pool FIle
        if re.search('Grp_[A-F]', templateVars['Site_Group']):
            Group_ID = '%s' % (templateVars['Site_Group'])
            site_group = ast.literal_eval(os.environ[Group_ID])
            for x in range(1, 13):
                sitex = 'Site_%s' % (x)
                if not site_group[sitex] == None:
                    Site_ID = 'Site_ID_%s' % (site_group[sitex])
                    site_dict = ast.literal_eval(os.environ[Site_ID])

                    # Create templateVars for Site_Name and APIC_URL
                    templateVars['Site_Name'] = site_dict.get('Site_Name')
                    templateVars['APIC_URL'] = site_dict.get('APIC_URL')

                    # Create Blank VLAN Pool VLAN(s) File
                    dest_file = './ACI/%s/VLANs/vlp_%s.tf' % (templateVars['Site_Name'], templateVars['Name'])
                    wr_file = open(dest_file, 'w')
                    wr_file.close()
                    dest_file = 'vlan_pool_%s.tf' % (templateVars['Name'])
                    dest_dir = 'VLANs'
                    template_file = "add_vlan_to_pool.template"
                    template = self.templateEnv.get_template(template_file)

                    for z in range(1, 3):
                        vgroup = 'VLAN_Grp%s' % (z)
                        vgrp = 'VGRP%s_Allocation' % (z)
                        templateVars['Allocation_Mode'] = templateVars[vgrp]
                        if re.search(r'\d+', str(templateVars[vgroup])):
                            vlan_list = vlan_list_full(templateVars[vgroup])
                            for v in vlan_list:
                                vlan = str(v)
                                if re.fullmatch(r'\d+', vlan):
                                    templateVars['VLAN_ID'] = int(vlan)

                                    # Add VLAN to VLAN Pool File
                                    create_tf_file('a+', dest_dir, dest_file, template, **templateVars)

        elif re.search(r'\d+', templateVars['Site_Group']):
            Site_ID = 'Site_ID_%s' % (templateVars['Site_Group'])
            site_dict = ast.literal_eval(os.environ[Site_ID])

            # Create templateVars for Site_Name and APIC_URL
            templateVars['Site_Name'] = site_dict.get('Site_Name')
            templateVars['APIC_URL'] = site_dict.get('APIC_URL')

            # Create Blank VLAN Pool VLAN(s) File
            dest_file = './ACI/%s/VLANs/vlan_pool_%s.tf' % (templateVars['Site_Name'], templateVars['Name'])
            wr_file = open(dest_file, 'w')
            wr_file.close()
            dest_file = 'vlan_pool_%s.tf' % (templateVars['Name'])
            dest_dir = 'VLANs'
            template_file = "add_vlan_to_pool.template"
            template = self.templateEnv.get_template(template_file)

            for z in range(1, 3):
                vgroup = 'VLAN_Grp%s' % (z)
                vgrp = 'VGRP%s_Allocation' % (z)
                templateVars['Allocation_Mode'] = templateVars[vgrp]
                if re.search(r'\d+', str(templateVars[vgroup])):
                    vlan_list = vlan_list_full(templateVars[vgroup])
                    for v in vlan_list:
                        vlan = str(v)
                        if re.fullmatch(r'\d+', vlan):
                            templateVars['VLAN_ID'] = int(vlan)

                            # Add VLAN to VLAN Pool File
                            create_tf_file('a+', dest_dir, dest_file, template, **templateVars)
        else:
            print(f"\n-----------------------------------------------------------------------------\n")
            print(f"   Error on Worksheet {ws.title}, Row {row_num} Site_Group, value {templateVars['Site_Group']}.")
            print(f"   Unable to Determine if this is a Single or Group of Site(s).  Exiting....")
            print(f"\n-----------------------------------------------------------------------------\n")
            exit()

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def vpc_pair(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'VPC_ID': '',
                         'Name': '',
                         'Node1_ID': '',
                         'Node2_ID': ''}
        optional_args = { }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.number_check(row_num, ws, 'VPC_ID', templateVars['VPC_ID'], 1, 1000)
            validating.number_check(row_num, ws, 'Node1_ID', templateVars['Node1_ID'], 101, 4001)
            validating.number_check(row_num, ws, 'Node2_ID', templateVars['Node2_ID'], 101, 4001)
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "vpc_domain.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'vpc_domain_%s.tf' % (templateVars['VPC_ID'])
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

# Terraform ACI Provider - Admin Policies
# Class must be instantiated with Variables
class Admin_Policies(object):
    def __init__(self, ws):
        self.templateLoader = jinja2.FileSystemLoader(
            searchpath=(aci_template_path + 'Admin_Policies/'))
        self.templateEnv = jinja2.Environment(loader=self.templateLoader)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def backup_host(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Remote_Host': '',
                         'Mgmt_EPG': '',
                         'Protocol': '',
                         'Remote_Path': '',
                         'Port': '',
                         'Auth_Type': '',
                         'Pwd_or_SSHPhrase': ''}
        optional_args = {'Description': '',
                         'Backup_User': '',
                         'SSH_Key': '',
                         'Description': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])

            if re.match(r'\:', templateVars['Remote_Host']):
                validating.ip_address(row_num, ws, 'Remote_Host', templateVars['Remote_Host'])
            elif re.match(r'[a-zA-Z]', templateVars['Remote_Host']):
                validating.dns_name(row_num, ws, 'Remote_Host', templateVars['Remote_Host'])
            else:
                validating.ip_address(row_num, ws, 'Remote_Host', templateVars['Remote_Host'])

            validating.sensitive_var(row_num, ws, 'Pwd_or_SSHPhrase', templateVars['Pwd_or_SSHPhrase'])
            if templateVars['Auth_Type'] == 'password':
                validating.sensitive_var(row_num, ws, 'Backup_User', templateVars['Backup_User'])
            else:
                validating.sensitive_var(row_num, ws, 'SSH_Key', templateVars['SSH_Key'])

            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])

            validating.number_check(row_num, ws, 'Port', templateVars['Port'], 1, 65535)
            validating.values(row_num, ws, 'Auth_Type', templateVars['Auth_Type'], ['password', 'ssh-key'])
            validating.values(row_num, ws, 'Protocol', templateVars['Protocol'], ['ftp', 'scp', 'sftp'])
            templateVars['Mgmt_EPG'] = validating.mgmt_epg(row_num, ws, 'Mgmt_EPG', templateVars['Mgmt_EPG'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        if re.search(r'\.', templateVars['Remote_Host']):
            templateVars['Remote_Host_'] = templateVars['Remote_Host'].replace('.', '-')
        else:
            templateVars['Remote_Host_'] = templateVars['Remote_Host'].replace(':', '-')
        if templateVars['Auth_Type'] == 'password':
            templateVars['Auth_Type'] = 'usePassword'
        elif templateVars['Auth_Type'] == 'ssh-key':
            templateVars['Auth_Type'] = 'useSshKeyContents'

        if not templateVars['Pwd_or_SSHPhrase'] == None:
            x = templateVars['Pwd_or_SSHPhrase'].split('r')
            key_number = x[1]
            templateVars['sensitive_var1'] = 'Pwd_or_SSHPhrase%s' % (key_number)

        if templateVars['Auth_Type'] == 'usePassword':
            if not templateVars['Backup_User'] == None:
                x = templateVars['Backup_User'].split('r')
                key_number = x[1]
                templateVars['sensitive_var2'] = 'Backup_User%s' % (key_number)
        else:
            if not templateVars['SSH_Key'] == None:
                x = templateVars['SSH_Key'].split('r')
                key_number = x[1]
                templateVars['sensitive_var2'] = 'SSH_Key%s' % (key_number)

        # Define the Template Source
        template_file = "backup_host.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'backup_host_%s.tf' % (templateVars['Remote_Host_'])
        dest_dir = 'Admin'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        template_file = "variables.template"
        template = self.templateEnv.get_template(template_file)

        # Create Variables File for the Sensitive Variables
        dest_file = '%s_variable.tf' % (templateVars['sensitive_var1'])
        dest_dir = 'Admin'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        templateVars['sensitive_var'] = templateVars['sensitive_var1']
        process_sensitive_var(wb, ws, row_num, dest_dir, dest_file, template, **templateVars)

        dest_file = '%s_variable.tf' % (templateVars['sensitive_var2'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        templateVars['sensitive_var'] = templateVars['sensitive_var2']
        process_sensitive_var(wb, ws, row_num, dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def backup_policy(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Scheduler_Type': '',
                         'Scheduler_Name': '',
                         'Days': '',
                         'Backup_Hour': '',
                         'Backup_Minute': '',
                         'Concurrent_Capacity': '',
                         'Export_Name': '',
                         'Format': '',
                         'Start_Now': '',
                         'Remote_Host': '',
                         'Encryption_Key': ''}
        optional_args = {'Scheduler_Descr': '',
                         'Export_Descr': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.days(row_num, ws, 'Days', templateVars['Days'])
            validating.number_check(row_num, ws, 'Backup_Hour', templateVars['Backup_Hour'], 0, 23)
            validating.number_check(row_num, ws, 'Backup_Minute', templateVars['Backup_Minute'], 0, 59)
            validating.values(row_num, ws, 'Concurrent_Capacity', templateVars['Concurrent_Capacity'], ['unlimited'])
            validating.values(row_num, ws, 'Format', templateVars['Format'], ['json', 'xml'])
            validating.values(row_num, ws, 'Start_Now', templateVars['Start_Now'], ['triggered', 'untriggered'])
            validating.sensitive_var(row_num, ws, 'Encryption_Key', templateVars['Encryption_Key'])
            if not templateVars['Scheduler_Descr'] == None:
                validating.description(row_num, ws, 'Scheduler_Descr', templateVars['Scheduler_Descr'])
            if not templateVars['Export_Descr'] == None:
                validating.description(row_num, ws, 'Export_Descr', templateVars['Export_Descr'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        if re.search(r'\.', templateVars['Remote_Host']):
            templateVars['Remote_Host_'] = templateVars['Remote_Host'].replace('.', '-')
        else:
            templateVars['Remote_Host_'] = templateVars['Remote_Host'].replace(':', '-')

        if not templateVars['Encryption_Key'] == None:
            x = templateVars['Encryption_Key'].split('r')
            key_number = x[1]
            templateVars['sensitive_var'] = 'Encryption_Key%s' % (key_number)

        # Define the Template Source
        template_file = "global_key.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'global_key.tf'
        dest_dir = 'Admin'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        template_file = "variables.template"
        template = self.templateEnv.get_template(template_file)

        # Create Variables File for the Sensitive Variables
        dest_file = '%s_variable.tf' % (templateVars['sensitive_var'])
        dest_dir = 'Admin'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)
        process_sensitive_var(wb, ws, row_num, dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def firmware(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'MG_Name': '',
                         'Admin_State': '',
                         'Admin_Notify': '',
                         'Graceful': '',
                         'Ignore_Compatability': '',
                         'Run_Mode': '',
                         'SW_Version': '',
                         'Ver_Check_Override': '',
                         'FW_Type': '',
                         'MG_Type': ''}
        optional_args = { }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'MG_Name', templateVars['MG_Name'])
            validating.values(row_num, ws, 'Admin_State', templateVars['Admin_State'], ['triggered', 'untriggered'])
            validating.values(row_num, ws, 'Admin_Notify', templateVars['Admin_Notify'], ['notifyAlwaysBetweenSets', 'notifyNever', 'notifyOnlyOnFailures'])
            validating.values(row_num, ws, 'Graceful', templateVars['Graceful'], ['no', 'yes'])
            validating.values(row_num, ws, 'Ignore_Compatability',templateVars['Ignore_Compatability'], ['no', 'yes'])
            validating.values(row_num, ws, 'Run_Mode', templateVars['Run_Mode'], ['pauseAlwaysBetweenSets', 'pauseNever', 'pauseOnlyOnFailures'])
            validating.values(row_num, ws, 'Ver_Check_Override', templateVars['Ver_Check_Override'], ['trigger', 'trigger-immediate', 'triggered', 'untriggered'])
            validating.values(row_num, ws, 'MG_Type', templateVars['MG_Type'], ['ALL', 'range'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "maintenance_group.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'maintenance_group_%s.tf' % (templateVars['MG_Name'])
        dest_dir = 'Admin'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def login_domain(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Login_Domain': '',
                         'Realm_Type': ''}
        optional_args = {'Description': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_complexity(row_num, ws, 'Login_Domain', templateVars['Login_Domain'])
            validating.values(row_num, ws, 'Realm_Type', templateVars['Realm_Type'], ['RADIUS', 'TACACS+'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        if templateVars['Realm_Type'] == 'RADIUS':
            template_file = "login_domain_radius.template"
        else:
            template_file = "login_domain_tacacs.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        if templateVars['Realm_Type'] == 'RADIUS':
            dest_file = 'login_domain_%s_%s.tf' % ('RADIUS', templateVars['Login_Domain'])
        else:
            dest_file = 'login_domain_%s_%s.tf' % ('TACACS', templateVars['Login_Domain'])
        dest_dir = 'Admin'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def radius(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'RADIUS_Server': '',
                         'Port': '',
                         'RADIUS_Secret': '',
                         'Authz_Proto': '',
                         'Timeout': '',
                         'Retry_Interval': '',
                         'Mgmt_EPG': '',
                         'Login_Domain': '',
                         'Domain_Order': ''}
        optional_args = {'Description': '',
                         'Domain_Descr': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.ip_address(row_num, ws, 'RADIUS_Server', templateVars['RADIUS_Server'])
            validating.name_complexity(row_num, ws, 'Login_Domain', templateVars['Login_Domain'])
            validating.number_check(row_num, ws, 'Domain_Order', templateVars['Domain_Order'], 0, 17)
            validating.number_check(row_num, ws, 'Port', templateVars['Port'], 1, 65535)
            validating.number_check(row_num, ws, 'Retry_Interval', templateVars['Retry_Interval'], 1, 5)
            validating.sensitive_var(row_num, ws, 'RADIUS_Secret', templateVars['RADIUS_Secret'])
            validating.timeout(row_num, ws, 'Timeout', templateVars['Timeout'])
            validating.values(row_num, ws, 'Authz_Proto', templateVars['Authz_Proto'], ['chap', 'mschap', 'pap'])
            templateVars['Mgmt_EPG'] = validating.mgmt_epg(row_num, ws, 'Mgmt_EPG', templateVars['Mgmt_EPG'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
            if not templateVars['Domain_Descr'] == None:
                validating.description(row_num, ws, 'Domain_Descr', templateVars['Domain_Descr'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        if re.search(r'\.', templateVars['RADIUS_Server']):
            templateVars['RADIUS_Server_'] = templateVars['RADIUS_Server'].replace('.', '-')
        else:
            templateVars['RADIUS_Server_'] = templateVars['RADIUS_Server'].replace(':', '-')

        if not templateVars['RADIUS_Secret'] == None:
            x = templateVars['RADIUS_Secret'].split('r')
            key_number = x[1]
            templateVars['sensitive_var'] = 'RADIUS_Secret%s' % (key_number)

        # Define the Template Source
        template_file = "radius.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'radius_server_%s.tf' % (templateVars['RADIUS_Server_'])
        dest_dir = 'Admin'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        template_file = "variables.template"
        template = self.templateEnv.get_template(template_file)

        # Create Variables File for the Sensitive Variables
        dest_file = '%s_variable.tf' % (templateVars['sensitive_var'])
        dest_dir = 'Admin'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)
        process_sensitive_var(wb, ws, row_num, dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def realm(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Auth_Realm': '',
                         'Domain_Type': ''}
        optional_args = {'Login_Domain': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.login_type(row_num, ws, 'Auth_Realm', templateVars['Auth_Realm'], 'Domain_Type', templateVars['Domain_Type'])
            if not templateVars['Domain_Type'] == 'local':
                validating.name_complexity(row_num, ws, 'Login_Domain', templateVars['Login_Domain'])
            validating.values(row_num, ws, 'Auth_Realm', templateVars['Auth_Realm'], ['console', 'default'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        if templateVars['Auth_Realm'] == 'console':
            templateVars['child_class'] = 'aaaConsoleAuth'
        elif templateVars['Auth_Realm'] == 'default':
            templateVars['child_class'] = 'aaaDefaultAuth'

        # Define the Template Source
        template_file = "realm.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        if templateVars['Auth_Realm'] == 'console':
            dest_file = 'realm_console.tf'
        else:
            dest_file = 'realm_default.tf'
        dest_dir = 'Admin'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def security(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Passwd_Strength': '',
                         'Enforce_Intv': '',
                         'Passwd_Intv': '',
                         'Number_Allowed': '',
                         'Passwd_Store': '',
                         'Lockout': '',
                         'Failed_Attempts': '',
                         'Time_Period': '',
                         'Dur_Lockout': '',
                         'Token_Timeout': '',
                         'Maximum_Valid': '',
                         'Web_Timeout': ''}
        optional_args = { }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.number_check(row_num, ws, 'Passwd_Intv', templateVars['Passwd_Intv'], 0, 745)
            validating.number_check(row_num, ws, 'Number_Allowed', templateVars['Number_Allowed'], 0, 10)
            validating.number_check(row_num, ws, 'Passwd_Store', templateVars['Passwd_Store'], 0, 15)
            validating.number_check(row_num, ws, 'Failed_Attempts', templateVars['Failed_Attempts'], 1, 15)
            validating.number_check(row_num, ws, 'Time_Period', templateVars['Time_Period'], 1, 720)
            validating.number_check(row_num, ws, 'Dur_Lockout', templateVars['Dur_Lockout'], 1, 1440)
            validating.number_check(row_num, ws, 'Token_Timeout', templateVars['Token_Timeout'], 300, 9600)
            validating.number_check(row_num, ws, 'Maximum_Valid', templateVars['Maximum_Valid'], 0, 24)
            validating.number_check(row_num, ws, 'Web_Timeout', templateVars['Web_Timeout'], 60, 65525)
            validating.values(row_num, ws, 'Enforce_Intv', templateVars['Enforce_Intv'], ['disable', 'enable'])
            validating.values(row_num, ws, 'Lockout', templateVars['Lockout'], ['disable', 'enable'])
            validating.values(row_num, ws, 'Passwd_Strength', templateVars['Passwd_Strength'], ['no', 'yes'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "security.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'security.tf'
        dest_dir = 'Admin'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def tacacs(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'TACACS_Server': '',
                         'Port': '',
                         'TACACS_Secret': '',
                         'Auth_Proto': '',
                         'Timeout': '',
                         'Retry_Interval': '',
                         'Mgmt_EPG': '',
                         'Login_Domain': '',
                         'Domain_Order': '',
                         'Acct_DestGrp_Name': ''}
        optional_args = {'Description': '',
                         'Domain_Descr': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.ip_address(row_num, ws, 'TACACS_Server', templateVars['TACACS_Server'])
            validating.name_complexity(row_num, ws, 'Login_Domain', templateVars['Login_Domain'])
            validating.number_check(row_num, ws, 'Domain_Order', templateVars['Domain_Order'], 0, 17)
            validating.number_check(row_num, ws, 'Port', templateVars['Port'], 1, 65535)
            validating.number_check(row_num, ws, 'Retry_Interval', templateVars['Retry_Interval'], 1, 5)
            validating.sensitive_var(row_num, ws, 'TACACS_Secret', templateVars['TACACS_Secret'])
            validating.timeout(row_num, ws, 'Timeout', templateVars['Timeout'])
            validating.values(row_num, ws, 'Auth_Proto', templateVars['Auth_Proto'], ['chap', 'mschap', 'pap'])
            templateVars['Mgmt_EPG'] = validating.mgmt_epg(row_num, ws, 'Mgmt_EPG', templateVars['Mgmt_EPG'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
            if not templateVars['Domain_Descr'] == None:
                validating.description(row_num, ws, 'Domain_Descr', templateVars['Domain_Descr'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        if re.search(r'\.', templateVars['TACACS_Server']):
            templateVars['TACACS_Server_'] = templateVars['TACACS_Server'].replace('.', '-')
        else:
            templateVars['TACACS_Server_'] = templateVars['TACACS_Server'].replace(':', '-')

        if not templateVars['TACACS_Secret'] == None:
            x = templateVars['TACACS_Secret'].split('r')
            key_number = x[1]
            templateVars['sensitive_var'] = 'TACACS_Secret%s' % (key_number)

        # Define the Template Source
        template_file = "tacacs.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'tacacs_server_%s.tf' % (templateVars['TACACS_Server_'])
        dest_dir = 'Admin'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        template_file = "variables.template"
        template = self.templateEnv.get_template(template_file)

        # Create Variables File for the Sensitive Variables
        dest_file = '%s_variable.tf' % (templateVars['sensitive_var'])
        dest_dir = 'Admin'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)
        process_sensitive_var(wb, ws, row_num, dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def tacacs_acct(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Acct_DestGrp_Name': '',
                         'Acct_SrcGrp_Name': ''}
        optional_args = {'Description': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Acct_DestGrp_Name', templateVars['Acct_DestGrp_Name'])
            validating.name_rule(row_num, ws, 'Acct_SrcGrp_Name', templateVars['Acct_SrcGrp_Name'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "tacacs_accounting.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'tacacs_accouting_grp_%s.tf' % (templateVars['Acct_DestGrp_Name'])
        dest_dir = 'Admin'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

# Terraform ACI Provider - Admin Policies
# Class must be instantiated with Variables
class Best_Practices(object):
    def __init__(self, ws):
        self.templateLoader = jinja2.FileSystemLoader(
            searchpath=(aci_template_path + 'Best_Practices/'))
        self.templateEnv = jinja2.Environment(loader=self.templateLoader)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def bgp_asn(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'BGP_ASN': ''}
        optional_args = {}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.number_check(row_num, ws, 'BGP_ASN', templateVars['BGP_ASN'], 1, 4294967295)
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "bgp_asn.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'bgp_asn_%s.tf' % (templateVars['BGP_ASN'])
        dest_dir = 'System'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def bgp_rr(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Node_ID': ''}
        optional_args = {}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.number_check(row_num, ws, 'Node_ID', templateVars['Node_ID'], 101, 4001)
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "bgp_rr.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'bgp_rr_%s.tf' % (templateVars['Node_ID'])
        dest_dir = 'System'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def ep_controls(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'EP_Loop_State': '',
                         'EP_Loop_Interval': '',
                         'EP_Loop_Multiplier': '',
                         'BD_Learn_Disable': '',
                         'Port_Disable': '',
                         'Rogue_State': '',
                         'Rogue_Interval': '',
                         'Rogue_Multiplier': '',
                         'Hold_Interval': '',
                         'IP_Aging_State': ''}
        optional_args = { }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.number_check(row_num, ws, 'EP_Loop_Interval', templateVars['EP_Loop_Interval'], 30, 300)
            validating.number_check(row_num, ws, 'EP_Loop_Multiplier', templateVars['EP_Loop_Multiplier'], 1, 255)
            validating.number_check(row_num, ws, 'Hold_Interval', templateVars['Hold_Interval'], 1800, 3600)
            validating.number_check(row_num, ws, 'Rogue_Interval', templateVars['Rogue_Interval'], 0, 65535)
            validating.number_check(row_num, ws, 'Rogue_Multiplier', templateVars['Rogue_Multiplier'], 2, 10)
            validating.values(row_num, ws, 'BD_Learn_Disable', templateVars['BD_Learn_Disable'], ['no', 'yes'])
            validating.values(row_num, ws, 'EP_Loop_State', templateVars['EP_Loop_State'], ['disabled', 'enabled'])
            validating.values(row_num, ws, 'IP_Aging_State', templateVars['IP_Aging_State'], ['disabled', 'enabled'])
            validating.values(row_num, ws, 'Port_Disable', templateVars['Port_Disable'], ['no', 'yes'])
            validating.values(row_num, ws, 'Rogue_State', templateVars['Rogue_State'], ['disabled', 'enabled'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        #Combine the Control Elements
        templateVars['action'] = ''
        action_count = 0
        if templateVars['BD_Learn_Disable'] == 'yes':
            templateVars['action'] = 'bd-learn-disable'
            action_count =+ 1
        if templateVars['Port_Disable'] == 'yes':
            if action_count == 0:
                templateVars['action'] = 'port-disable'
                scope_count =+ 1
            else:
                templateVars['action'] = 'bd-learn-disable,port-disable'

        # Define the Template Source
        template_file = "ep_controls.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'ep_controls.tf'
        dest_dir = 'System'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def error_recovery(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Recovery_Interval': '',
                         'EP_Move': '',
                         'BPDU_Guard': '',
                         'MCP_Loop': ''}
        optional_args = { }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.number_check(row_num, ws, 'Recovery_Interval', templateVars['Recovery_Interval'], 30, 65535)
            validating.values(row_num, ws, 'EP_Move', templateVars['EP_Move'], ['no', 'yes'])
            validating.values(row_num, ws, 'BPDU_Guard', templateVars['BPDU_Guard'], ['no', 'yes'])
            validating.values(row_num, ws, 'MCP_Loop', templateVars['MCP_Loop'], ['no', 'yes'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "error_recovery.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'error_recovery.tf'
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def fabric_settings(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'MGMT_Pref': '',
                         'Type': '',
                         'Enable_DOM': '',
                         'Feature_Selection': '',
                         'BFD_ISIS_Policy': '',
                         'Preserve_CoS': ''}
        optional_args = {'Description': '',
                         'Alias': '',
                         'L3_Description': '',}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.number_check(row_num, ws, 'Enable_DOM', templateVars['Enable_DOM'], 0, 1)
            validating.values(row_num, ws, 'BFD_ISIS_Policy', templateVars['BFD_ISIS_Policy'], ['disabled', 'enabled'])
            validating.values(row_num, ws, 'MGMT_Pref', templateVars['MGMT_Pref'], ['inband', 'ooband'])
            validating.values(row_num, ws, 'Preserve_CoS', templateVars['Preserve_CoS'], ['no', 'yes'])
            validating.values(row_num, ws, 'Type', templateVars['Type'], ['compatible', 'strict'])
            validating.values(row_num, ws, 'Feature_Selection', templateVars['Feature_Selection'], ['analytics', 'netflow', 'telemetry'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
            if not templateVars['L3_Description'] == None:
                validating.description(row_num, ws, 'L3_Description', templateVars['L3_Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Convert the Preserve_CoS value to APIC Format
        if templateVars['Preserve_CoS'] == 'yes':
            templateVars['Preserve_CoS'] = 'dot1p-preserve'
        else:
            templateVars['Preserve_CoS'] = None

        # Define the Template Source
        template_file = "apic_preference.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'apic_preference.tf'
        dest_dir = 'System'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        template_file = "coop_policy.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'coop_policy.tf'
        dest_dir = 'System'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        template_file = "l3_interface.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'l3_interface.tf'
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        template_file = "node_control.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'node_control.tf'
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        template_file = "preserve_cos.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'preserve_cos.tf'
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def fabric_wide(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Disable_Remote_EP_Learn': '',
                         'Enforce_Subnet': '',
                         'EPG_VLAN_Validate': '',
                         'Domain_Validation': '',
                         'Opflex_Auth': '',
                         'Reallocate_Gipo': '',
                         'Restrict_Infra_VLAN': '',
                         'Tracking_State': '',
                         'Delay_Timer': '',
                         'Min_Links': '',
                         'APIC_Ports': ''}
        optional_args = { }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.number_check(row_num, ws, 'Delay_Timer', templateVars['Delay_Timer'], 1, 300)
            validating.number_check(row_num, ws, 'Min_Links', templateVars['Min_Links'], 0, 12)
            validating.values(row_num, ws, 'Disable_Remote_EP_Learn', templateVars['Disable_Remote_EP_Learn'], ['no', 'yes'])
            validating.values(row_num, ws, 'Enforce_Subnet', templateVars['Enforce_Subnet'], ['no', 'yes'])
            validating.values(row_num, ws, 'EPG_VLAN_Validate', templateVars['EPG_VLAN_Validate'], ['no', 'yes'])
            validating.values(row_num, ws, 'Domain_Validation', templateVars['Domain_Validation'], ['no', 'yes'])
            validating.values(row_num, ws, 'Opflex_Auth', templateVars['Opflex_Auth'], ['no', 'yes'])
            validating.values(row_num, ws, 'Reallocate_Gipo', templateVars['Reallocate_Gipo'], ['no', 'yes'])
            validating.values(row_num, ws, 'Restrict_Infra_VLAN', templateVars['Restrict_Infra_VLAN'], ['no', 'yes'])
            validating.values(row_num, ws, 'Tracking_State', templateVars['Tracking_State'], ['on', 'off'])
            validating.values(row_num, ws, 'APIC_Ports', templateVars['APIC_Ports'], ['no', 'yes'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "fabric_wide.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'fabric_wide.tf'
        dest_dir = 'System'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        template_file = "port_tracking.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'port_tracking.tf'
        dest_dir = 'System'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def isis_policy(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'ISIS_MTU': '',
                         'ISIS_Metric': '',
                         'LSP_Flood_Mode': '',
                         'LSP_Initial_Interval': '',
                         'LSP_Max_Interval': '',
                         'LSP_Second_Interval': '',
                         'SPF_Initial_Interval': '',
                         'SPF_Max_Interval': '',
                         'SPF_Second_Interval': ''}
        optional_args = { }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.number_check(row_num, ws, 'ISIS_MTU', templateVars['ISIS_MTU'], 128, 4352)
            validating.number_check(row_num, ws, 'ISIS_Metric', templateVars['ISIS_Metric'], 1, 63)
            validating.number_check(row_num, ws, 'LSP_Initial_Interval', templateVars['LSP_Initial_Interval'], 50, 120000)
            validating.number_check(row_num, ws, 'LSP_Max_Interval', templateVars['LSP_Max_Interval'], 50, 120000)
            validating.number_check(row_num, ws, 'LSP_Second_Interval', templateVars['LSP_Second_Interval'], 50, 120000)
            validating.number_check(row_num, ws, 'SPF_Initial_Interval', templateVars['SPF_Initial_Interval'], 50, 120000)
            validating.number_check(row_num, ws, 'SPF_Max_Interval', templateVars['SPF_Max_Interval'], 50, 120000)
            validating.number_check(row_num, ws, 'SPF_Second_Interval', templateVars['SPF_Second_Interval'], 50, 120000)
            validating.values(row_num, ws, 'LSP_Flood_Mode', templateVars['LSP_Flood_Mode'], ['disabled', 'enabled'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "isis_policy.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'isis_policy.tf'
        dest_dir = 'System'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def mcp_policy(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Admin_State': '',
                         'Control': '',
                         'MCP_Key': '',
                         'Detect_Multiplier': '',
                         'Loop_Action': '',
                         'Initial_Delay': '',
                         'Frequency_Seconds': '',
                         'Frequency_msec': ''}
        optional_args = {'Description': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.number_check(row_num, ws, 'Detect_Multiplier', templateVars['Detect_Multiplier'], 1, 255)
            validating.number_check(row_num, ws, 'Initial_Delay', templateVars['Initial_Delay'], 0, 1800)
            validating.number_check(row_num, ws, 'Frequency_Seconds', templateVars['Frequency_Seconds'], 0, 300)
            validating.number_check(row_num, ws, 'Frequency_msec', templateVars['Frequency_msec'], 0, 999)
            validating.values(row_num, ws, 'Admin_State', templateVars['Admin_State'], ['disabled', 'enabled'])
            validating.values(row_num, ws, 'Control', templateVars['Control'], ['no', 'yes'])
            validating.values(row_num, ws, 'Loop_Action', templateVars['Loop_Action'], ['no', 'yes'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        #Modify the Control Value
        if templateVars['Control'] == 'yes':
            templateVars['Control'] = 'pdu-per-vlan'
        else:
            templateVars['Control'] = None

        # Convert the Loop_Action value to APIC Format
        if templateVars['Loop_Action'] == 'yes':
            templateVars['Loop_Action'] = 'port-disable'
        else:
            templateVars['Loop_Action'] = None

        if not templateVars['MCP_Key'] == None:
            x = templateVars['MCP_Key'].split('r')
            key_number = x[1]
            templateVars['sensitive_var'] = 'MCP_Key%s' % (key_number)

        # Define the Template Source
        template_file = "mcp_policy.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'mcp_policy.tf'
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # site_dict = ast.literal_eval(os.environ[Site_ID])

        # Define the Template Source
        template_file = "variables.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = '%s_variable.tf' % (templateVars['sensitive_var'])
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        process_sensitive_var(wb, ws, row_num, dest_dir, dest_file, template, **templateVars)

# Terraform ACI Provider - Fabric Policies
# Class must be instantiated with Variables
class Fabric_Policies(object):
    def __init__(self, ws):
        self.templateLoader = jinja2.FileSystemLoader(
            searchpath=(aci_template_path + 'Fabric_Policies/'))
        self.templateEnv = jinja2.Environment(loader=self.templateLoader)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def date_time(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Admin_State': '',
                         'Server_State': '',
                         'Master_Mode': '',
                         'Auth_State': ''}
        optional_args = {'Description': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            validating.values(row_num, ws, 'Admin_State', templateVars['Admin_State'], ['disabled', 'enabled'])
            validating.values(row_num, ws, 'Server_State', templateVars['Server_State'], ['disabled', 'enabled'])
            validating.values(row_num, ws, 'Master_Mode', templateVars['Master_Mode'], ['disabled', 'enabled'])
            validating.values(row_num, ws, 'Auth_State', templateVars['Auth_State'], ['disabled', 'enabled'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        if templateVars['Server_State'] == 'disabled':
            templateVars['Master_Mode'] = 'disabled'

        # Define the Template Source
        template_file = "date_time_profile.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'date_time_profile_%s.tf' % (templateVars['Name'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def dns(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'DNS_Profile': '',
                         'DNS_Server': '',
                         'Preferred': ''}
        optional_args = {}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.ip_address(row_num, ws, 'DNS_Server', templateVars['DNS_Server'])
            validating.values(row_num, ws, 'Preferred', templateVars['Preferred'], ['no', 'yes'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        if re.search(r'\.', templateVars['DNS_Server']):
            templateVars['DNS_Server_'] = templateVars['DNS_Server'].replace('.', '-')
        else:
            templateVars['DNS_Server_'] = templateVars['DNS_Server'].replace(':', '-')

        # Define the Template Source
        template_file = "dns.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'dns_profile_%s_%s.tf' % (templateVars['DNS_Profile'], templateVars['DNS_Server_'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def dns_profile(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Name': '',
                         'Mgmt_EPG': ''}
        optional_args = {'Description': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            templateVars['Mgmt_EPG'] = validating.mgmt_epg(row_num, ws, 'Mgmt_EPG', templateVars['Mgmt_EPG'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "dns_profile.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'dns_profile_%s.tf' % (templateVars['Name'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def domain(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'DNS_Profile': '',
                         'Default_Domain': '',
                         'Domain': ''}
        optional_args = {}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.domain(row_num, ws, 'Domain', templateVars['Domain'])
            validating.values(row_num, ws, 'Default_Domain', templateVars['Default_Domain'], ['no', 'yes'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        templateVars['Domain_'] = templateVars['Domain'].replace('.', '-')

        # Define the Template Source
        template_file = "domain.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'dns_profile_%s_domain_%s.tf' % (templateVars['DNS_Profile'], templateVars['Domain_'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def pod_policy(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Policy_Group': '',
                         'Pod_Profile': '',
                         'Date_Time_Policy': '',
                         'ISIS_Policy': '',
                         'COOP_Group_Policy': '',
                         'BGP_RR_Policy': '',
                         'Mgmt_Access_Policy': '',
                         'SNMP_Policy': '',
                         'MACsec_Policy': ''}
        optional_args = {'PG_Description': '',
                         'Profile_Descr': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Policy_Group', templateVars['Policy_Group'])
            validating.name_rule(row_num, ws, 'Pod_Profile', templateVars['Pod_Profile'])
            if not templateVars['PG_Description'] == None:
                validating.description(row_num, ws, 'PG_Description', templateVars['PG_Description'])
            if not templateVars['Profile_Descr'] == None:
                validating.description(row_num, ws, 'Profile_Descr', templateVars['Profile_Descr'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        templateVars['ISIS_Policy'] = 'default'
        templateVars['COOP_Group_Policy'] = 'default'
        templateVars['BGP_RR_Policy'] = 'default'

        # Define the Template Source
        template_file = "pod_profile.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'pod_profile_%s.tf' % (templateVars['Pod_Profile'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def ntp(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Date_Policy': '',
                         'NTP_Server': '',
                         'Preferred': '',
                         'Mgmt_EPG': ''}
        optional_args = {'Description': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.ip_address(row_num, ws, 'NTP_Server', templateVars['NTP_Server'])
            validating.values(row_num, ws, 'Preferred', templateVars['Preferred'], ['no', 'yes'])
            templateVars['Mgmt_EPG'] = validating.mgmt_epg(row_num, ws, 'Mgmt_EPG', templateVars['Mgmt_EPG'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        if re.search(r'\.', templateVars['NTP_Server']):
            templateVars['NTP_Server_'] = templateVars['NTP_Server'].replace('.', '-')
        else:
            templateVars['NTP_Server_'] = templateVars['NTP_Server'].replace(':', '-')

        # Define the Template Source
        template_file = "ntp.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'ntp_%s.tf' % (templateVars['NTP_Server_'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def sch_dstgrp(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'DestGrp_Name': '',
                         'Admin_State': '',
                         'SMTP_Port': '',
                         'SMTP_Relay': '',
                         'Mgmt_EPG': '',
                         'From_Email': '',
                         'Reply_Email': '',
                         'To_Email': '',
                         'Contract_Id': '',
                         'Customer_Id': '',
                         'Site_Id': ''}
        optional_args = {'Description': '',
                         'Phone_Number': '',
                         'Contact_Info': '',
                         'Street_Address': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            if re.match(r'\:', templateVars['SMTP_Relay']):
                validating.ip_address(row_num, ws, 'SMTP_Relay', templateVars['SMTP_Relay'])
            elif re.match(r'[a-zA-Z]', templateVars['SMTP_Relay']):
                validating.dns_name(row_num, ws, 'SMTP_Relay', templateVars['SMTP_Relay'])
            else:
                validating.ip_address(row_num, ws, 'SMTP_Relay', templateVars['SMTP_Relay'])
            validating.email(row_num, ws, 'From_Email', templateVars['From_Email'])
            validating.email(row_num, ws, 'Reply_Email', templateVars['Reply_Email'])
            validating.email(row_num, ws, 'To_Email', templateVars['To_Email'])
            validating.name_rule(row_num, ws, 'DestGrp_Name', templateVars['DestGrp_Name'])
            validating.values(row_num, ws, 'Admin_State', templateVars['Admin_State'], ['disabled', 'enabled'])
            templateVars['Mgmt_EPG'] = validating.mgmt_epg(row_num, ws, 'Mgmt_EPG', templateVars['Mgmt_EPG'])
            if not templateVars['Contact_Info'] == None:
                validating.description(row_num, ws, 'Contact_Info', templateVars['Contact_Info'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
            if not templateVars['Phone_Number'] == None:
                validating.phone(row_num, ws, 'Phone_Number', templateVars['Phone_Number'])
            if not templateVars['Street_Address'] == None:
                validating.description(row_num, ws, 'Street_Address', templateVars['Street_Address'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "smartcallhome_dg.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'smartcallhome_dg_%s.tf' % (templateVars['DestGrp_Name'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def sch_receiver(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'DestGrp_Name': '',
                         'Receiver_Name': '',
                         'Admin_State': '',
                         'Email': '',
                         'Format': '',
                         'RFC_Compliant': '',
                         'Audit': '',
                         'Events': '',
                         'Faults': '',
                         'Session': ''}
        optional_args = { }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.email(row_num, ws, 'Email', templateVars['Email'])
            validating.name_rule(row_num, ws, 'DestGrp_Name', templateVars['DestGrp_Name'])
            validating.name_rule(row_num, ws, 'Receiver_Name', templateVars['Receiver_Name'])
            validating.values(row_num, ws, 'Admin_State', templateVars['Admin_State'], ['disabled', 'enabled'])
            validating.values(row_num, ws, 'RFC_Compliant', templateVars['RFC_Compliant'], ['no', 'yes'])
            validating.values(row_num, ws, 'Audit', templateVars['Audit'], ['no', 'yes'])
            validating.values(row_num, ws, 'Events', templateVars['Events'], ['no', 'yes'])
            validating.values(row_num, ws, 'Faults', templateVars['Faults'], ['no', 'yes'])
            validating.values(row_num, ws, 'Session', templateVars['Session'], ['no', 'yes'])
            validating.values(row_num, ws, 'Format', templateVars['Format'], ['aml', 'short-txt', 'xml'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        incl_list = ''
        if not templateVars['Audit'] == 'no':
            incl_list = 'audit'
        if not templateVars['Events'] == 'no':
            if incl_list == '':
                incl_list = 'events'
            else:
                incl_list = incl_list + ',events'
        if not templateVars['Faults'] == 'no':
            if incl_list == '':
                incl_list = 'faults'
            else:
                incl_list = incl_list + ',faults'
        if not templateVars['Session'] == 'no':
            if incl_list == '':
                incl_list = 'session'
            else:
                incl_list = incl_list + ',session'
        templateVars['Included_Types'] = incl_list

        # Define the Template Source
        template_file = "smartcallhome_receiver.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'smartcallhome_dg_%s_receiver_%s.tf' % (templateVars['DestGrp_Name'], templateVars['Receiver_Name'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        template_file = "smartcallhome_source.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'smartcallhome_source.tf'
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def snmp_client(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'SNMP_Policy': '',
                         'Client_Group': '',
                         'SNMP_Client': '',
                         'SNMP_Client_Name': ''}
        optional_args = { }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.ip_address(row_num, ws, 'SNMP_Client', templateVars['SNMP_Client'])
            validating.name_rule(row_num, ws, 'SNMP_Client_Name', templateVars['SNMP_Client_Name'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        if re.search(r'\.', templateVars['SNMP_Client']):
            templateVars['SNMP_Client_'] = templateVars['SNMP_Client'].replace('.', '-')
        else:
            templateVars['SNMP_Client_'] = templateVars['SNMP_Client'].replace(':', '-')

        # Define the Template Source
        template_file = "snmp_client.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'snmp_policy_%s_clientgroup_%s_%s.tf' % (templateVars['SNMP_Policy'], templateVars['Client_Group'], templateVars['SNMP_Client_'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def snmp_comm(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'SNMP_Policy': '',
                         'SNMP_Community': ''}
        optional_args = {'Description': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.snmp_string(row_num, ws, 'SNMP_Community', templateVars['SNMP_Community'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        if not templateVars['SNMP_Community'] == None:
            x = templateVars['SNMP_Community'].split('r')
            key_number = x[1]
            templateVars['sensitive_var'] = 'SNMP_Community%s' % (key_number)

        # Define the Template Source
        template_file = "snmp_comm.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'snmp_policy_%s_comm_%s.tf' % (templateVars['SNMP_Policy'], templateVars['SNMP_Community'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # site_dict = ast.literal_eval(os.environ[Site_ID])

        # Define the Template Source
        template_file = "variables.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = '%s_variable.tf' % (templateVars['sensitive_var'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        process_sensitive_var(wb, ws, row_num, dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def snmp_clgrp(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'SNMP_Policy': '',
                         'Client_Group': '',
                         'Mgmt_EPG': ''}
        optional_args = {'Description': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Client_Group', templateVars['Client_Group'])
            validating.name_rule(row_num, ws, 'SNMP_Policy', templateVars['SNMP_Policy'])
            templateVars['Mgmt_EPG'] = validating.mgmt_epg(row_num, ws, 'Mgmt_EPG', templateVars['Mgmt_EPG'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "snmp_client_group.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'snmp_client_group_%s.tf' % (templateVars['Client_Group'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def snmp_policy(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Policy_Name': '',
                         'Admin_State': ''}
        optional_args = {'Description': '',
                         'SNMP_Contact': '',
                         'SNMP_Location': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Policy_Name', templateVars['Policy_Name'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
            if not templateVars['SNMP_Contact'] == None:
                validating.description(row_num, ws, 'SNMP_Contact', templateVars['SNMP_Contact'])
            if not templateVars['SNMP_Location'] == None:
                validating.description(row_num, ws, 'SNMP_Location', templateVars['SNMP_Location'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "snmp_policy.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'snmp_policy_%s.tf' % (templateVars['Policy_Name'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def snmp_trap(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'SNMP_Policy': '',
                         'SNMP_Trap_DG': '',
                         'Trap_Server': '',
                         'Destination_Port': '',
                         'Version': '',
                         'Community_or_Username': '',
                         'Security_Level': '',
                         'Mgmt_EPG': ''}
        optional_args = { }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        # Set noauth if v1 or v2c
        if re.search('(v1|v2c)', templateVars['Version']):
            templateVars['Security_Level'] = 'noauth'

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.ip_address(row_num, ws, 'Trap_Server', templateVars['Trap_Server'])
            validating.number_check(row_num, ws, 'Destination_Port', templateVars['Destination_Port'], 1, 65535)
            validating.values(row_num, ws, 'Version', templateVars['Version'], ['v1', 'v2c', 'v3'])
            validating.values(row_num, ws, 'Security_Level', templateVars['Security_Level'], ['auth', 'noauth', 'priv'])
            validating.snmp_string(row_num, ws, 'Community_or_Username', templateVars['Community_or_Username'])
            templateVars['Mgmt_EPG'] = validating.mgmt_epg(row_num, ws, 'Mgmt_EPG', templateVars['Mgmt_EPG'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        if re.search(r'\.', templateVars['Trap_Server']):
            templateVars['Trap_Server_'] = templateVars['Trap_Server'].replace('.', '-')
        else:
            templateVars['Trap_Server_'] = templateVars['Trap_Server'].replace(':', '-')

        if not templateVars['Community_or_Username'] == None:
            x = templateVars['Community_or_Username'].split('r')
            key_number = x[1]
            templateVars['sensitive_var'] = 'Community_or_Username%s' % (key_number)

        # Define the Template Source
        template_file = "snmp_trap.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'snmp_trap_server_%s.tf' % (templateVars['Trap_Server_'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # site_dict = ast.literal_eval(os.environ[Site_ID])

        # Define the Template Source
        template_file = "variables.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = '%s_variable.tf' % (templateVars['sensitive_var'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        process_sensitive_var(wb, ws, row_num, dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def snmp_user(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'SNMP_Policy': '',
                         'SNMP_User': '',
                         'Authorization_Type': '',
                         'Authorization_Key': ''}
        optional_args = {'Privacy_Type': '',
                         'Privacy_Key': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            auth_type = templateVars['Authorization_Type']
            auth_key = templateVars['Authorization_Key']
            validating.snmp_auth(row_num, ws, templateVars['Privacy_Type'], templateVars['Privacy_Key'], auth_type, auth_key)
            validating.snmp_string(row_num, ws, 'SNMP_User', templateVars['SNMP_User'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Modify User Input of templateVars['Privacy_Type'] or templateVars['Authorization_Type'] to send to APIC
        if templateVars['Privacy_Type'] == 'none':
            templateVars['Privacy_Type'] = None
        if templateVars['Authorization_Type'] == 'md5':
            templateVars['Authorization_Type'] = None
        if templateVars['Authorization_Type'] == 'sha1':
            templateVars['Authorization_Type'] = 'hmac-sha1-96'

        # Define the Template Source
        template_file = "snmp_user.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'snmp_policy_%s_user_%s.tf' % (templateVars['SNMP_Policy'], templateVars['SNMP_User'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        if not templateVars['Privacy_Key'] == None:
            x = templateVars['Privacy_Key'].split('r')
            key_number = x[1]
            templateVars['sensitive_var1'] = 'Privacy_Key%s' % (key_number)

        if not templateVars['Authorization_Key'] == None:
            x = templateVars['Authorization_Key'].split('r')
            key_number = x[1]
            templateVars['sensitive_var2'] = 'Authorization_Key%s' % (key_number)

        # Define the Template Source
        template_file = "snmp_user.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'snmp_policy_%s_user_%s.tf' % (templateVars['SNMP_Policy'], templateVars['SNMP_User'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # site_dict = ast.literal_eval(os.environ[Site_ID])

        # Define the Template Source
        template_file = "variables.template"
        template = self.templateEnv.get_template(template_file)

        # Create Variables File for the Privacy Key & Authorization Key
        if not templateVars['Privacy_Key'] == None:
            dest_file = '%s_variable.tf' % (templateVars['sensitive_var1'])
            dest_dir = 'Fabric'
            process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

            templateVars['sensitive_var'] = templateVars['sensitive_var1']
            process_sensitive_var(wb, ws, row_num, dest_dir, dest_file, template, **templateVars)

        if not templateVars['Authorization_Key'] == None:
            dest_file = '%s_variable.tf' % (templateVars['sensitive_var2'])
            dest_dir = 'Fabric'
            process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

            templateVars['sensitive_var'] = templateVars['sensitive_var2']
            process_sensitive_var(wb, ws, row_num, dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def syslog_dg(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Dest_Grp_Name': '',
                         'Minimum_Level': '',
                         'Log_Format': '',
                         'Console': '',
                         'Console_Level': '',
                         'Local': '',
                         'Local_Level': '',
                         'Include_msec': '',
                         'Include_timezone': '',
                         'Audit': '',
                         'Events': '',
                         'Faults': '',
                         'Session': ''}
        optional_args = { }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.log_level(row_num, ws, 'Minimum_Level', templateVars['Minimum_Level'])
            validating.log_level(row_num, ws, 'Local_Level', templateVars['Local_Level'])
            validating.log_level(row_num, ws, 'Console_Level', templateVars['Console_Level'])
            validating.name_rule(row_num, ws, 'Dest_Grp_Name', templateVars['Dest_Grp_Name'])
            validating.values(row_num, ws, 'Console', templateVars['Console'], ['disabled', 'enabled'])
            validating.values(row_num, ws, 'Local', templateVars['Local'], ['disabled', 'enabled'])
            validating.values(row_num, ws, 'Include_msec', templateVars['Include_msec'], ['no', 'yes'])
            validating.values(row_num, ws, 'Include_timezone', templateVars['Include_timezone'], ['no', 'yes'])
            validating.values(row_num, ws, 'Audit', templateVars['Audit'], ['no', 'yes'])
            validating.values(row_num, ws, 'Events', templateVars['Events'], ['no', 'yes'])
            validating.values(row_num, ws, 'Faults', templateVars['Faults'], ['no', 'yes'])
            validating.values(row_num, ws, 'Session', templateVars['Session'], ['no', 'yes'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)
        incl_list = ''
        if not templateVars['Audit'] == 'no':
            incl_list = 'audit'
        if not templateVars['Events'] == 'no':
            if incl_list == '':
                incl_list = 'events'
            else:
                incl_list = incl_list + ',events'
        if not templateVars['Faults'] == 'no':
            if incl_list == '':
                incl_list = 'faults'
            else:
                incl_list = incl_list + ',faults'
        if not templateVars['Session'] == 'no':
            if incl_list == '':
                incl_list = 'session'
            else:
                incl_list = incl_list + ',session'
        templateVars['Included_Types'] = incl_list

        # Define the Template Source
        template_file = "syslog_dg.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'syslog_dg_%s.tf' % (templateVars['Dest_Grp_Name'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def syslog_rmt(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Dest_Grp_Name': '',
                         'Syslog_Server': '',
                         'Port': '',
                         'Mgmt_EPG': '',
                         'Severity': '',
                         'Facility': ''}
        optional_args = { }

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.ip_address(row_num, ws, 'Syslog_Server', templateVars['Syslog_Server'])
            validating.log_level(row_num, ws, 'Severity', templateVars['Severity'])
            validating.number_check(row_num, ws, 'Port', templateVars['Port'], 1, 65535)
            validating.syslog_fac(row_num, ws, 'Facility', templateVars['Facility'])
            templateVars['Mgmt_EPG'] = validating.mgmt_epg(row_num, ws, 'Mgmt_EPG', templateVars['Mgmt_EPG'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        if re.search(r'\.', templateVars['Syslog_Server']):
            templateVars['Syslog_Server_'] = templateVars['Syslog_Server'].replace('.', '-')
        else:
            templateVars['Syslog_Server_'] = templateVars['Syslog_Server'].replace(':', '-')

        # Define the Template Source
        template_file = "syslog_rmt.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'syslog_rmt_%s.tf' % (templateVars['Syslog_Server_'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def trap_groups(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'SNMP_Trap_DG': '',
                         'SNMP_Source': '',
                         'Audit': '',
                         'Events': '',
                         'Faults': '',
                         'Session': ''}
        optional_args = {'Description': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'SNMP_Trap_DG', templateVars['SNMP_Trap_DG'])
            validating.name_rule(row_num, ws, 'SNMP_Source', templateVars['SNMP_Source'])
            validating.values(row_num, ws, 'Audit', templateVars['Audit'], ['no', 'yes'])
            validating.values(row_num, ws, 'Events', templateVars['Events'], ['no', 'yes'])
            validating.values(row_num, ws, 'Faults', templateVars['Faults'], ['no', 'yes'])
            validating.values(row_num, ws, 'Session', templateVars['Session'], ['no', 'yes'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        incl_list = ''
        if not templateVars['Audit'] == 'no':
            incl_list = 'audit'
        if not templateVars['Events'] == 'no':
            if incl_list == '':
                incl_list = 'events'
            else:
                incl_list = incl_list + ',events'
        if not templateVars['Faults'] == 'no':
            if incl_list == '':
                incl_list = 'faults'
            else:
                incl_list = incl_list + ',faults'
        if not templateVars['Session'] == 'no':
            if incl_list == '':
                incl_list = 'session'
            else:
                incl_list = incl_list + ',session'
        templateVars['Included_Types'] = incl_list

        # Define the Template Source
        template_file = "snmp_trap_dest_grp.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'snmp_trap_dg_%s.tf' % (templateVars['SNMP_Trap_DG'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        template_file = "snmp_trap_source.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'snmp_trap_source_%s.tf' % (templateVars['SNMP_Source'])
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

# Terraform ACI Provider - Tenants Policies
# Class must be instantiated with Variables
class L3Out_Policies(object):
    def __init__(self, ws):
        self.templateLoader = jinja2.FileSystemLoader(
            searchpath=(aci_template_path + 'L3Out_Policies/'))
        self.templateEnv = jinja2.Environment(loader=self.templateLoader)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def add_l3out(self, wb, ws, row_num, **kwargs):
        # Open the Network Policies Worksheet
        ws_net = wb['Network Policies']
        rowcount = ws_net.max_row

        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Tenant': '',
                         'L3Out': '',
                         'VRF_Tenant': '',
                         'VRF': '',
                         'L3_Domain': '',
                         'L3Out_Policy': '',
                         'enforce_rtctrl': '',
                         'target_dscp': ''}
        optional_args = {'Alias': '',
                         'Description': '',
                         'Tags': '',
                         'cons_vzBrCP': '',
                         'vzCPIf': '',
                         'Master_fvEPg': '',
                         'prov_vzBrCP': '',
                         'vzTaboo': '',
                         'exception_tag': '',
                         'rtctrlProfile': '',
                         'sub_rtctrlProfile': '',
                         'rtsumARtSummPol': ''}


        # Get the L3Out Policies from the Network Policies Tab
        func = 'l3Out'
        count = countKeys(ws_net, func)
        l3_count = ''
        var_dict = findVars(ws_net, func, rowcount, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('L3Out_Policy'):
                l3_count = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.dscp(l3_count, ws_net, 'target_dscp', templateVars['target_dscp'])
            validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
            validating.name_rule(row_num, ws, 'L3Out', templateVars['L3Out'])
            validating.name_rule(row_num, ws, 'VRF', templateVars['VRF'])
            validating.name_rule(row_num, ws, 'VRF_Tenant', templateVars['VRF_Tenant'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "l3out.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'l3out_%s.tf' % (templateVars['L3Out'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def add_extepg(self, wb, ws, row_num, **kwargs):
        # Open the Network Policies Worksheet
        ws_net = wb['Network Policies']
        rowcount = ws_net.max_row

        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Tenant': '',
                         'L3Out': '',
                         'Ext_EPG': '',
                         'Ext_EPG_Policy': '',
                         'Subnets': '',
                         'Ext_Subnet_Policy': '',
                         'prio': '',
                         'target_dscp': '',
                         'pref_gr_memb': '',
                         'match_t': '',
                         'flood': '',
                         'export-rtctrl': '',
                         'import-rtctrl': '',
                         'import-security': '',
                         'shared-security': '',
                         'shared-rtctrl': '',
                         'agg-export': '',
                         'agg-import': '',
                         'agg-shared': ''}
        optional_args = {'Alias': '',
                         'Description': '',
                         'Tags': '',
                         'cons_vzBrCP': '',
                         'vzCPIf': '',
                         'Master_fvEPg': '',
                         'prov_vzBrCP': '',
                         'vzTaboo': '',
                         'exception_tag': '',
                         'rtctrlProfile': '',
                         'sub_rtctrlProfile': '',
                         'rtsumARtSummPol': ''}


        # Get the L3Out Policies from the Network Policies Tab
        func = 'ext_epg'
        count = countKeys(ws_net, func)
        epg_count = ''
        var_dict = findVars(ws_net, func, rowcount, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('Ext_EPG_Policy'):
                epg_count = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}

        func = 'ext_subnet'
        count = countKeys(ws_net, func)
        sub_count = ''
        var_dict = findVars(ws_net, func, rowcount, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('Ext_Subnet_Policy'):
                sub_count = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.dscp(epg_count, ws_net, 'epg_target_dscp', templateVars['epg_target_dscp'])
            validating.match_t(epg_count, ws_net, 'match_t', templateVars['match_t'])
            validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
            if not templateVars['Subnet'] == None:
                if re.search(',', templateVars['Subnet']):
                    sx = templateVars['Subnet'].split(',')
                    for x in sx:
                        validating.ip_address(row_num, ws, 'Subnet', x)
            validating.qos_priority(epg_count, ws_net, 'prio', templateVars['prio'])
            validating.values(sub_count, ws_net, 'agg-export', templateVars['agg-export'], ['no', 'yes'])
            validating.values(sub_count, ws_net, 'agg-import', templateVars['agg-import'], ['no', 'yes'])
            validating.values(sub_count, ws_net, 'agg-shared', templateVars['agg-shared'], ['no', 'yes'])
            validating.values(sub_count, ws_net, 'export-rtctrl', templateVars['export-rtctrl'], ['no', 'yes'])
            validating.values(sub_count, ws_net, 'import-rtctrl', templateVars['import-rtctrl'], ['no', 'yes'])
            validating.values(sub_count, ws_net, 'import-security', templateVars['import-security'], ['no', 'yes'])
            validating.values(sub_count, ws_net, 'shared-security', templateVars['shared-security'], ['no', 'yes'])
            validating.values(sub_count, ws_net, 'shared-rtctrl', templateVars['shared-rtctrl'], ['no', 'yes'])
            validating.values(epg_count, ws_net, 'flood', templateVars['flood'], ['disabled', 'enabled'])
            validating.values(epg_count, ws_net, 'pref_gr_memb', templateVars['pref_gr_memb'], ['exclude', 'include'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Create aggregate templateVars
        aggregate = ''
        aggregate_count = 0
        if templateVars['agg-export'] == 'yes':
            aggregate = '"agg-export"'
            aggregate_count =+ 1
        if templateVars['agg-import'] == 'yes':
            if aggregate_count == 0:
                aggregate = '"agg-import"'
                aggregate_count =+ 1
            else:
                aggregate = aggregate + ', ' + '"agg-import"'
                aggregate_count =+ 1
        if templateVars['agg-shared'] == 'yes':
            if aggregate_count == 0:
                aggregate = '"agg-import"'
                aggregate_count =+ 1
            else:
                aggregate = aggregate + ', ' + '"agg-shared"'
                aggregate_count =+ 1

        if aggregate_count == 0:
            templateVars['aggregate'] = None
        else:
            templateVars['aggregate'] = '[%s]' % (aggregate)

        # Create scope templateVars
        scope = ''
        scope_count = 0
        if templateVars['export-rtctrl'] == 'yes':
            scope = '"export-rtctrl"'
            scope_count =+ 1
        if templateVars['import-rtctrl'] == 'yes':
            if scope_count == 0:
                scope = '"import-rtctrl"'
                scope_count =+ 1
            else:
                scope = scope + ', ' + '"import-rtctrl"'
                scope_count =+ 1
        if templateVars['import-security'] == 'yes':
            if scope_count == 0:
                scope = '"import-security"'
                scope_count =+ 1
            else:
                scope = scope + ', ' + '"import-security"'
                scope_count =+ 1
        if templateVars['shared-security'] == 'yes':
            if scope_count == 0:
                scope = '"shared-security"'
                scope_count =+ 1
            else:
                scope = scope + ', ' + '"shared-security"'
                scope_count =+ 1
        if templateVars['shared-rtctrl'] == 'yes':
            if scope_count == 0:
                scope = '"shared-rtctrl"'
                scope_count =+ 1
            else:
                scope = scope + ', ' + '"shared-rtctrl"'
                scope_count =+ 1

        if scope_count == 0:
            templateVars['scope'] = None
        else:
            templateVars['scope'] = '[%s]' % (scope)

        # Define the Template Source
        template_file = "ext_epg.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'l3out_%s_epg_%s.tf' % (templateVars['L3Out'], templateVars['Ext_EPG'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        if re.search(',', templateVars['Subnet']):
            sx = templateVars['Subnet'].split(',')
            for x in sx:
                templateVars['Subnet'] = x
                templateVars['Subnet_'] = x.replace('.', '-')
                templateVars['Subnet_'] = templateVars['Subnet_'].replace('/', '_')

                # Define the Template Source
                template_file = "ext_subnet.template"
                template = self.templateEnv.get_template(template_file)

                # Process the template through the Sites
                dest_file = 'l3out_%s_epg_%s_subnet_%s.tf' % (templateVars['L3Out'], templateVars['Ext_EPG'], templateVars['Subnet_'])
                dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
                process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)
        else:
            templateVars['Subnet_'] = templateVars['Subnet'].replace('.', '-')
            templateVars['Subnet_'] = templateVars['Subnet_'].replace('/', '_')

            # Define the Template Source
            template_file = "ext_subnet.template"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = 'l3out_%s_epg_%s_subnet_%s.tf' % (templateVars['L3Out'], templateVars['Ext_EPG'], templateVars['Subnet_'])
            dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
            process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def bgp_profile(self, wb, ws, row_num, **kwargs):
        # Open the Network Policies Worksheet
        ws_net = wb['Network Policies']
        rowcount = ws_net.max_row

        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Node_Intf_ID': '',
                         'Address': '',
                         'Remote_ASN': '',
                         'EBGP_Multihop_TTL': '', # (1-255)
                         'Admin_State': '', # (enabled|disabled)
                         'BGP_Policy': '',
                         'Policy_Name': '',
                         'allow_self_as': '', # ctrl [send-com,send-ext-com,nh-self,dis-peer-as-check,allow-self-as,as-override,segment-routing-disable]. default None
                         'as_override': '',
                         'peer_as_check': '',
                         'next_hop_self': '',
                         'send_community': '',
                         'send_ext_community': '',
                         'allowed_self_as_count': '', # count is between 1 and 10
                         'bfd': '', # peerCtrl [bfd,dis-conn-check]
                         'disable_connected_check': '',
                         'AF_Mcast': '', # addrTCtrl [af-ucast,af-mcast,af-label-ucast]
                         'AF_Ucast': '',
                         'remove_all_private_as': '', # privateASctrl [remove-exclusive,remove-all,replace-as]
                         'remove_private_as': '',
                          'private_to_local': ''}
        optional_args = {'Password': '',
                         'Description': '',
                         'BGP_Peer_Prefix_Policy': '', #  uni/tn-common/bgpPfxP-default
                         'Local_ASN_Config': '', # (no-prepen|replace-as|dual-as)
                         'Local_ASN': '',
                         'Weight': ''} # 0 to 65535

        # Get the Node Policies from the Network Policies Tab
        func = 'Interface_Policies'
        count = countKeys(ws_net, func)
        row_count = ''
        var_dict = findVars(ws_net, func, rowcount, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('Interface_Policies'):
                row_count = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.number_check(row_num, ws, 'Node_Intf_ID', templateVars['Node_Intf_ID'], 1, 100)
            if not templateVars['Node2_ID'] == None:
                validating.number_check(row_num, ws, 'Node2_ID', templateVars['Node2_ID'], 101, 4001)
                validating.ip_address(row_num, ws, 'Node2_IP', templateVars['Node2_IP'])
            if not templateVars['VLAN'] == None:
                validating.vlans(row_num, ws, 'VLAN', templateVars['VLAN'])
            validating.qos_priority(row_count, ws_net, 'prio', templateVars['prio'])
            if not templateVars['tag'] == None:
                validating.tag_check(row_count, ws_net, 'tag', templateVars['tag'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "nodep.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'l3out_%s_nodep_%s.tf' % (templateVars['L3Out'], templateVars['NodeP_Name'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        template_file = "nodep_to_fabric_node.template"
        template = self.templateEnv.get_template(template_file)

        # Modify Variables for Template
        templateVars['Node_ID'] = templateVars['Node1_ID']
        templateVars['rtr_id'] = templateVars['Node1_Rotuer_ID']
        templateVars['rtr_id_loop_back'] = templateVars['Node1_Loopback']
        templateVars['tDn'] = 'topology/pod-%s/node-%s' % (templateVars['Node_ID'], templateVars['Node_ID'])

        # Process the template through the Sites
        dest_file = 'l3out_%s_nodep_%s.tf' % (templateVars['L3Out'], templateVars['NodeP_Name'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

        if not templateVars['Node2_ID'] == None:
            # Modify Variables for Template
            templateVars['Node_ID'] = templateVars['Node2_ID']
            templateVars['rtr_id'] = templateVars['Node2_Rotuer_ID']
            templateVars['rtr_id_loop_back'] = templateVars['Node2_Loopback']
            templateVars['tDn'] = 'topology/pod-%s/node-%s' % (templateVars['Node_ID'], templateVars['Node_ID'])

            # Process the template through the Sites
            dest_file = 'l3out_%s_nodep_%s.tf' % (templateVars['L3Out'], templateVars['NodeP_Name'])
            dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
            process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def intf_prof(self, wb, ws, row_num, **kwargs):
        # Open the Network Policies Worksheet
        ws_net = wb['Network Policies']
        rowcount = ws_net.max_row

        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Tenant': '',
                         'L3Out': '',
                         'Node_Profile': '',
                         'NodeIf_Profile': '',
                         'Node_Profile': '',
                         'Interface_Type': '',
                         'Interface_Policies': '',
                         'Node1_ID': '',
                         'Node1_Intf': '',
                         'Node1_IP': '',
                         'prio': ''}
        optional_args = {'Description': '',
                         'Alias': '',
                         'Node2_ID': '',
                         'Node2_Intf': '',
                         'VLAN': '',
                         'Node2_IP': '',
                         'BGP_Profile': '',
                         'EIGRP_Profile': '',
                         'OSPF_Profile': '',
                         'name_alias': '',
                         'annotation': '',
                         'mtu': '',
                         'tag': '',
                         'arpIfPol': '',
                         'egress_qosDppPol': '',
                         'ingress_qosDppPol': '',
                         'qosCustomPol': '',
                         'ndIfPol': '',
                         'netflowMonitorPol': ''}

        # Get the Node Policies from the Network Policies Tab
        func = 'Interface_Policies'
        count = countKeys(ws_net, func)
        row_count = ''
        var_dict = findVars(ws_net, func, rowcount, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('Interface_Policies'):
                row_count = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.number_check(row_num, ws, 'Node1_ID', templateVars['Node1_ID'], 101, 4001)
            if not templateVars['Node2_ID'] == None:
                validating.number_check(row_num, ws, 'Node2_ID', templateVars['Node2_ID'], 101, 4001)
                validating.ip_address(row_num, ws, 'Node2_IP', templateVars['Node2_IP'])
            if not templateVars['VLAN'] == None:
                validating.vlans(row_num, ws, 'VLAN', templateVars['VLAN'])
            validating.qos_priority(row_count, ws_net, 'prio', templateVars['prio'])
            if not templateVars['tag'] == None:
                validating.tag_check(row_count, ws_net, 'tag', templateVars['tag'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "nodep.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'l3out_%s_nodep_%s.tf' % (templateVars['L3Out'], templateVars['NodeP_Name'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        template_file = "nodep_to_fabric_node.template"
        template = self.templateEnv.get_template(template_file)

        # Modify Variables for Template
        templateVars['Node_ID'] = templateVars['Node1_ID']
        templateVars['rtr_id'] = templateVars['Node1_Rotuer_ID']
        templateVars['rtr_id_loop_back'] = templateVars['Node1_Loopback']
        templateVars['tDn'] = 'topology/pod-%s/node-%s' % (templateVars['Node_ID'], templateVars['Node_ID'])

        # Process the template through the Sites
        dest_file = 'l3out_%s_nodep_%s.tf' % (templateVars['L3Out'], templateVars['NodeP_Name'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

        if not templateVars['Node2_ID'] == None:
            # Modify Variables for Template
            templateVars['Node_ID'] = templateVars['Node2_ID']
            templateVars['rtr_id'] = templateVars['Node2_Rotuer_ID']
            templateVars['rtr_id_loop_back'] = templateVars['Node2_Loopback']
            templateVars['tDn'] = 'topology/pod-%s/node-%s' % (templateVars['Node_ID'], templateVars['Node_ID'])

            # Process the template through the Sites
            dest_file = 'l3out_%s_nodep_%s.tf' % (templateVars['L3Out'], templateVars['NodeP_Name'])
            dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
            process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def node_prof(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Tenant': '',
                         'L3Out': '',
                         'Node_Profile': '',
                         'Target_DSCP': '',
                         'Color_Tag': '',
                         'Pod_ID': '',
                         'Node1_ID': '',
                         'Node1_Router_ID': '',
                         'Node1_Loopback': ''}
        optional_args = {'Alias': '',
                         'Description': '',
                         'Node2_ID': '',
                         'Node2_Router_ID': '',
                         'Node2_Loopback': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.dscp(row_num, ws, 'Target_DSCP', templateVars['Target_DSCP'])
            validating.ip_address(row_num, ws, 'Node1_Router_ID', templateVars['Node1_Router_ID'])
            validating.name_rule(row_num, ws, 'Node_Profile', templateVars['Node_Profile'])
            validating.number_check(row_num, ws, 'Node1_ID', templateVars['Node1_ID'], 101, 4001)
            validating.tag_check(row_num, ws, 'Color_Tag', templateVars['Color_Tag'])
            validating.values(row_num, ws, 'Node1_Loopback', templateVars['Node1_Loopback'], ['no', 'yes'])
            if not templateVars['Node2_ID'] == None:
                validating.number_check(row_num, ws, 'Node2_ID', templateVars['Node2_ID'], 101, 4001)
                validating.ip_address(row_num, ws, 'Node2_Router_ID', templateVars['Node2_Router_ID'])
                validating.values(row_num, ws, 'Node2_Loopback', templateVars['Node2_Loopback'], ['no', 'yes'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "nodep.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'l3out_%s_nodep_%s.tf' % (templateVars['L3Out'], templateVars['Node_Profile'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        template_file = "nodep_to_fabric_node.template"
        template = self.templateEnv.get_template(template_file)

        # Modify Variables for Template
        templateVars['Node_ID'] = templateVars['Node1_ID']
        templateVars['rtr_id'] = templateVars['Node1_Router_ID']
        templateVars['rtr_id_loopback'] = templateVars['Node1_Loopback']
        templateVars['tDn'] = 'topology/pod-%s/node-%s' % (templateVars['Pod_ID'], templateVars['Node_ID'])

        # Process the template through the Sites
        dest_file = 'l3out_%s_nodep_%s.tf' % (templateVars['L3Out'], templateVars['Node_Profile'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

        if not templateVars['Node2_ID'] == None:
            # Modify Variables for Template
            templateVars['Node_ID'] = templateVars['Node2_ID']
            templateVars['rtr_id'] = templateVars['Node2_Router_ID']
            templateVars['rtr_id_loop_back'] = templateVars['Node2_Loopback']
            templateVars['tDn'] = 'topology/pod-%s/node-%s' % (templateVars['Pod_ID'], templateVars['Node_ID'])

            # Process the template through the Sites
            dest_file = 'l3out_%s_nodep_%s.tf' % (templateVars['L3Out'], templateVars['Node_Profile'])
            dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
            process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

# Terraform ACI Provider - Site Policies
# Class must be instantiated with Variables
class Site_Policies(object):
    def __init__(self, ws):
        self.templateLoader = jinja2.FileSystemLoader(
            searchpath=(aci_template_path))
        self.templateEnv = jinja2.Environment(loader=self.templateLoader)

    # Method must be called with the following kwargs.
    # Site_ID: Required.  Number to Represeent the Site
    # Site_Name: Required.  A Name for the Site.  Must only contain alphanumeric and underscore
    # APIC_URL: Required.  URL for the APIC for the Site
    # BGP_AS: Required.  Autonomous System for BGP Process
    # SNMP_Location: Required.  SNMP Location for the APIC Cluster
    # Contract_ID: Required.  Contract for Equipment to be used with Smart CallHome Function
    # Customer_Identifier: Required.  Customer Identifier to be used with Smart CallHome Function
    # Site_Identifier: Required.  Site Identifier to be used with Smart CallHome Function
    # Street_Address: Optional.  Street Address for the Site to be used with Smart CallHome Function
    def site_id(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_ID': '',
                         'Site_Name': '',
                         'APIC_URL': '',
                         'APIC_Version': '',
                         'APIC_Auth_Type': '',
                         'Provider_Version': '',
                         'Run_Location': '',
                         'State_Location': ''}
        optional_args = {'Terraform_Cloud_Org': '',
                         'Workspace_Prefix': '',
                         'VCS_Base_Repo': '',
                         'Terraform_Version': '',
                         'Terraform_Cloud_Agent': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Variables
            validating.name_complexity(row_num, ws, 'Site_Name', templateVars['Site_Name'])
            validating.url(row_num, ws, 'APIC_URL', templateVars['APIC_URL'])
            validating.values(row_num, ws, 'APIC_Version', templateVars['APIC_Version'], ['3.X', '4.X', '5.X'])
            validating.values(row_num, ws, 'APIC_Auth_Type', templateVars['APIC_Auth_Type'], ['ssh-key', 'user_pass'])
            validating.values(row_num, ws, 'State_Location', templateVars['State_Location'], ['Local', 'Terraform_Cloud'])
            validating.values(row_num, ws, 'Run_Location', templateVars['Run_Location'], ['Local', 'Terraform_Cloud'])
            if templateVars['State_Location'] == 'Terraform_Cloud':
                validating.not_empty(row_num, ws, 'Terraform_Cloud_Org', templateVars['Terraform_Cloud_Org'])
                validating.not_empty(row_num, ws, 'VCS_Base_Repo', templateVars['VCS_Base_Repo'])
                validating.not_empty(row_num, ws, 'Terraform_Version', templateVars['Terraform_Version'])
                validating.not_empty(row_num, ws, 'Terraform_Cloud_Agent', templateVars['Terraform_Cloud_Agent'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Save the Site Information into Environment Variables
        Site_ID = 'Site_ID_%s' % (templateVars['Site_ID'])
        os.environ[Site_ID] = '%s' % (templateVars)

        # Check to see if the oauth_token is already set in the Environment, and if not set it.
        if templateVars['Run_Location'] == 'Terraform_Cloud':
            if os.environ.get('tfcloud_token') is None:
                print(f'\n-----------------------------------------------------------------------------------\n')
                print(f'  The Run_Location was set to {templateVars["Run_Location"]}.')
                print(f'  To Store the Data in Terraform Cloud we will need a User or Org Token to ')
                print(f'  authenticate to Terraform Cloud.  If you have not already obtained a token see ')
                print(f'  instructions in how to obtain a token Here:')
                print(f'   - https://www.terraform.io/docs/cloud/users-teams-organizations/api-tokens.html')
                print(f'  Please Select "C" to Continue or "Q" to Exit:')
                print(f'\n-----------------------------------------------------------------------------------\n')
                while True:
                    user_response = input('  Please Enter ["C" or "Q"]: ')
                    if re.search('^C$', user_response):
                        break
                    elif user_response == 'Q':
                        exit()
                    else:
                        print(f'\n-----------------------------------------------------------------------------\n')
                        print(f'  A Valid Response is either "C", "Q"...')
                        print(f'\n-----------------------------------------------------------------------------\n')

                while True:
                    try:
                        secure_value = getpass.getpass(prompt=f'Enter the value for the Terraform Cloud Token: ')
                        break
                    except Exception as e:
                        print('Something went wrong. Error received: {}'.format(e))

                # Add the Variable to the Environment
                os.environ['tfcloud_token'] = '%s' % (secure_value)
            else:
                tfcloud_token = os.environ.get('tfcloud_token')

        # Validate if Workspaces Exist and if not Create them.
        if templateVars['State_Location'] == 'Terraform_Cloud':
            orgx = templateVars['TF_Cloud_Org']
            sitex = templateVars['Site_Name']
            if not templateVars['Workspace_Prefix'] == None:
                wspx = templateVars['Workspace_Prefix']
                workspace_list = ['%s_ACI_%s_%s'] % (wspx, sitex, 'Access')
                workspace_list = workspace_list + ['%s_ACI_%s_%s'] % (wspx, sitex, 'Admin')
                workspace_list = workspace_list + ['%s_ACI_%s_%s'] % (wspx, sitex, 'Tenant_common')
                workspace_list = workspace_list + ['%s_ACI_%s_%s'] % (wspx, sitex, 'Tenant_infra')
                workspace_list = workspace_list + ['%s_ACI_%s_%s'] % (wspx, sitex, 'Tenant_mgmt')
                workspace_list = workspace_list + ['%s_ACI_%s_%s'] % (wspx, sitex, 'VLANs')
            else:
                workspace_list = ['ACI_%s_%s'] % (sitex, 'Access')
                workspace_list = workspace_list + ['ACI_%s_%s'] % (sitex, 'Admin')
                workspace_list = workspace_list + ['ACI_%s_%s'] % (sitex, 'Tenant_common')
                workspace_list = workspace_list + ['ACI_%s_%s'] % (sitex, 'Tenant_infra')
                workspace_list = workspace_list + ['ACI_%s_%s'] % (sitex, 'Tenant_mgmt')
                workspace_list = workspace_list + ['ACI_%s_%s'] % (sitex, 'VLANs')
            org = templateVars['TF_Cloud_Org']
            url = 'https://app.terraform.io/api/v2/organizations/%s/workspaces' %  (org)
            tf_token = 'Bearer %s' % (tfcloud_token)
            MyHeaders = {'Authorization': tf_token,
                    'Content-Type': 'application/vnd.api+json'
            }

            get_var = ''
            while get_var == '':
                try:
                    get_var = requests.get(url, headers=MyHeaders)
                    status = get_var.status_code
                except requests.exceptions.ConnectionError as e:
                    print("Connection error, pausing before retrying. Error: {}"
                        .format(e))
                    time.sleep(5)
                except Exception as e:
                    print("failed. Exception: {}".format(e))
                    status = 666
                    return(status)

            # Use this for Troubleshooting
            # if print_response_always:
            #     print(get_var.text)
            # if status != 200 and print_response_on_fail:
            #     print(get_var.text)

            workspace = 'tyscott-aci-admin'
            json_data = get_var.json()
            key_count = 0
            for keys in json_data['data']:
                if keys['attributes']['name'] == workspace:
                    print(keys['id'])
                    workspace_id = keys['id']
                    key_count =+ 1

        # Write the .gitignore to the Appropriate Directories
        template_file = ".gitignore"
        template = self.templateEnv.get_template(template_file)
        dest_file = '.gitignore'
        create_tf_file('w', 'Access', dest_file, template, **templateVars)
        create_tf_file('w', 'VLANs', dest_file, template, **templateVars)
        create_tf_file('w', 'Admin', dest_file, template, **templateVars)
        create_tf_file('w', 'Fabric', dest_file, template, **templateVars)
        create_tf_file('w', 'System', dest_file, template, **templateVars)
        create_tf_file('w', 'Tenant_common', dest_file, template, **templateVars)
        create_tf_file('w', 'Tenant_infra', dest_file, template, **templateVars)
        create_tf_file('w', 'Tenant_mgmt', dest_file, template, **templateVars)

        # Write the main.tf to the Appropriate Directories
        template_file = "main.template"
        template = self.templateEnv.get_template(template_file)
        dest_file = 'main.tf'
        create_tf_file('w', 'Access', dest_file, template, **templateVars)
        #       (wr_method, dest_dir, dest_file, template, **templateVars)
        create_tf_file('w', 'VLANs', dest_file, template, **templateVars)
        create_tf_file('w', 'Admin', dest_file, template, **templateVars)
        create_tf_file('w', 'Fabric', dest_file, template, **templateVars)
        create_tf_file('w', 'System', dest_file, template, **templateVars)
        create_tf_file('w', 'Tenant_common', dest_file, template, **templateVars)
        create_tf_file('w', 'Tenant_infra', dest_file, template, **templateVars)
        create_tf_file('w', 'Tenant_mgmt', dest_file, template, **templateVars)

        # Write the variables.tf to the Appropriate Directories
        template_file = "variables.template"
        dest_file = 'variables.tf'
        create_tf_file('w', 'Access', dest_file, template, **templateVars)
        create_tf_file('w', 'VLANs', dest_file, template, **templateVars)
        create_tf_file('w', 'Admin', dest_file, template, **templateVars)
        create_tf_file('w', 'Fabric', dest_file, template, **templateVars)
        create_tf_file('w', 'System', dest_file, template, **templateVars)
        create_tf_file('w', 'Tenant_common', dest_file, template, **templateVars)
        create_tf_file('w', 'Tenant_infra', dest_file, template, **templateVars)
        create_tf_file('w', 'Tenant_mgmt', dest_file, template, **templateVars)

        # Create Directories and default Terraform Files for Tenants in the Tenants and Bridge_Domains Tab if Needed
        ws_names = ['Tenants', 'Bridge_Domains']
        for ws_name in ws_names:
            ws_tenants = wb[ws_name]
            rows = ws_tenants.max_row
            func_regex = re.compile('^add_')
            func_list = findKeys(ws_tenants, func_regex)
            for func in func_list:
                count = countKeys(ws_tenants, func)
                var_dict = findVars(ws_tenants, func, rows, count)
                for pos in var_dict:
                    row_num = var_dict[pos]['row']
                    del var_dict[pos]['row']
                    for x in list(var_dict[pos].keys()):
                        if var_dict[pos][x] == '':
                            del var_dict[pos][x]
                    if not var_dict[pos].get('Tenant') == None:
                        tenant_dir = 'Tenant_%s' % var_dict[pos].get('Tenant')
                        create_tf_file('w', tenant_dir, template_file, template, **templateVars)

                        # Write the main.tf to the Appropriate Directories
                        template_file = "main.template"
                        template = self.templateEnv.get_template(template_file)
                        dest_file = 'main.tf'

                        create_tf_file('w', tenant_dir, dest_file, template, **templateVars)

                        # Write the variables.tf to the Appropriate Directories
                        template_file = "variables.template"
                        template = self.templateEnv.get_template(template_file)
                        dest_file = 'variables.tf'

                        create_tf_file('w', tenant_dir, dest_file, template, **templateVars)

        site_wb = '%s_intf_selectors.xlsx' % (templateVars['Site_Name'])
        if not os.path.isfile(site_wb):
            wb.save(filename=site_wb)
            wb_wr = load_workbook(site_wb)
            ws_wr = wb_wr.get_sheet_names()
            for sheetName in ws_wr:
                if sheetName not in ['Sites']:
                    sheetToDelete = wb_wr.get_sheet_by_name(sheetName)
                    wb_wr.remove_sheet(sheetToDelete)
            wb_wr.save(filename=site_wb)

    # Method must be called with the following kwargs.
    # Group: Required.  A Group Name to represent a list of Site_ID's
    # Site_1: Required.  The Site_ID for the First Site
    # Site_2: Required.  The Site_ID for the Second Site
    # Site_[3-12]: Optional.  The Site_ID for the 3rd thru the 12th Site(s)
    def group_id(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Group': '',
                         'Site_1': '',
                         'Site_2': ''}
        optional_args = {'Site_3': '',
                         'Site_4': '',
                         'Site_5': '',
                         'Site_6': '',
                         'Site_7': '',
                         'Site_8': '',
                         'Site_9': '',
                         'Site_10': '',
                         'Site_11': '',
                         'Site_12': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        for x in range(1, 13):
            site = 'Site_%s' % (x)
            if not templateVars[site] == None:
                validating.site_group(wb, ws, site, templateVars[site])

        grp_count = 0
        for x in list(map(chr, range(ord('A'), ord('F')+1))):
            grp = 'Grp_%s' % (x)
            if templateVars['Group'] == grp:
                grp_count += 1
        if grp_count == 0:
            print(f'\n-----------------------------------------------------------------------------\n')
            print(f'   Error on Worksheet {ws.title}, Row {row_num} Group, Group_Name "{templateVars["Group"]}" is invalid.')
            print(f'   A valid Group Name is Grp_A thru Grp_F.  Exiting....')
            print(f'\n-----------------------------------------------------------------------------\n')
            exit()

        # Save the Site Information into Environment Variables
        Group_ID = '%s' % (templateVars['Group'])
        os.environ[Group_ID] = '%s' % (templateVars)

# Terraform ACI Provider - Tenants Policies
# Class must be instantiated with Variables
class Tenant_Policies(object):
    def __init__(self, ws):
        self.templateLoader = jinja2.FileSystemLoader(
            searchpath=(aci_template_path + 'Tenant_Policies/'))
        self.templateEnv = jinja2.Environment(loader=self.templateLoader)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def add_app(self, wb, ws, row_num, **kwargs):
        # Open the Network Policies Worksheet
        ws_net = wb['Network Policies']
        rowcount = ws_net.max_row

        # Dicts for Application Profile; required and optional args
        required_args = {'Site_Group': '',
                         'Tenant': '',
                         'App_Profile': '',
                         'App_Policy': '',
                         'Policy_Name': '',
                         'prio': '',
                         'monEPGPol': ''}
        optional_args = {'App_Alias': '',
                         'App_Descr': '',
                         'App_Tags': ''}

        # Get the Application Profile Policies from the Network Policies Tab
        func = 'app'
        count = countKeys(ws_net, func)
        row_count = ''
        var_dict = findVars(ws_net, func, rowcount, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('App_Policy'):
                row_count = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}
                break

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
            validating.name_rule(row_num, ws, 'App_Profile', templateVars['App_Profile'])
            validating.qos_priority(row_count, ws_net, 'prio', templateVars['prio'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        if templateVars['monEPGPol'] == 'default':
            templateVars['monEPGPol'] = 'uni/tn-common/monepg-default'

        if re.search('^(common|mgmt|infra)$', templateVars['Tenant']):
            templateVars['Tenant_Dn'] = 'data.aci_tenant.%s' % (templateVars['Tenant'])
        else:
            templateVars['Tenant_Dn'] = 'aci_tenant.%s' % (templateVars['Tenant'])

        # Define the Template Source
        template_file = "app.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'app_%s.tf' % (templateVars['App_Profile'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def add_bd(self, wb, ws, row_num, **kwargs):
        # Assign the kwargs to a initial var for each process
        initial_kwargs = kwargs

        # Open the Network Policies Worksheet
        ws_net = wb['Network Policies']
        rowcount = ws_net.max_row

        # Dicts for Bridge Domain required and optional args
        required_args = {'Site_Group': '',
                         'Tenant': '',
                         'Bridge_Domain': '',
                         'BD_Policy': '',
                         'VRF_Tenant': '',
                         'VRF': '',
                         'Policy_Name': '',
                         'bd_type': '',
                         'host_routing': '',
                         'ep_clear': '',
                         'unk_mac': '',
                         'unk_mcast': '',
                         'v6unk_mcast': '',
                         'multi_dst': '',
                         'mcast_allow': '',
                         'ipv6_mcast': '',
                         'arp_flood': '',
                         'limit_learn': '',
                         'fvEpRetPol': '',
                         'unicast_route': '',
                         'intersight_l2': '',
                         'intersight_bum': '',
                         'optimize_wan': '',
                         'monEPGPol': '',
                         'ip_dp_learning': ''}
        optional_args = {'Alias': '',
                         'Description': '',
                         'Tags': '',
                         'Custom_MAC': '',
                         'Link_Local_IPv6': '',
                         'L3Out_Tenant': '',
                         'L3Out': '',
                         'dhcpRelayP': '',
                         'igmpIfPol': '',
                         'igmpSnoopPol': '',
                         'mldSnoopPol': '',
                         'ep_move': '',
                         'rtctrlProfile': '',
                         'ndIfPol': '',
                         'fhsBDPol': '',
                         'netflowMonitorPol': ''}

        # Get the BD Policies from the Network Policies Tab
        func = 'bd'
        row_count = ''
        count = countKeys(ws_net, func)
        var_dict = findVars(ws_net, func, rowcount, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('BD_Policy'):
                row_count = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}
                break

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
            validating.name_rule(row_num, ws, 'Bridge_Domain', templateVars['Bridge_Domain'])
            validating.name_rule(row_num, ws, 'VRF_Tenant', templateVars['VRF_Tenant'])
            validating.name_rule(row_num, ws, 'VRF', templateVars['VRF'])
            if not templateVars['L3Out_Tenant'] == None:
                validating.name_rule(row_num, ws, 'L3Out_Tenant', templateVars['L3Out_Tenant'])
            if not templateVars['L3Out'] == None:
                validating.name_rule(row_num, ws, 'L3Out', templateVars['L3Out'])
            validating.values(row_count, ws_net, 'bd_type', templateVars['bd_type'], ['fc', 'regular'])
            validating.values(row_count, ws_net, 'ep_clear', templateVars['ep_clear'], ['no', 'yes'])
            validating.values(row_count, ws_net, 'host_routing', templateVars['host_routing'], ['no', 'yes'])
            validating.values(row_count, ws_net, 'mcast_allow', templateVars['mcast_allow'], ['no', 'yes'])
            validating.values(row_count, ws_net, 'ipv6_mcast', templateVars['ipv6_mcast'], ['no', 'yes'])
            validating.values(row_count, ws_net, 'arp_flood', templateVars['arp_flood'], ['no', 'yes'])
            validating.values(row_count, ws_net, 'limit_learn', templateVars['limit_learn'], ['no', 'yes'])
            validating.values(row_count, ws_net, 'unicast_route', templateVars['unicast_route'], ['no', 'yes'])
            validating.values(row_count, ws_net, 'limit_learn', templateVars['limit_learn'], ['no', 'yes'])
            validating.values(row_count, ws_net, 'intersight_l2', templateVars['intersight_l2'], ['no', 'yes'])
            validating.values(row_count, ws_net, 'intersight_bum', templateVars['intersight_bum'], ['no', 'yes'])
            validating.values(row_count, ws_net, 'optimize_wan', templateVars['optimize_wan'], ['no', 'yes'])
            validating.values(row_count, ws_net, 'ip_dp_learning', templateVars['ip_dp_learning'], ['no', 'yes'])
            validating.values(row_count, ws_net, 'unk_mac', templateVars['unk_mac'], ['flood', 'proxy'])
            validating.values(row_count, ws_net, 'unk_mcast', templateVars['unk_mcast'], ['flood', 'opt-flood'])
            validating.values(row_count, ws_net, 'v6unk_mcast', templateVars['v6unk_mcast'], ['flood', 'opt-flood'])
            validating.values(row_count, ws_net, 'multi_dst', templateVars['multi_dst'], ['bd-flood', 'drop', 'encap-flood'])
            if not templateVars['ep_move'] == None:
                validating.values(row_count, ws_net, 'ep_move', templateVars['ep_move'], ['garp'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        if templateVars['dhcpRelayP'] == 'default':
            templateVars['dhcpRelayP'] = 'uni/tn-common/relayp-default'
        if templateVars['fhsBDPol'] == 'default':
            templateVars['fhsBDPol'] = 'uni/tn-common/bdpol-default'
        if templateVars['fvEpRetPol'] == 'default':
            templateVars['fvEpRetPol'] = 'uni/tn-common/epRPol-default'
        if templateVars['igmpIfPol'] == 'default':
            templateVars['igmpIfPol'] = 'uni/tn-common/igmpIfPol-default'
        if templateVars['igmpSnoopPol'] == 'default':
            templateVars['igmpSnoopPol'] = 'uni/tn-common/snPol-default'
        if templateVars['mldSnoopPol'] == 'default':
            templateVars['mldSnoopPol'] = 'uni/tn-common/mldsnoopPol-default'
        if templateVars['monEPGPol'] == 'default':
            templateVars['monEPGPol'] = 'uni/tn-common/monepg-default'
        if templateVars['ndIfPol'] == 'default':
            templateVars['ndIfPol'] = 'uni/tn-common/ndifpol-default'
        if templateVars['netflowMonitorPol'] == 'default':
            templateVars['netflowMonitorPol'] = 'uni/tn-common/monitorpol-default'

        if re.search('^(common|mgmt|infra)$', templateVars['Tenant']):
            templateVars['Tenant_Dn'] = 'data.aci_tenant.%s' % (templateVars['Tenant'])
        else:
            templateVars['Tenant_Dn'] = 'aci_tenant.%s' % (templateVars['Tenant'])

        if not templateVars['Tenant'] == templateVars['VRF_Tenant']:
            templateVars['vrfDn'] = 'data.aci_tenant.%s,data.aci_vrf.%s' % (templateVars['VRF_Tenant'], templateVars['VRF'])
            templateVars['rel_VRF'] = 'data.aci_vrf.%s' % (templateVars['VRF'])
            # Define the Template Source
            template_file = "data_vrf.template"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = 'data_tenant_%s_vrf_%s.tf' % (templateVars['VRF_Tenant'], templateVars['VRF'])
            dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
            process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

            # Process the template through the Sites
            templateVars['data_Tenant'] = templateVars['VRF_Tenant']
            template_file = "data_tenant.template"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = 'data_tenant_%s.tf' % (templateVars['VRF_Tenant'])
            dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
            process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        else:
            templateVars['vrfDn'] = 'aci_vrf.%s' % (templateVars['VRF'])
            templateVars['rel_VRF'] = 'aci_vrf.%s' % (templateVars['VRF'])

        if not (templateVars['L3Out'] == None and templateVars['Tenant'] == templateVars['L3Out_Tenant']):
            # Process the template through the Sites
            templateVars['data_Tenant'] = templateVars['L3Out_Tenant']
            template_file = "data_tenant.template"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = 'data_tenant_%s.tf' % (templateVars['L3Out_Tenant'])
            dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
            process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

            template_file = "data_l3out.template"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = 'data_tenant_%s_l3out_%s.tf' % (templateVars['L3Out_Tenant'], templateVars['L3Out'])
            dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
            process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

            # Create the Distinguished Name for the L3Out
            templateVars['L3Out_Dn'] = 'data.aci_l3_outside.%s_%s' % ((templateVars['L3Out_Tenant'], templateVars['L3Out']))

        elif not templateVars['L3Out'] == None:
            # Create the Distinguished Name for the L3Out
            templateVars['L3Out_Dn'] = 'aci_l3_outside.%s_%s' % ((templateVars['L3Out_Tenant'], templateVars['L3Out']))

        # Define the Template Source
        template_file = "bd.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'bd_%s.tf' % (templateVars['Bridge_Domain'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

       # Reset kwargs back to initial kwargs
        kwargs = initial_kwargs

        # Initialize the Class
        aci_lib_ref = 'Tenant_Policies'
        class_init = '%s(ws)' % (aci_lib_ref)

        # Create the Subnet if it Exists
        if not kwargs.get('Subnet') == None:
            eval("%s.%s(wb, ws, row_num, **kwargs)" % (class_init, 'add_subnet'))

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def add_epg(self, wb, ws, row_num, **kwargs):
        # Assign the kwargs to a initial var for each process
        initial_kwargs = kwargs

        # Initialize the Class
        aci_lib_ref = 'Tenant_Policies'
        class_init = '%s(ws)' % (aci_lib_ref)

        # Create the Application Profile if it Exists
        if not kwargs.get('App_Profile') == None:
            eval("%s.%s(wb, ws, row_num, **kwargs)" % (class_init, 'add_app'))

        # Reset kwargs back to initial kwargs
        kwargs = initial_kwargs

        # Open the Network Policies Worksheet
        ws_net = wb['Network Policies']
        rowcount = ws_net.max_row

        # Dicts for Bridge Domain required and optional args
        required_args = {'Site_Group': '',
                         'Tenant': '',
                         'Bridge_Domain': '',
                         'App_Profile': '',
                         'EPG': '',
                         'EPG_Policy': '',
                         'Policy_Name': '',
                         'is_attr_based': '',
                         'prio': '',
                         'pc_enf_pref': '',
                         'fwd_ctrl': '',
                         'pref_gr_memb': '',
                         'flood': '',
                         'match_t': '',
                         'monEPGPol': '',
                         'shutdown': '',
                         'has_mcast': ''}
        optional_args = {'Alias': '',
                         'EPG_Description': '',
                         'Tags': '',
                         'Physical_Domains': '',
                         'VMM_Domains': '',
                         'VLAN': '',
                         'PVLAN': '',
                         'EPG_to_AAEP': '',
                         'cons_vzBrCP': '',
                         'vzCPIf': '',
                         'Master_fvEPg': '',
                         'prov_vzBrCP': '',
                         'vzCtrctEPgCont': '',
                         'vzTaboo': '',
                         'exception_tag': '',
                         'qosCustomPol': '',
                         'qosDppPol': '',
                         'intra_vzBrCP': '',
                         'fhsTrustCtrlPol': '',
                         'vzGraphCont': '',
                         'FC_Domain': '',}

        # Get the EPG Policies from the Network Policies Tab
        func = 'epg'
        count = countKeys(ws_net, func)
        row_count = ''
        var_dict = findVars(ws_net, func, rowcount, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('EPG_Policy'):
                row_count = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
            validating.name_rule(row_num, ws, 'Bridge_Domain', templateVars['Bridge_Domain'])
            if not templateVars['VLAN'] == None:
                validating.vlans(row_num, ws, 'VLAN', templateVars['VLAN'])
            if not templateVars['PVLAN'] == None:
                validating.vlans(row_num, ws, 'PVLAN', templateVars['PVLAN'])
            validating.match_t(row_count, ws_net, 'match_t', templateVars['match_t'])
            validating.values(row_count, ws_net, 'fwd_ctrl', templateVars['fwd_ctrl'], ['none', 'proxy-arp'])
            validating.qos_priority(row_count, ws_net, 'prio', templateVars['prio'])
            validating.values(row_count, ws_net, 'flood', templateVars['flood'], ['disabled', 'enabled'])
            validating.values(row_count, ws_net, 'is_attr_based', templateVars['is_attr_based'], ['no', 'yes'])
            validating.values(row_count, ws_net, 'pc_enf_pref', templateVars['pc_enf_pref'], ['enforced', 'unenforced'])
            validating.values(row_count, ws_net, 'pref_gr_memb', templateVars['pref_gr_memb'], ['exclude', 'include'])
            validating.values(row_count, ws_net, 'shutdown', templateVars['shutdown'], ['no', 'yes'])
            if not templateVars['EPG_to_AAEP'] == None:
                validating.name_rule(row_num, ws, 'EPG_to_AAEP', templateVars['EPG_to_AAEP'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        if templateVars['cons_vzBrCP'] == 'default':
            templateVars['cons_vzBrCP'] = 'uni/tn-common/brc-default'
        if templateVars['prov_vzBrCP'] == 'default':
            templateVars['prov_vzBrCP'] = 'uni/tn-common/brc-default'
        if templateVars['vzCPIf'] == 'default':
            templateVars['vzCPIf'] = 'uni/tn-common/cif-default'
        # if templateVars['vzCtrctEPgCont'] == 'default':
        #     templateVars['vzCtrctEPgCont'] = 'uni/tn-common/mldsnoopPol-default'
        if templateVars['vzTaboo'] == 'default':
            templateVars['vzTaboo'] = 'uni/tn-common/taboo-default'
        if templateVars['qosCustomPol'] == 'default':
            templateVars['qosCustomPol'] = 'uni/tn-common/qoscustom-default'
        if templateVars['qosDppPol'] == 'default':
            templateVars['qosDppPol'] = 'uni/tn-common/qosdpppol-default'
        if templateVars['intra_vzBrCP'] == 'default':
            templateVars['intra_vzBrCP'] = 'uni/tn-common/brc-default'
        if templateVars['monEPGPol'] == 'default':
            templateVars['monEPGPol'] = 'uni/tn-common/monepg-default'
        if templateVars['fhsTrustCtrlPol'] == 'default':
            templateVars['fhsTrustCtrlPol'] = 'uni/tn-common/trustctrlpol-default'
        if templateVars['fwd_ctrl'] == 'none':
            templateVars['fwd_ctrl'] = None
        # if templateVars['vzGraphCont'] == 'default':
        #     templateVars['vzGraphCont'] = 'uni/tn-common/monitorpol-default'

        if re.search('^(common|mgmt|infra)$', templateVars['Tenant']):
            templateVars['Tenant_Dn'] = 'data.aci_tenant.%s' % (templateVars['Tenant'])
        else:
            templateVars['Tenant_Dn'] = 'aci_tenant.%s' % (templateVars['Tenant'])

        # Define the Template Source
        template_file = "epg.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'app_%s_epg_%s.tf' % (templateVars['App_Profile'], templateVars['EPG'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        if not templateVars['Physical_Domains'] == None:
            if ',' in templateVars['Physical_Domains']:
                splitx = templateVars['Physical_Domains'].split(',')
                for x in splitx:
                    templateVars['Domain'] = 'phys-%s' % (x)
                    # Define the Template Source
                    template_file = "domain_to_epg.template"
                    template = self.templateEnv.get_template(template_file)

                    # Process the template through the Sites
                    dest_file = 'epg_%s_%s.tf' % (templateVars['App_Profile'], templateVars['EPG'])
                    dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
                    process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)
            else:
                templateVars['Domain'] = 'phys-%s' % (templateVars['Physical_Domains'])
                # Define the Template Source
                template_file = "domain_to_epg.template"
                template = self.templateEnv.get_template(template_file)

                # Process the template through the Sites
                dest_file = 'epg_%s_%s.tf' % (templateVars['App_Profile'], templateVars['EPG'])
                dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
                process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

        if not templateVars['VMM_Domains'] == None:
            if ',' in templateVars['VMM_Domains']:
                splitx = templateVars['VMM_Domains'].split(',')
                for x in splitx:
                    templateVars['Domain'] = 'vmm-%s' % (x)
                    # Define the Template Source
                    template_file = "domain_to_epg.template"
                    template = self.templateEnv.get_template(template_file)

                    # Process the template through the Sites
                    dest_file = 'app_%s_epg_%s.tf' % (templateVars['App_Profile'], templateVars['EPG'])
                    dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
                    process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)
            else:
                templateVars['Domain'] = 'vmm-%s' % (templateVars['VMM_Domains'])
                # Define the Template Source
                template_file = "domain_to_epg.template"
                template = self.templateEnv.get_template(template_file)

                # Process the template through the Sites
                dest_file = 'app_%s_epg_%s.tf' % (templateVars['App_Profile'], templateVars['EPG'])
                dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
                process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

        if not templateVars['VLAN'] == None:
            # Define the Template Source
            template_file = "static_path.template"
            template = self.templateEnv.get_template(template_file)

            dest_file = 'app_%s_epg_%s.tf' % (templateVars['App_Profile'], templateVars['EPG'])
            dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
            process_workbook(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

        if not templateVars['EPG_to_AAEP'] == None:
            if re.search(',', templateVars['EPG_to_AAEP']):
                # Define the Template Source
                aep_list = templateVars['EPG_to_AAEP'].split(',')
                for aep in aep_list:
                    templateVars['AAEP'] = aep

                    # Define the Template Source
                    template_file = "policies_global_aep_generic.template"
                    template = self.templateEnv.get_template(template_file)

                    # Process the template through the Sites
                    dest_file = 'policies_global_aep_%s_generic.tf' % (templateVars['AAEP'])
                    dest_dir = 'Access'
                    process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

                    # Define the Template Source
                    template_file = "data_access_generic.template"
                    template = self.templateEnv.get_template(template_file)

                    # Process the template through the Sites
                    dest_file = 'data_aep_%s.tf' % (templateVars['AAEP'])
                    dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
                    process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

                    # Define the Template Source
                    template_file = "epgs_using_function.template"
                    template = self.templateEnv.get_template(template_file)

                    # Process the template through the Sites
                    dest_file = 'app_%s_epg_%s_aep_%s.tf' % (templateVars['App_Profile'], templateVars['EPG'], templateVars['AAEP'])
                    dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
                    process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

        # dest_file = 'epg_%s_%s_static_bindings.tf' % (templateVars['App_Profile'], templateVars['EPG'])
        # dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        # create_file(wb, ws, row_num, 'w', dest_dir, dest_file, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def add_subnet(self, wb, ws, row_num, **kwargs):
        # Open the Network Policies Worksheet
        ws_net = wb['Network Policies']
        rowcount = ws_net.max_row

        # Dicts for Subnet required and optional args
        required_args = {'Site_Group': '',
                         'Tenant': '',
                         'Bridge_Domain': '',
                         'Subnet': '',
                         'Subnet_Policy': '',
                         'Policy_Name': '',
                         'virtual': '',
                         'preferred': '',
                         'scope': '',
                         'nd': '',
                         'no-default-gateway': '',
                         'querier': ''}
        optional_args = {'Subnet_Description': '',
                         'l3extOut': '',
                         'rtctrlProfile': '',
                         'ndPfxPol': ''}

        # Get the Subnet Policies from the Network Policies Tab
        func = 'subnet'
        count = countKeys(ws_net, func)
        row_count = ''
        var_dict = findVars(ws_net, func, rowcount, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('Subnet_Policy'):
                row_count = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}
                break

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.ip_address(row_num, ws, 'Subnet', templateVars['Subnet'])
            validating.values(row_count, ws_net, 'nd', templateVars['nd'], ['no', 'yes'])
            validating.values(row_count, ws_net, 'no-default-gateway', templateVars['no-default-gateway'], ['no', 'yes'])
            validating.values(row_count, ws_net, 'preferred', templateVars['preferred'], ['no', 'yes'])
            validating.values(row_count, ws_net, 'querier', templateVars['querier'], ['no', 'yes'])
            validating.values(row_count, ws_net, 'virtual', templateVars['virtual'], ['no', 'yes'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        if templateVars['l3extOut'] == 'default':
            templateVars['l3extOut'] = 'uni/tn-common/out-default'
        if templateVars['ndPfxPol'] == 'default':
            templateVars['ndPfxPol'] = 'uni/tn-common/ndpfxpol-default'

        # Create ctrl templateVars
        ctrl = ''
        if templateVars['nd'] == 'yes':
            ctrl = ctrl + '"nd"'
        if templateVars['no-default-gateway'] == 'yes':
            ctrl = ctrl + ', ' + '"no-default-gateway"'
        if templateVars['querier'] == 'yes':
            ctrl = ctrl + ', ' + '"querier"'

        if ctrl == '':
            templateVars['ctrl'] = '["unspecified"]'
        else:
            templateVars['ctrl'] = '[%s]' % (ctrl)

        # Modify scope templateVars
        if re.search('^(private|public|shared)$', templateVars['scope']):
            templateVars['scope'] = '"%s"' % (templateVars['scope'])
        elif re.search('^(private|public)\\-shared$', templateVars['scope']):
            x = templateVars['scope'].split('-')
            templateVars['scope'] = '"%s", "%s"' & (x[0], x[1])

        # As period and colon are not allowed in description need to modify Subnet to work for description and filename
        if ':' in templateVars['Subnet']:
            network = "%s" % (ipaddress.IPv6Network(templateVars['Subnet'], strict=False))
            templateVars['Subnet_'] = network
            templateVars['Subnet_'] = templateVars['Subnet_'].replace(':', '-')
            templateVars['Subnet_'] = templateVars['Subnet_'].replace('/', '_')
        else:
            network = "%s" % (ipaddress.IPv4Network(templateVars['Subnet'], strict=False))
            templateVars['Subnet_'] = network
            templateVars['Subnet_'] = templateVars['Subnet_'].replace('.', '-')
            templateVars['Subnet_'] = templateVars['Subnet_'].replace('/', '_')

        # Define the Template Source
        template_file = "subnet.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'bd_%s_subnet_%s.tf' % (templateVars['Bridge_Domain'], templateVars['Subnet_'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def add_tenant(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Tenant': ''}
        optional_args = {'Alias': '',
                         'Description': '',
                         'Tags': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "tenant.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'tenant_%s.tf' % (templateVars['Tenant'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def add_vrf(self, wb, ws, row_num, **kwargs):
        # Open the Network Policies Worksheet
        ws_net = wb['Network Policies']
        rowcount = ws_net.max_row

        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Tenant': '',
                         'VRF': '',
                         'VRF_Policy': '',
                         'Policy_Name': '',
                         'pc_enf_pref': '',
                         'pc_enf_dir': '',
                         'bd_enforce': '',
                         'enf_type': '',
                         'fvEpRetPol': '',
                         'monEPGPol': '',
                         'ip_dp_learning': '',
                         'knw_mcast_act': ''}
        optional_args = {'Alias': '',
                         'Description': '',
                         'Tags': '',
                         'cons_vzBrCP': '',
                         'vzCPIf': '',
                         'prov_vzBrCP': '',
                         'bgpCtxPol': '',
                         'bgpCtxAfPol': '',
                         'ospfCtxPol': '',
                         'ospfCtxAfPol': '',
                         'eigrpCtxAfPol': '',
                         'l3extRouteTagPol': '',
                         'l3extVrfValidationPol': ''}


        # Get the VRF Policies from the Network Policies Tab
        func = 'VRF'
        count = countKeys(ws_net, func)
        row_count = ''
        var_dict = findVars(ws_net, func, rowcount, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('VRF_Policy'):
                row_count = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
            validating.name_rule(row_num, ws, 'VRF', templateVars['VRF'])
            validating.values(row_count, ws_net, 'bd_enforce', templateVars['bd_enforce'], ['no', 'yes'])
            validating.values(row_count, ws_net, 'ip_dp_learning', templateVars['ip_dp_learning'], ['disabled', 'enabled'])
            validating.values(row_count, ws_net, 'knw_mcast_act', templateVars['knw_mcast_act'], ['deny', 'permit'])
            validating.values(row_count, ws_net, 'pc_enf_dir', templateVars['pc_enf_dir'], ['egress', 'ingress'])
            validating.values(row_count, ws_net, 'pc_enf_pref', templateVars['pc_enf_pref'], ['enforced', 'unenforced'])
            validating.values(row_count, ws_net, 'enf_type', templateVars['enf_type'], ['contract', 'pref_grp', 'vzAny'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        if templateVars['cons_vzBrCP'] == 'default':
            templateVars['cons_vzBrCP'] = 'uni/tn-common/brc-default'
        if templateVars['prov_vzBrCP'] == 'default':
            templateVars['prov_vzBrCP'] = 'uni/tn-common/brc-default'
        if templateVars['vzCPIf'] == 'default':
            templateVars['vzCPIf'] = 'uni/tn-common/cif-default'
        if templateVars['bgpCtxPol'] == 'default':
            templateVars['bgpCtxPol'] = 'uni/tn-common/bgpCtxP-default'
        if templateVars['bgpCtxAfPol'] == 'default':
            templateVars['bgpCtxAfPol'] = 'uni/tn-common/bgpCtxAfP-default'
        if templateVars['eigrpCtxAfPol'] == 'default':
            templateVars['eigrpCtxAfPol'] = 'uni/tn-common/eigrpCtxAfP-default'
        if templateVars['ospfCtxPol'] == 'default':
            templateVars['ospfCtxPol'] = 'uni/tn-common/ospfCtxP-default'
        if templateVars['ospfCtxAfPol'] == 'default':
            templateVars['ospfCtxAfPol'] = 'uni/tn-common/ospfCtxP-default'
        if templateVars['fvEpRetPol'] == 'default':
            templateVars['fvEpRetPol'] = 'uni/tn-common/epRPol-default'
        if templateVars['monEPGPol'] == 'default':
            templateVars['monEPGPol'] = 'uni/tn-common/monepg-default'
        if templateVars['l3extVrfValidationPol'] == 'default':
            templateVars['l3extVrfValidationPol'] = 'uni/tn-common/vrfvalidationpol-default'

        if re.search('^(common|mgmt|infra)$', templateVars['Tenant']):
            templateVars['Tenant_resource'] = 'data.aci_tenant.%s' % (templateVars['Tenant'])
        else:
            templateVars['Tenant_resource'] = 'aci_tenant.%s' % (templateVars['Tenant'])

        # Define the Template Source
        template_file = "vrf.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'vrf_%s.tf' % (templateVars['VRF'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        if templateVars['enf_type'] == 'pref_grp':
            # Define the Template Source
            template_file = "pref_grp.template"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = 'vrf_%s.tf' % (templateVars['VRF'])
            dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
            process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

        elif templateVars['enf_type'] == 'vzAny':
            # Define the Template Source
            template_file = "vzAny.template"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = 'vrf_%s.tf' % (templateVars['VRF'])
            dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
            process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        template_file = "snmp_ctx.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'vrf_%s.tf' % (templateVars['VRF'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def ctx_comm(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Tenant': '',
                         'VRF': '',
                         'Ctx_Community': ''}
        optional_args = {'Description': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
            validating.name_rule(row_num, ws, 'VRF', templateVars['VRF'])
            validating.snmp_string(row_num, ws, 'Ctx_Community', templateVars['Ctx_Community'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "snmp_ctx_community.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'vrf_%s.tf' % (templateVars['VRF'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def contract_add(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Tenant': '',
                         'Contract': '',
                         'Scope': '',
                         'QoS_Class': '',
                         'Target_DSCP': ''}
        optional_args = {'Description': '',
                         'Alias': '',
                         'Tags': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.dscp(row_num, ws, 'Target_DSCP', templateVars['Target_DSCP'])
            validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
            validating.name_rule(row_num, ws, 'Contract', templateVars['Contract'])
            validating.qos_priority(row_num, ws, 'QoS_Class', templateVars['QoS_Class'])
            validating.values(row_num, ws, 'Scope', templateVars['Scope'], ['application-profile', 'context', 'global', 'tenant'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "contract.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'contract_%s.tf' % (templateVars['Contract'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def contract_mgmt(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Tenant': '',
                         'Contract': '',
                         'Scope': '',
                         'QoS_Class': '',
                         'Target_DSCP': '',
                         'Subject': '',
                         'Reverse_Filter_Ports': '',
                         'Filter_1': ''}
        optional_args = {'Description': '',
                         'Subject_Descr': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.dscp(row_num, ws, 'Target_DSCP', templateVars['Target_DSCP'])
            validating.name_rule(row_num, ws, 'Contract', templateVars['Contract'])
            validating.name_rule(row_num, ws, 'Filter_1', templateVars['Filter_1'])
            validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
            validating.name_rule(row_num, ws, 'Subject', templateVars['Subject'])
            validating.qos_priority(row_num, ws, 'QoS_Class', templateVars['QoS_Class'])
            validating.values(row_num, ws, 'Reverse_Filter_Ports', templateVars['Reverse_Filter_Ports'], ['no', 'yes'])
            validating.values(row_num, ws, 'Scope', templateVars['Scope'], ['application-profile', 'context', 'global', 'tenant'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
            if not templateVars['Subject_Descr'] == None:
                validating.description(row_num, ws, 'Subject_Descr', templateVars['Subject_Descr'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "contract_oob_mgmt.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'contract_oob_mgmt_%s.tf' % (templateVars['Contract'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        template_file = "contract_subj_oob.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'contract_oob_mgmt_%s.tf' % (templateVars['Contract'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        template_file = "contract_filter_oob.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        templateVars['Filter'] == templateVars['Filter_1']
        dest_file = 'contract_oob_mgmt_%s.tf' % (templateVars['Contract'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def filter_add(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Tenant': '',
                         'Filter': ''}
        optional_args = {'Description': '',
                         'Alias': '',
                         'Tags': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Filter', templateVars['Filter'])
            validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "contract_filter.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'contract_filter_%s.tf' % (templateVars['Filter'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def filter_entry(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Tenant': '',
                         'Filter': '',
                         'Filter_Entry': '',
                         'EtherType': '',
                         'IP_Protocol': '',
                         'ARP_Flag': '',
                         'ICMPv4_Type': '',
                         'ICMPv6_Type': '',
                         'Match_DSCP': '',
                         'Match_Only_Frags': '',
                         'Source_From': '',
                         'Source_To': '',
                         'Dest_From': '',
                         'Dest_To': '',
                         'Stateful': '',
                         'TCP_Session_Rules': ''}
        optional_args = {'Description': '',
                         'Alias': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.dscp(row_num, ws, 'Match_DSCP', templateVars['Match_DSCP'])
            validating.filter_ports(row_num, ws, 'Source_From', templateVars['Source_From'])
            validating.filter_ports(row_num, ws, 'Source_To', templateVars['Source_To'])
            validating.filter_ports(row_num, ws, 'Dest_From', templateVars['Dest_From'])
            validating.filter_ports(row_num, ws, 'Dest_To', templateVars['Dest_To'])
            validating.name_rule(row_num, ws, 'Filter', templateVars['Filter'])
            validating.name_rule(row_num, ws, 'Filter_Entry', templateVars['Filter_Entry'])
            validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
            validating.values(row_num, ws, 'Match_Only_Frags', templateVars['Match_Only_Frags'], ['no', 'yes'])
            validating.values(row_num, ws, 'Stateful', templateVars['Stateful'], ['no', 'yes'])
            validating.values(row_num, ws, 'EtherType', templateVars['EtherType'], ['arp', 'fcoe', 'ip', 'ipv4', 'ipv6', 'trill', 'mac_security', 'mpls_ucast', 'unspecified'])
            validating.values(row_num, ws, 'IP_Protocol', templateVars['IP_Protocol'], ['egp', 'eigrp', 'igp', 'icmp', 'icmpv6', 'igmp', 'l2tp', 'ospfigp', 'pim', 'tcp', 'udp', 'unspecified'])
            validating.values(row_num, ws, 'ARP_Flag', templateVars['ARP_Flag'], ['req', 'reply', 'unspecified'])
            validating.values(row_num, ws, 'ICMPv4_Type', templateVars['ICMPv4_Type'], ['dst-unreach', 'echo', 'echo-rep', 'src-quench', 'time-exceeded', 'unspecified'])
            validating.values(row_num, ws, 'ICMPv6_Type', templateVars['ICMPv6_Type'], ['dst-unreach', 'echo-req', 'echo-rep', 'nbr-solicit', 'nbr-advert', 'redirect', 'time-exceeded', 'unspecified'])
            validating.values(row_num, ws, 'TCP_Session_Rules', templateVars['TCP_Session_Rules'], ['ack', 'est', 'fin', 'rst', 'syn', 'unspecified'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "contract_filter_entry.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'contract_filter_%s.tf' % (templateVars['Filter'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def mgmt_epg(self, wb, ws, row_num, **kwargs):
        # Dicts for Bridge Domain required and optional args
        required_args = {'Site_Group': '',
                         'Type': '',
                         'EPG': '',
                         'EPG': '',
                         'QoS_Class': ''}
        optional_args = {'Tags': '',
                         'VLAN': '',
                         'Bridge_Domain': '',
                         'Contract_Tenant': '',
                         'consumed_Contract': '',
                         'provided_Contract': '',}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'EPG', templateVars['EPG'])
            validating.qos_priority(row_num, ws, 'QoS_Class', templateVars['QoS_Class'])
            validating.values(row_num, ws, 'Type', templateVars['Type'], ['In-Band EPG', 'Out-of-Band EPG'])
            if templateVars['Type'] == 'In-Band EPG':
                validating.vlans(row_num, ws, 'VLAN', templateVars['VLAN'])
                validating.name_rule(row_num, ws, 'Bridge_Domain', templateVars['Bridge_Domain'])
            if not templateVars['Contract_Tenant'] == None:
                validating.name_rule(row_num, ws, 'Contract_Tenant', templateVars['Contract_Tenant'])
                validating.not_empty(row_num, ws, 'provided_Contract', templateVars['provided_Contract'])
                if templateVars['Type'] == 'In-Band EPG':
                    validating.not_empty(row_num, ws, 'consumed_Contract', templateVars['consumed_Contract'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        if templateVars['Type'] == 'In-Band EPG':
            # Define the Template Source and Destination File
            template_file = "epg_inband.template"
            template = self.templateEnv.get_template(template_file)
            dest_file = 'mgmt_epg_%s_%s.tf' % ('inband', templateVars['EPG'])
        else:
            # Define the Template Source and Destination File
            template_file = "epg_oob.template"
            template = self.templateEnv.get_template(template_file)
            dest_file = 'mgmt_epg_%s_%s.tf' % ('oob', templateVars['EPG'])

        # Process the template through the Sites
        dest_dir = 'Tenant_mgmt'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        if not templateVars['Contract_Tenant'] == None:
            if templateVars['Type'] == 'In-Band EPG':
                # Define the Template Source
                template_file = "epg_to_contract_inb.template"
                template = self.templateEnv.get_template(template_file)

                # Process the template through the Sites
                dest_file = 'epg_inband_%s_contracts.tf' % (templateVars['EPG'])
                dest_dir = 'Tenant_mgmt'
                process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)
            else:
                # Define the Template Source
                template_file = "epg_to_contract_oob.template"
                template = self.templateEnv.get_template(template_file)

                # Process the template through the Sites
                dest_file = 'epg_oob_%s_contract_%s.tf' % (templateVars['EPG'], templateVars['provided_Contract'])
                dest_dir = 'Tenant_mgmt'
                process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source and Destination File
        template_file = "var_mgmt.template"
        template = self.templateEnv.get_template(template_file)

        if templateVars['Type'] == 'In-Band EPG':
            templateVars['var_name'] = 'var_inb'
            templateVars['var_value'] = 'uni/tn-mgmt/mgmtp-default/inb-{{EPG}}'
            dest_file = 'var_mgmt_%s_epg.tf' % ('inb')
        else:
            templateVars['var_name'] = 'var_oob'
            templateVars['var_value'] = 'uni/tn-mgmt/mgmtp-default/inb-{{EPG}}'
            dest_file = 'var_mgmt_%s_epg.tf' % ('oob')

        # Process the template through the Sites
        dest_dir = 'Access'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)
        dest_dir = 'Admin'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)
        dest_dir = 'Fabric'
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def subject_add(self, wb, ws, row_num, **kwargs):
        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Tenant': '',
                         'Contract': '',
                         'Subject': '',
                         'Reverse_Filter_Ports': '',
                         'QoS_Class': '',
                         'Target_DSCP': '',
                         'Filter_1': ''}
        optional_args = {'Description': '',
                         'Alias': '',
                         'Tags': '',
                         'Filter_2': '',
                         'Filter_3': '',
                         'Filter_4': '',
                         'Filter_5': '',
                         'Filter_6': '',
                         'Filter_7': '',
                         'Filter_8': '',
                         'Filter_9': ''}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.dscp(row_num, ws, 'Target_DSCP', templateVars['Target_DSCP'])
            validating.name_rule(row_num, ws, 'Contract', templateVars['Contract'])
            validating.name_rule(row_num, ws, 'Filter_1', templateVars['Filter_1'])
            validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
            validating.name_rule(row_num, ws, 'Subject', templateVars['Subject'])
            validating.qos_priority(row_num, ws, 'QoS_Class', templateVars['QoS_Class'])
            validating.values(row_num, ws, 'Reverse_Filter_Ports', templateVars['Reverse_Filter_Ports'], ['no', 'yes'])
            if not templateVars['Alias'] == None:
                validating.name_rule(row_num, ws, 'Alias', templateVars['Alias'])
            if not templateVars['Description'] == None:
                validating.description(row_num, ws, 'Description', templateVars['Description'])
            if not templateVars['Filter_2'] == None:
                validating.name_rule(row_num, ws, 'Filter_2', templateVars['Filter_2'])
            if not templateVars['Filter_3'] == None:
                validating.name_rule(row_num, ws, 'Filter_3', templateVars['Filter_3'])
            if not templateVars['Filter_4'] == None:
                validating.name_rule(row_num, ws, 'Filter_4', templateVars['Filter_4'])
            if not templateVars['Filter_5'] == None:
                validating.name_rule(row_num, ws, 'Filter_5', templateVars['Filter_5'])
            if not templateVars['Filter_6'] == None:
                validating.name_rule(row_num, ws, 'Filter_6', templateVars['Filter_6'])
            if not templateVars['Filter_7'] == None:
                validating.name_rule(row_num, ws, 'Filter_7', templateVars['Filter_7'])
            if not templateVars['Filter_8'] == None:
                validating.name_rule(row_num, ws, 'Filter_8', templateVars['Filter_8'])
            if not templateVars['Filter_9'] == None:
                validating.name_rule(row_num, ws, 'Filter_9', templateVars['Filter_9'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Define the Template Source
        template_file = "contract_subject.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'contract_%s_subj_%s.tf' % (templateVars['Contract'], templateVars['Subject'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'a+', dest_dir, dest_file, template, **templateVars)

# Terraform ACI Provider - Tenants Policies
# Class must be instantiated with Variables
class VMM_Policies(object):
    def __init__(self, ws):
        self.templateLoader = jinja2.FileSystemLoader(
            searchpath=(aci_template_path + 'Tenant_Policies/'))
        self.templateEnv = jinja2.Environment(loader=self.templateLoader)

    # Method must be called with the following kwargs.
    # Please Refer to the Input Spreadsheet "Notes" in the relevant column headers
    # for Detailed information on the Arguments used by this Method.
    def add_vmm(self, wb, ws, row_num, **kwargs):
        # Open the Network Policies Worksheet
        ws_net = wb['Network Policies']
        rowcount = ws_net.max_row

        # Dicts for required and optional args
        required_args = {'Site_Group': '',
                         'Tenant': '',
                         'VRF': '',
                         'Name': '',
                         'L3Out_Policy': '',
                         'L3_Domain': '',
                         'Ext_EPG': '',
                         'Ext_EPG_Policy': '',
                         'Subnet': '',
                         'Ext_Subnet_Policy': '',
                         'target_dscp': '',
                         'enforce_rtctrl': '',
                         'prio': '',
                         'epg_target_dscp': '',
                         'pref_gr_memb': '',
                         'match_t': '',
                         'flood': '',
                         'export-rtctrl': '',
                         'import-rtctrl': '',
                         'import-security': '',
                         'shared-security': '',
                         'shared-rtctrl': '',
                         'agg-export': '',
                         'agg-import': '',
                         'agg-shared': ''}
        optional_args = {'Description': '',
                         'EPG_Description': '',
                         'annotation': '',
                         'name_alias': '',
                         'leak_rtctrlProfile': '',
                         'damp_rtctrlProfile': '',
                         'fvBDPublicSubnetHolder': '',
                         'epg_annotation': '',
                         'epg_name_alias': '',
                         'cons_vzBrCP': '',
                         'vzCPIf': '',
                         'Master_fvEPg': '',
                         'prov_vzBrCP': '',
                         'vzTaboo': '',
                         'exception_tag': '',
                         'rtctrlProfile': '',
                         'sub_annotation': '',
                         'sub_name_alias': '',
                         'sub_rtctrlProfile': '',
                         'rtsumARtSummPol': ''}


        # Get the L3Out Policies from the Network Policies Tab
        func = 'L3Out_Policy'
        count = countKeys(ws_net, func)
        l3_count = ''
        var_dict = findVars(ws_net, func, rowcount, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('L3Out_Policy'):
                l3_count = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}

        func = 'Ext_EPG_Policy'
        count = countKeys(ws_net, func)
        epg_count = ''
        var_dict = findVars(ws_net, func, rowcount, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('Ext_EPG_Policy'):
                epg_count = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}

        func = 'Ext_Subnet_Policy'
        count = countKeys(ws_net, func)
        sub_count = ''
        var_dict = findVars(ws_net, func, rowcount, count)
        for pos in var_dict:
            if var_dict[pos].get('Policy_Name') == kwargs.get('Ext_Subnet_Policy'):
                sub_count = var_dict[pos]['row']
                del var_dict[pos]['row']
                kwargs = {**kwargs, **var_dict[pos]}

        # Validate inputs, return dict of template vars
        templateVars = process_kwargs(required_args, optional_args, **kwargs)

        try:
            # Validate Required Arguments
            validating.site_group(row_num, ws, 'Site_Group', templateVars['Site_Group'])
            validating.name_rule(row_num, ws, 'Tenant', templateVars['Tenant'])
            validating.name_rule(row_num, ws, 'VRF', templateVars['VRF'])
            validating.name_rule(row_num, ws, 'Name', templateVars['Name'])
            if not templateVars['Subnet'] == None:
                if re.search(',', templateVars['Subnet']):
                    sx = templateVars['Subnet'].split(',')
                    for x in sx:
                        validating.ip_address(row_num, ws, 'Subnet', x)
            validating.dscp(l3_count, ws_net, 'target_dscp', templateVars['target_dscp'])
            validating.dscp(epg_count, ws_net, 'epg_target_dscp', templateVars['epg_target_dscp'])
            validating.match_t(epg_count, ws_net, 'match_t', templateVars['match_t'])
            validating.qos_priority(epg_count, ws_net, 'prio', templateVars['prio'])
            validating.values(sub_count, ws_net, 'agg-export', templateVars['agg-export'], ['no', 'yes'])
            validating.values(sub_count, ws_net, 'agg-import', templateVars['agg-import'], ['no', 'yes'])
            validating.values(sub_count, ws_net, 'agg-shared', templateVars['agg-shared'], ['no', 'yes'])
            validating.values(sub_count, ws_net, 'export-rtctrl', templateVars['export-rtctrl'], ['no', 'yes'])
            validating.values(sub_count, ws_net, 'import-rtctrl', templateVars['import-rtctrl'], ['no', 'yes'])
            validating.values(sub_count, ws_net, 'import-security', templateVars['import-security'], ['no', 'yes'])
            validating.values(sub_count, ws_net, 'shared-security', templateVars['shared-security'], ['no', 'yes'])
            validating.values(sub_count, ws_net, 'shared-rtctrl', templateVars['shared-rtctrl'], ['no', 'yes'])
            validating.values(l3_count, ws_net, 'enforce_rtctrl', templateVars['enforce_rtctrl'], ['export', 'export-import'])
            validating.values(epg_count, ws_net, 'flood', templateVars['flood'], ['disabled', 'enabled'])
            validating.values(epg_count, ws_net, 'pref_gr_memb', templateVars['pref_gr_memb'], ['exclude', 'include'])
        except Exception as err:
            Error_Return = '%s\nError on Worksheet %s Row %s.  Please verify Input Information.' % (SystemExit(err), ws, row_num)
            raise ErrException(Error_Return)

        # Create aggregate templateVars
        aggregate = ''
        if templateVars['agg-export'] == 'yes':
            aggregate = aggregate + '"export-rtctrl"'
        if templateVars['agg-import'] == 'yes':
            aggregate = aggregate + ', ' + '"import-rtctrl"'
        if templateVars['agg-shared'] == 'yes':
            aggregate = aggregate + ', ' + '"shared-rtctrl"'

        else:
            templateVars['aggregate'] = '[%s]' % (aggregate)

        # Create scope templateVars
        scope = ''
        if templateVars['export-rtctrl'] == 'yes':
            scope = scope + '"export-rtctrl"'
        if templateVars['import-rtctrl'] == 'yes':
            scope = scope + ', ' + '"import-rtctrl"'
        if templateVars['import-security'] == 'yes':
            scope = scope + ', ' + '"import-security"'
        if templateVars['shared-security'] == 'yes':
            scope = scope + ', ' + '"shared-security"'
        if templateVars['shared-rtctrl'] == 'yes':
            scope = scope + ', ' + '"shared-rtctrl"'

        else:
            templateVars['scope'] = '[%s]' % (scope)

        # Define the Template Source
        template_file = "l3out.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'l3out_%s.tf' % (templateVars['Name'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        # Define the Template Source
        template_file = "ext_epg.template"
        template = self.templateEnv.get_template(template_file)

        # Process the template through the Sites
        dest_file = 'l3out_%s_epg_%s.tf' % (templateVars['Name'], templateVars['Ext_EPG'])
        dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
        process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

        if re.search(',', templateVars['Subnet']):
            sx = templateVars['Subnet'].split(',')
            for x in sx:
                templateVars['Subnet'] = x
                templateVars['Subnet_'] = x.replace('.', '-')
                templateVars['Subnet_'] = x.replace('/', '_')

                # Define the Template Source
                template_file = "ext_subnet.template"
                template = self.templateEnv.get_template(template_file)

                # Process the template through the Sites
                dest_file = 'l3out_%s_epg_%s_subnet_%s.tf' % (templateVars['Name'], templateVars['Ext_EPG'], templateVars['Subnet'])
                dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
                process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)
        else:
            templateVars['Subnet_'] = templateVars['Subnet'].replace('.', '-')
            templateVars['Subnet_'] = templateVars['Subnet'].replace('/', '_')

            # Define the Template Source
            template_file = "ext_subnet.template"
            template = self.templateEnv.get_template(template_file)

            # Process the template through the Sites
            dest_file = 'l3out_%s_epg_%s_subnet_%s.tf' % (templateVars['Name'], templateVars['Ext_EPG'], templateVars['Subnet'])
            dest_dir = 'Tenant_%s' % (templateVars['Tenant'])
            process_method(wb, ws, row_num, 'w', dest_dir, dest_file, template, **templateVars)

# Function to Create Destination Files
def create_file(wb, ws, row_num, wr_method, dest_dir, dest_file, **templateVars):
    if re.search('Grp_[A-F]', templateVars['Site_Group']):
        Group_ID = '%s' % (templateVars['Site_Group'])
        site_group = ast.literal_eval(os.environ[Group_ID])
        for x in range(1, 13):
            sitex = 'Site_%s' % (x)
            if not site_group[sitex] == None:
                Site_ID = 'Site_ID_%s' % (site_group[sitex])
                site_dict = ast.literal_eval(os.environ[Site_ID])

                # Create templateVars for Site_Name and APIC_URL
                templateVars['Site_Name'] = site_dict.get('Site_Name')
                templateVars['APIC_URL'] = site_dict.get('APIC_URL')

                # Create Terraform file from Template
                tf_file = './ACI/%s/%s/%s' % (templateVars['Site_Name'], dest_dir, dest_file)
                wr_file = open(tf_file, wr_method)
                wr_file.close()

    elif re.search(r'\d+', templateVars['Site_Group']):
        Site_ID = 'Site_ID_%s' % (templateVars['Site_Group'])
        site_dict = ast.literal_eval(os.environ[Site_ID])

        # Create templateVars for Site_Name and APIC_URL
        templateVars['Site_Name'] = site_dict.get('Site_Name')
        templateVars['APIC_URL'] = site_dict.get('APIC_URL')

        # Create Terraform file from Template
        tf_file = './ACI/%s/%s/%s' % (templateVars['Site_Name'], dest_dir, dest_file)
        wr_file = open(tf_file, wr_method)
        wr_file.close()
    else:
        print(f"\n-----------------------------------------------------------------------------\n")
        print(f"   Error on Worksheet {ws.title}, Row {row_num} Site_Group, value {templateVars['Site_Group']}.")
        print(f"   Unable to Determine if this is a Single or Group of Site(s).  Exiting....")
        print(f"\n-----------------------------------------------------------------------------\n")
        exit()

# Function to Create Terraform Files
def create_tf_file(wr_method, dest_dir, dest_file, template, **templateVars):
    # Make sure the Destination Directory Exists
    dest_dir = './ACI/%s/%s' % (templateVars['Site_Name'], dest_dir)
    if not os.path.isdir(dest_dir):
        mk_dir = 'mkdir -p %s' % (dest_dir)
        os.system(mk_dir)
    # Create File for the Template in the Destination Folder
    tf_file = '%s/%s' % (dest_dir, dest_file)
    wr_file = open(tf_file, wr_method)

    # Render Payload and Write to File
    payload = template.render(templateVars)
    wr_file.write(payload + '\n\n')
    wr_file.close()

# Function to Create Interface Selectors
def create_selector(ws_sw, ws_sw_row_count, **templateVars):
    print(templateVars['port_count'])
    Port_Selector = ''
    for port in range(1, int(templateVars['port_count']) + 1):
        if port < 10:
            Port_Selector = 'Eth%s-0%s' % (templateVars['module'], port)
        elif port < 100:
            Port_Selector = 'Eth%s-%s' % (templateVars['module'], port)
        elif port > 99:
            Port_Selector = 'Eth%s_%s' % (templateVars['module'], port)
        modp = '%s/%s' % (templateVars['module'],port)
        # Copy the Port Selector to the Worksheet
        data = ['intf_selector',templateVars['Pod_ID'],templateVars['Node_ID'],templateVars['Name'],Port_Selector,modp,'','','','','','']
        ws_sw.append(data)
        rc = '%s:%s' % (ws_sw_row_count, ws_sw_row_count)
        for cell in ws_sw[rc]:
            if ws_sw_row_count % 2 == 0:
                cell.style = 'ws_odd'
            else:
                cell.style = 'ws_even'
        dv1_cell = 'A%s' % (ws_sw_row_count)
        dv2_cell = 'H%s' % (ws_sw_row_count)
        dv3_cell = 'I%s' % (ws_sw_row_count)
        templateVars['dv1'].add(dv1_cell)
        templateVars['dv2'].add(dv2_cell)
        templateVars['dv3'].add(dv3_cell)
        ws_sw_row_count += 1
    return ws_sw_row_count

# Function to Create Static Paths within EPGs
def create_static_paths(wb, wb_sw, row_num, wr_method, dest_dir, dest_file, template, **templateVars):
    wsheets = wb_sw.get_sheet_names()
    tf_file = ''
    for wsheet in wsheets:
        ws = wb_sw[wsheet]
        for row in ws.rows:
            if not (row[12].value == None or row[13].value == None):
                vlan_test = ''
                if re.search('^(individual|port-channel|vpc)$', row[7].value) and (re.search(r'\d+', str(row[12].value)) or re.search(r'\d+', str(row[13].value))):
                    if not row[12].value == None:
                        vlan = row[12].value
                        vlan_test = vlan_range(vlan, **templateVars)
                        if 'true' in vlan_test:
                            templateVars['mode'] = 'native'
                    if not 'true' in vlan_test:
                        templateVars['mode'] = 'regular'
                        if not row[13].value == None:
                            vlans = row[13].value
                            vlan_test = vlan_range(vlans, **templateVars)
                if vlan_test == 'true':
                    templateVars['Pod_ID'] = row[1].value
                    templateVars['Node_ID'] = row[2].value
                    templateVars['Interface_Profile'] = row[3].value
                    templateVars['Interface_Selector'] = row[4].value
                    templateVars['Port'] = row[5].value
                    templateVars['Policy_Group'] = row[6].value
                    templateVars['Port_Type'] = row[7].value
                    templateVars['Bundle_ID'] = row[9].value
                    Site_Group = templateVars['Site_Group']
                    pod = templateVars['Pod_ID']
                    node_id =  templateVars['Node_ID']
                    if templateVars['Port_Type'] == 'vpc':
                        ws_vpc = wb['Inventory']
                        for rx in ws_vpc.rows:
                            if rx[0].value == 'vpc_pair' and int(rx[1].value) == int(Site_Group) and str(rx[4].value) == str(node_id):
                                node1 = templateVars['Node_ID']
                                node2 = rx[5].value
                                templateVars['Policy_Group'] = '%s_vpc%s' % (row[3].value, templateVars['Bundle_ID'])
                                templateVars['tDn'] = 'topology/pod-%s/protpaths-%s-%s/pathep-[%s]' % (pod, node1, node2, templateVars['Policy_Group'])
                                templateVars['Static_Path'] = 'rspathAtt-[topology/pod-%s/protpaths-%s-%s/pathep-[%s]' % (pod, node1, node2, templateVars['Policy_Group'])
                                templateVars['GUI_Static'] = 'Pod-%s/Node-%s-%s/%s' % (pod, node1, node2, templateVars['Policy_Group'])
                                templateVars['Static_descr'] = 'Pod-%s_Nodes-%s-%s_%s' % (pod, node1, node2, templateVars['Policy_Group'])
                                tf_file = './ACI/%s/%s/%s' % (templateVars['Site_Name'], dest_dir, dest_file)
                                read_file = open(tf_file, 'r')
                                read_file.seek(0)
                                static_path_descr = 'resource "aci_epg_to_static_path" "%s_%s_%s"' % (templateVars['App_Profile'], templateVars['EPG'], templateVars['Static_descr'])
                                if not static_path_descr in read_file.read():
                                    create_tf_file(wr_method, dest_dir, dest_file, template, **templateVars)

                    elif templateVars['Port_Type'] == 'port-channel':
                        templateVars['Policy_Group'] = '%s_pc%s' % (row[3].value, templateVars['Bundle_ID'])
                        templateVars['tDn'] = 'topology/pod-%s/paths-%s/pathep-[%s]' % (pod, templateVars['Node_ID'], templateVars['Policy_Group'])
                        templateVars['Static_Path'] = 'rspathAtt-[topology/pod-%s/paths-%s/pathep-[%s]' % (pod, templateVars['Node_ID'], templateVars['Policy_Group'])
                        templateVars['GUI_Static'] = 'Pod-%s/Node-%s/%s' % (pod, templateVars['Node_ID'], templateVars['Policy_Group'])
                        templateVars['Static_descr'] = 'Pod-%s_Node-%s_%s' % (pod, templateVars['Node_ID'], templateVars['Policy_Group'])
                        read_file = open(tf_file, 'r')
                        read_file.seek(0)
                        static_path_descr = 'resource "aci_epg_to_static_path" "%s_%s_%s"' % (templateVars['App_Profile'], templateVars['EPG'], templateVars['Static_descr'])
                        if not static_path_descr in read_file.read():
                            create_tf_file(wr_method, dest_dir, dest_file, template, **templateVars)

                    elif templateVars['Port_Type'] == 'individual':
                        port = 'eth%s' % (templateVars['Port'])
                        templateVars['tDn'] = 'topology/pod-%s/paths-%s/pathep-[%s]' % (pod, templateVars['Node_ID'], port)
                        templateVars['Static_Path'] = 'rspathAtt-[topology/pod-%s/paths-%s/pathep-[%s]' % (pod, templateVars['Node_ID'], port)
                        templateVars['GUI_Static'] = 'Pod-%s/Node-%s/%s' % (pod, templateVars['Node_ID'], port)
                        templateVars['Static_descr'] = 'Pod-%s_Node-%s_%s' % (pod, templateVars['Node_ID'], templateVars['Interface_Selector'])
                        read_file = open(tf_file, 'r')
                        read_file.seek(0)
                        static_path_descr = 'resource "aci_epg_to_static_path" "%s_%s_%s"' % (templateVars['App_Profile'], templateVars['EPG'], templateVars['Static_descr'])
                        if not static_path_descr in read_file.read():
                            create_tf_file(wr_method, dest_dir, dest_file, template, **templateVars)

# Function to Count the Number of Keys
def countKeys(ws, func):
    count = 0
    for i in ws.rows:
        if any(i):
            if str(i[0].value) == func:
                count += 1
    return count

# Function to find the Keys for each Section
def findKeys(ws, func_regex):
    func_list = OrderedSet()
    for i in ws.rows:
        if any(i):
            if re.search(func_regex, str(i[0].value)):
                func_list.add(str(i[0].value))
    return func_list

def findVars(ws, func, rows, count):
    var_list = []
    var_dict = {}
    for i in range(1, rows + 1):
        if (ws.cell(row=i, column=1)).value == func:
            print()
            try:
                for x in range(2, 34):
                    if (ws.cell(row=i - 1, column=x)).value:
                        var_list.append(str(ws.cell(row=i - 1, column=x).value))
                    else:
                        x += 1
            except Exception as e:
                e = e
                pass
            break
    vcount = 1
    while vcount <= count:
        var_dict[vcount] = {}
        var_count = 0
        for z in var_list:
            var_dict[vcount][z] = ws.cell(row=i + vcount - 1, column=2 + var_count).value
            var_count += 1
        var_dict[vcount]['row'] = i + vcount - 1
        vcount += 1
    return var_dict

# Function to execute HTTP Post
def post(uri, oauth_token, payload, section=''):
    # Use this for Troubleshooting
    if print_payload:
        print(payload)

    tf_token = 'Bearer %s' % (oauth_token)
    MyHeaders = {'Authorization': tf_token,
               'Content-Type': 'application/vnd.api+json'
    }

    r = ''
    while r == '':
        try:
            r = requests.post('https://app.terraform.io/api/v2/{}', data=json.dumps(payload), headers=MyHeaders)%(uri)
            status = r.status_code
        except requests.exceptions.ConnectionError as e:
            print("Connection error, pausing before retrying. Error: {}"
                  .format(e))
            time.sleep(5)
        except Exception as e:
            print("Method {} failed. Exception: {}".format(section[:-5], e))
            status = 666
            return(status)

    # Use this for Troubleshooting
    if print_response_always:
        print(r.text)
    if status != 200 and print_response_on_fail:
        print(r.text)

    return status

# Function to see if the State Location is an Environment Variable and if not Set it
def process_check_env_vars(**templateVars):
    sensitive_var = 'TF_VAR_%s' % (templateVars['sensitive_var'])

    # Check to see if the Variable is already set in the Environment, and if not set it.
    if os.environ.get(sensitive_var) is None:
        print(f'\n----------------------------------------------------------------------------------\n')
        print(f'  The State Location is set to "Local", which means that sensitive variables')
        print(f'  need to be stored locally in the Environment Variables.  The Script did not')
        print(f'  find {sensitive_var} as an "environment" variable.  To not be prompted for the ')
        print(f'  value of {sensitive_var} each time add the following to your local environemnt:\n')
        print(f'   - export {sensitive_var}="{sensitive_var}_value"')
        print(f'\n----------------------------------------------------------------------------------\n')
        while True:
            try:
                secure_value = getpass.getpass(prompt=f'Enter the value for the {sensitive_var}: ')
                break
            except Exception as e:
                print('Something went wrong. Error received: {}'.format(e))

        # Add the Variable to the Environment
        os.environ[sensitive_var] = '%s' % (secure_value)

# Function to Determine if Data Should be Stored in TC4B or Local
def process_check_tf_cloud_vars(tfcloud_token, sensitive_var, **templateVars):

    uri = 'lmnop'
    # Check to see if the Variable is already set in Terraform Cloud, and if not set it.
    if uri is not 'xyz':
        print(f'\n-----------------------------------------------------------------------------\n')
        print(f'  The Secure Variable Location is set to "tf_cloud" for {sensitive_var}.')
        print(f'  The Script was not able to find this Variable in the the Workspace.')
        print(f'\n-----------------------------------------------------------------------------\n')
        while True:
            try:
                secure_value = getpass.getpass(prompt=f'Enter the value for the {sensitive_var}: ')
                break
            except Exception as e:
                print('Something went wrong. Error received: {}'.format(e))

    org = templateVars['TF_Cloud_Org']
    url = 'https://app.terraform.io/api/v2/organizations/%s/workspaces' %  (org)
    tf_token = 'Bearer %s' % (tfcloud_token)
    MyHeaders = {'Authorization': tf_token,
               'Content-Type': 'application/vnd.api+json'
    }

    get_ws = ''
    while get_ws == '':
        try:
            get_ws = requests.get(url, headers=MyHeaders)
            status = get_ws.status_code
        except requests.exceptions.ConnectionError as e:
            print("Connection error, pausing before retrying. Error: {}"
                  .format(e))
            time.sleep(5)
        except Exception as e:
            print("failed. Exception: {}".format(e))
            status = 666
            return(status)

    # Use this for Troubleshooting
    # if print_response_always:
    #     print(get_ws.text)
    # if status != 200 and print_response_on_fail:
    #     print(get_ws.text)

    workspace = 'tyscott-aci-admin'
    json_data = get_ws.json()
    key_count = 0
    for keys in json_data['data']:
        if keys['attributes']['name'] == workspace:
            print(keys['id'])
            workspace_id = keys['id']
            key_count =+ 1

    if not key_count == 1:

        templateLoader = jinja2.FileSystemLoader(searchpath=(aci_template_path + 'tf_cloud/'))
        templateEnv = jinja2.Environment(loader=templateLoader)

        template_file = "sensitive_variable.json"
        template = templateEnv.get_template(template_file)
        payload = template.render(templateVars)

        post(uri, tfcloud_token, payload, template_file)
    exit()

# Function to Process Sensitive Variables
def process_sensitive_var(wb, ws, row_num, dest_dir, dest_file, template, **templateVars):
    # Check the Sites Tab for Variable Location
    if re.search('Grp_[A-F]', templateVars['Site_Group']):
        Group_ID = '%s' % (templateVars['Site_Group'])
        site_group = ast.literal_eval(os.environ[Group_ID])
        for x in range(1, 13):
            sitex = 'Site_%s' % (x)
            if not site_group[sitex] == None:
                Site_ID = 'Site_ID_%s' % (site_group[sitex])
                site_dict = ast.literal_eval(os.environ[Site_ID])

                # Create templateVars for Environment Information
                templateVars['Site_Name'] = site_dict.get('Site_Name')
                templateVars['APIC_URL'] = site_dict.get('APIC_URL')
                templateVars['APIC_Version'] = site_dict.get('APIC_Version')
                templateVars['APIC_Auth_Type'] = site_dict.get('APIC_Auth_Type')
                templateVars['Terraform_Version'] = site_dict.get('Terraform_Version')
                templateVars['Provider_Version'] = site_dict.get('Provider_Version')
                templateVars['Run_Location'] = site_dict.get('Run_Location')
                templateVars['State_Location'] = site_dict.get('State_Location')
                templateVars['Terraform_Cloud_Org'] = site_dict.get('Terraform_Cloud_Org')
                templateVars['Workspace_Prefix'] = site_dict.get('Workspace_Prefix')
                templateVars['VCS_Base_Repo'] = site_dict.get('VCS_Base_Repo')
                templateVars['Terraform_Cloud_Agent'] = site_dict.get('Terraform_Cloud_Agent')

                if not templateVars['Workspace_Prefix'] == None:
                    templateVars['workspace'] = '%s_ACI_%s_%s' % (templateVars['Workspace_Prefix'], templateVars['Site_Name'], dest_dir)
                else:
                    templateVars['workspace'] = 'ACI_%s_%s' % (templateVars['Site_Name'], dest_dir)


                if templateVars['Run_Location'] == 'Local':
                    process_check_env_vars(**templateVars)
                else:
                    uri = 'workspaces/%s/vars' % (templateVars['workspace'])
                    process_check_tf_cloud_vars(**templateVars)

    elif re.search(r'\d+', templateVars['Site_Group']):
        Site_ID = 'Site_ID_%s' % (templateVars['Site_Group'])
        site_dict = ast.literal_eval(os.environ[Site_ID])

        # Create templateVars for Environment Information
        templateVars['Site_Name'] = site_dict.get('Site_Name')
        templateVars['APIC_URL'] = site_dict.get('APIC_URL')
        templateVars['APIC_Version'] = site_dict.get('APIC_Version')
        templateVars['APIC_Auth_Type'] = site_dict.get('APIC_Auth_Type')
        templateVars['Terraform_Version'] = site_dict.get('Terraform_Version')
        templateVars['Provider_Version'] = site_dict.get('Provider_Version')
        templateVars['Run_Location'] = site_dict.get('Run_Location')
        templateVars['State_Location'] = site_dict.get('State_Location')
        templateVars['Terraform_Cloud_Org'] = site_dict.get('Terraform_Cloud_Org')
        templateVars['Workspace_Prefix'] = site_dict.get('Workspace_Prefix')
        templateVars['VCS_Base_Repo'] = site_dict.get('VCS_Base_Repo')
        templateVars['Terraform_Cloud_Agent'] = site_dict.get('Terraform_Cloud_Agent')

        if templateVars['Run_Location'] == 'Local':
            process_check_env_vars(**templateVars)
        else:
            process_check_tf_cloud_vars(**templateVars)

    else:
        print(f"\n-----------------------------------------------------------------------------\n")
        print(f"   Error on Worksheet {ws.title}, Row {row_num} Site_Group, value {templateVars['Site_Group']}.")
        print(f"   Unable to Determine if this is a Single or Group of Site(s).  Exiting....")
        print(f"\n-----------------------------------------------------------------------------\n")
        exit()

# Function to validate input for each method
def process_kwargs(required_args, optional_args, **kwargs):
    # Validate all required kwargs passed
    if all(item in kwargs for item in required_args.keys()) is not True:
        error_ = '\n***ERROR***\nREQUIRED ARGS ARE:\n "%s"\nOPTIONAL ARGS ARE:\n "%s"\nPROVIDED ARGS ARE:\n"%s"\nInsufficient required arguments.' % (required_args, optional_args, kwargs)
        raise InsufficientArgs(error_)

    # Load all required args values from kwargs
    for item in kwargs:
        if item in required_args.keys():
            required_args[item] = kwargs[item]
    for item in kwargs:
        if item in optional_args.keys():
            optional_args[item] = kwargs[item]

    # Combine option and required dicts for Jinja template render
    templateVars = {**required_args, **optional_args}
    return(templateVars)

# Function to Apply Method to Site(s)
def process_method(wb, ws, row_num, wr_method, dest_dir, dest_file, template, **templateVars):
    if re.search('Grp_[A-F]', templateVars['Site_Group']):
        Group_ID = '%s' % (templateVars['Site_Group'])
        site_group = ast.literal_eval(os.environ[Group_ID])
        for x in range(1, 13):
            sitex = 'Site_%s' % (x)
            if not site_group[sitex] == None:
                Site_ID = 'Site_ID_%s' % (site_group[sitex])
                site_dict = ast.literal_eval(os.environ[Site_ID])

                # Create templateVars for Site_Name and APIC_URL
                templateVars['Site_Name'] = site_dict.get('Site_Name')
                templateVars['APIC_URL'] = site_dict.get('APIC_URL')

                # Create Terraform file from Template
                create_tf_file(wr_method, dest_dir, dest_file, template, **templateVars)

    elif re.search(r'\d+', templateVars['Site_Group']):
        Site_ID = 'Site_ID_%s' % (templateVars['Site_Group'])
        site_dict = ast.literal_eval(os.environ[Site_ID])

        # Create templateVars for Site_Name and APIC_URL
        templateVars['Site_Name'] = site_dict.get('Site_Name')
        templateVars['APIC_URL'] = site_dict.get('APIC_URL')

        # Create Terraform file from Template
        create_tf_file(wr_method, dest_dir, dest_file, template, **templateVars)
    else:
        print(f"\n-----------------------------------------------------------------------------\n")
        print(f"   Error on Worksheet {ws.title}, Row {row_num} Site_Group, value {templateVars['Site_Group']}.")
        print(f"   Unable to Determine if this is a Single or Group of Site(s).  Exiting....")
        print(f"\n-----------------------------------------------------------------------------\n")
        exit()

# Function to Add Static Port Bindings to Bridge Domains Terraform Files
def process_workbook(wb, ws, row_num, wr_method, dest_dir, dest_file, template, **templateVars):
    if re.search('Grp_[A-F]', templateVars['Site_Group']):
        Group_ID = '%s' % (templateVars['Site_Group'])
        site_group = ast.literal_eval(os.environ[Group_ID])
        for x in range(1, 13):
            sitex = 'Site_%s' % (x)
            if not site_group[sitex] == None:
                Site_ID = 'Site_ID_%s' % (site_group[sitex])
                site_dict = ast.literal_eval(os.environ[Site_ID])

                # Create templateVars for Site_Name and APIC_URL
                templateVars['Site_Name'] = site_dict.get('Site_Name')
                templateVars['Site_Group'] = site_dict.get('Site_ID')
                templateVars['APIC_URL'] = site_dict.get('APIC_URL')

                # Pull in the Site Workbook
                excel_workbook = '%s_intf_selectors.xlsx' % (templateVars['Site_Name'])
                try:
                    wb_sw = load_workbook(excel_workbook)
                except Exception as e:
                    print(f"Something went wrong while opening the workbook - {excel_workbook}... ABORT!")
                    sys.exit(e)

                # Process the Interface Selectors for Static Port Paths
                create_static_paths(wb, wb_sw, row_num, wr_method, dest_dir, dest_file, template, **templateVars)

    elif re.search(r'\d+', templateVars['Site_Group']):
        Site_ID = 'Site_ID_%s' % (templateVars['Site_Group'])
        site_dict = ast.literal_eval(os.environ[Site_ID])

        # Create templateVars for Site_Name and APIC_URL
        templateVars['Site_Name'] = site_dict.get('Site_Name')
        templateVars['Site_Group'] = site_dict.get('Site_ID')
        templateVars['APIC_URL'] = site_dict.get('APIC_URL')

        # Pull in the Site Workbook
        excel_workbook = '%s_intf_selectors.xlsx' % (templateVars['Site_Name'])
        try:
            wb_sw = load_workbook(excel_workbook)
        except Exception as e:
            print(f"Something went wrong while opening the workbook - {excel_workbook}... ABORT!")
            sys.exit(e)


        # Process the Interface Selectors for Static Port Paths
        create_static_paths(wb, wb_sw, row_num, wr_method, dest_dir, dest_file, template, **templateVars)

    else:
        print(f"\n-----------------------------------------------------------------------------\n")
        print(f"   Error on Worksheet {ws.title}, Row {row_num} Site_Group, value {templateVars['Site_Group']}.")
        print(f"   Unable to Determine if this is a Single or Group of Site(s).  Exiting....")
        print(f"\n-----------------------------------------------------------------------------\n")
        exit()

# Function to Determine Port Count on Modules
def query_module_type(row_num, module_type):
    if re.search('^M4', module_type):
        port_count = '4'
    elif re.search('^M6', module_type):
        port_count = '6'
    elif re.search('^M12', module_type):
        port_count = '12'
    elif re.search('X9716D-GX', module_type):
        port_count = '16'
    elif re.search('X9732C-EX', module_type):
        port_count = '32'
    elif re.search('X9736', module_type):
        port_count = '36'
    else:
        print(f'\n-----------------------------------------------------------------------------\n')
        print(f'   Error on Row {row_num}.  Unknown Switch Model {module_type}')
        print(f'   Please verify Input Information.  Exiting....')
        print(f'\n-----------------------------------------------------------------------------\n')
        exit()
    return port_count

# Function to Determine Port count from Switch Model
def query_switch_model(row_num, switch_type):
    modules = ''
    switch_type = str(switch_type)
    if re.search('^9396', switch_type):
        modules = '2'
        port_count = '48'
    elif re.search('^93', switch_type):
        modules = '1'

    if re.search('^9316', switch_type):
        port_count = '16'
    elif re.search('^(93120)', switch_type):
        port_count = '102'
    elif re.search('^(93108|93120|93216|93360)', switch_type):
        port_count = '108'
    elif re.search('^(93180|93240|9348|9396)', switch_type):
        port_count = '54'
    elif re.search('^(93240)', switch_type):
        port_count = '60'
    elif re.search('^9332', switch_type):
        port_count = '34'
    elif re.search('^(9336|93600)', switch_type):
        port_count = '36'
    elif re.search('^9364C-GX', switch_type):
        port_count = '64'
    elif re.search('^9364', switch_type):
        port_count = '66'
    elif re.search('^95', switch_type):
        port_count = '36'
        if switch_type == '9504':
            modules = '4'
        elif switch_type == '9508':
            modules = '8'
        elif switch_type == '9516':
            modules = '16'
        else:
            print(f'\n-----------------------------------------------------------------------------\n')
            print(f'   Error on Row {row_num}.  Unknown Switch Model {switch_type}')
            print(f'   Please verify Input Information.  Exiting....')
            print(f'\n-----------------------------------------------------------------------------\n')
            exit()
    else:
        print(f'\n-----------------------------------------------------------------------------\n')
        print(f'   Error on Row {row_num}.  Unknown Switch Model {switch_type}')
        print(f'   Please verify Input Information.  Exiting....')
        print(f'\n-----------------------------------------------------------------------------\n')
        exit()
    return modules,port_count

# Function to Read Excel Workbook Data
def read_in(excel_workbook):
    try:
        wb = load_workbook(excel_workbook)
        print("Workbook Loaded.")
    except Exception as e:
        print(f"Something went wrong while opening the workbook - {excel_workbook}... ABORT!")
        sys.exit(e)
    return wb

# Function to Define stdout_log output
def stdout_log(sheet, line):
    if log_level == 0:
        return
    elif ((log_level == (1) or log_level == (2)) and
            (sheet) and (line is None)):
        #print('*' * 80)
        print(f'\n-----------------------------------------------------------------------------\n')
        print(f'   Starting work on {sheet.title} Worksheet')
        print(f'\n-----------------------------------------------------------------------------\n')
        #print('*' * 80)
    elif log_level == (2) and (sheet) and (line is not None):
        print('Evaluating line %s from %s Worksheet...' % (line, sheet.title))
    else:
        return

# Function to Expand a VLAN List to Individual VLANs
def vlan_list_full(vlan_list):
    full_vlan_list = []
    if re.search(r',', str(vlan_list)):
        vlist = vlan_list.split(',')
        for v in vlist:
            if re.fullmatch('^\\d{1,4}\\-\\d{1,4}$', v):
                a,b = v.split('-')
                a = int(a)
                b = int(b)
                vrange = range(a,b+1)
                for vl in vrange:
                    full_vlan_list.append(vl)
            elif re.fullmatch('^\\d{1,4}$', v):
                full_vlan_list.append(v)
    elif re.search('\\-', str(vlan_list)):
        a,b = vlan_list.split('-')
        a = int(a)
        b = int(b)
        vrange = range(a,b+1)
        for v in vrange:
            full_vlan_list.append(v)
    else:
        full_vlan_list.append(vlan_list)
    return full_vlan_list

# Add Prefix to VLAN Numbers for BD/EPG Names
def vlan_to_netcentric(vlan):
    vlan = int(vlan)
    if vlan < 10:
        vlan = str(vlan)
        netcentric = 'v000' + vlan
        return netcentric
    elif vlan < 100:
        vlan = str(vlan)
        netcentric = 'v00' + vlan
        return netcentric
    elif vlan < 1000:
        vlan = str(vlan)
        netcentric = 'v0' + vlan
        return netcentric
    else:
        vlan = str(vlan)
        netcentric = 'v' + vlan
        return netcentric

# Function to Expand a VLAN Range to a VLAN List
def vlan_range(vlan_list, **templateVars):
    results = 'unknown'
    while results == 'unknown':
        if re.search(',', str(vlan_list)):
            vx = vlan_list.split(',')
            for vrange in vx:
                if re.search('-', vrange):
                    vl = vrange.split('-')
                    min_ = int(vl[0])
                    max_ = int(vl[1])
                    if (int(templateVars['VLAN']) >= min_ and int(templateVars['VLAN']) <= max_):
                        results = 'true'
                        return results
                else:
                    if templateVars['VLAN'] == vrange:
                        results = 'true'
                        return results
            results = 'false'
            return results
        elif re.search('-', str(vlan_list)):
            vl = vlan_list.split('-')
            min_ = int(vl[0])
            max_ = int(vl[1])
            if (int(templateVars['VLAN']) >= min_ and int(templateVars['VLAN']) <= max_):
                results = 'true'
                return results
        else:
            if int(templateVars['VLAN']) == int(vlan_list):
                results = 'true'
                return results
        results = 'false'
        return results
